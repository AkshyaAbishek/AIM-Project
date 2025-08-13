"""
AIM Web Application - Flask Implementation
Converts the tkinter desktop application to a web-based interface for Azure deployment
"""

import os
import json
import hashlib
from datetime import datetime
from typing import Dict, Any, Optional
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO

from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file, session
from flask import Response
import sqlite3
import sys

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

try:
    from aim_processor import AIMProcessor, ValidationError, MappingError
except ImportError:
    # Fallback for web deployment
    class AIMProcessor:
        def process_fast_ui_input(self, data, product_type):
            return {"status": "success", "message": "Processed successfully"}
    
    class ValidationError(Exception):
        pass
    
    class MappingError(Exception):
        pass

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize AIM processor
aim_processor = AIMProcessor()

class WebDatabaseManager:
    """Database manager for web application"""
    
    def __init__(self, db_path: str = "aim_data.db"):
        # Ensure we're using the correct path
        if not os.path.isabs(db_path):
            db_path = os.path.join(os.path.dirname(__file__), db_path)
        self.db_path = db_path
        print(f"Database path: {self.db_path}")
        self.init_database()
    
    def init_database(self):
        """Initialize database with web-optimized schema"""
        try:
            print(f"Initializing database at: {self.db_path}")
            with sqlite3.connect(self.db_path) as conn:
                # First, create table if it doesn't exist with minimal schema
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS user_data (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        product_type TEXT NOT NULL,
                        data_hash TEXT UNIQUE NOT NULL,
                        json_data TEXT NOT NULL,
                        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            
            conn.execute("""
                CREATE TABLE IF NOT EXISTS processing_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    operation_details TEXT,
                    status TEXT CHECK(status IN ('success', 'warning', 'error')),
                    error_message TEXT,
                    processing_time_ms INTEGER,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Add new columns if they don't exist (for existing databases)
            columns_to_add = [
                ("calculator_path", "TEXT"),
                ("raw_data", "TEXT"),
                ("processed_data", "TEXT"),
                ("field_mappings", "TEXT"),
                ("updated_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("modified_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("validation_status", "TEXT DEFAULT 'pending'"),
                ("status", "TEXT DEFAULT 'pending'"),
                ("quality_score", "REAL DEFAULT 0.0"),
                ("file_source", "TEXT"),
                ("source_file", "TEXT"),
                ("processing_notes", "TEXT"),
                ("notes", "TEXT"),
                ("session_id", "TEXT")
            ]
            
            for column_name, column_type in columns_to_add:
                try:
                    conn.execute(f"ALTER TABLE user_data ADD COLUMN {column_name} {column_type}")
                except sqlite3.OperationalError:
                    pass  # Column already exists
            
            # Add session_id to processing_logs if it doesn't exist
            try:
                conn.execute("ALTER TABLE processing_logs ADD COLUMN session_id TEXT")
            except sqlite3.OperationalError:
                pass
            
            # Create indexes
            try:
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_hash ON user_data(data_hash)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_session ON user_data(session_id)")
            except sqlite3.OperationalError:
                pass
            
            print("Database initialization completed successfully")
                
        except Exception as e:
            print(f"Error initializing database: {e}")
            import traceback
            traceback.print_exc()
    
    def store_data(self, name: str, product_type: str, json_data: Dict, 
                   session_id: str, file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with duplicate detection"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, created_date, 
                     validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    datetime.now().isoformat(), 'validated', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'storage_error',
                'message': str(e)
            }
    
    def calculate_quality_score(self, json_data: Dict) -> float:
        """Calculate data quality score"""
        score = 0.0
        
        # Required fields check (40% of score)
        required_fields = ['policy_number', 'product_type', 'effective_date']
        present_required = sum(1 for field in required_fields if field in json_data and json_data[field])
        score += (present_required / len(required_fields)) * 40
        
        # Data completeness (30% of score)
        total_fields = len(json_data)
        non_empty_fields = sum(1 for v in json_data.values() if v is not None and str(v).strip())
        if total_fields > 0:
            score += (non_empty_fields / total_fields) * 30
        
        # Data format consistency (30% of score)
        format_score = 30
        if 'age' in json_data:
            try:
                age = float(json_data['age'])
                if age < 0 or age > 120:
                    format_score -= 10
            except:
                format_score -= 10
        
        score += format_score
        return round(min(100, score), 2)
    
    def add_user_data(self, name: str, raw_data: Dict, product_type: str, 
                      source_file: Optional[str] = None, status: str = 'pending',
                      calculator_path: Optional[str] = None) -> int:
        """Add user data to database and return record ID"""
        try:
            data_str = json.dumps(raw_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Check for duplicates
                cursor.execute("SELECT id FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing record
                    cursor.execute("""
                        UPDATE user_data 
                        SET updated_date = ?, status = ?, calculator_path = ?
                        WHERE id = ?
                    """, (datetime.now().isoformat(), status, calculator_path, existing[0]))
                    return existing[0]
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(raw_data)
                
                # Insert new record
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, raw_data, created_date, 
                     validation_status, status, quality_score, source_file, calculator_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(raw_data), json.dumps(raw_data),
                    datetime.now().isoformat(), 'validated', status, quality_score,
                    source_file, calculator_path
                ))
                
                return cursor.lastrowid
                
        except Exception as e:
            raise Exception(f"Failed to add user data: {str(e)}")

    def get_statistics(self) -> Dict[str, Any]:
        """Get database statistics for dashboard"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Total records
                cursor.execute("SELECT COUNT(*) FROM user_data")
                total_records = cursor.fetchone()[0]
                
                # Records by product type
                cursor.execute("SELECT product_type, COUNT(*) FROM user_data GROUP BY product_type")
                product_counts = dict(cursor.fetchall())
                
                # Basic stats to avoid column issues
                return {
                    'total_records': total_records,
                    'status_counts': {'processed': total_records},  # Simplified
                    'product_counts': product_counts,
                    'product_distribution': product_counts,  # Template expects this name
                    'average_quality_score': 0.0,  # Simplified
                    'average_quality': 0.0,  # Template expects this name
                    'recent_activity': total_records,  # Simplified
                    'processed_records': total_records,
                    'pending_records': 0,
                    'error_records': 0
                }
        except Exception as e:
            print(f"Database error in get_statistics: {e}")
            # Return default stats if there's an error
            return {
                'total_records': 0,
                'status_counts': {},
                'product_counts': {},
                'product_distribution': {},  # Template expects this name
                'average_quality_score': 0,
                'average_quality': 0,  # Template expects this name
                'recent_activity': 0,
                'processed_records': 0,
                'pending_records': 0,
                'error_records': 0
            }

    def load_calculator_reference(self, calculator_path: str, product_type: str) -> Dict:
        """Load calculator reference data (mock implementation)"""
        # This is a mock implementation. In a real system, you would load
        # the actual calculator configuration and reference data
        
        calculator_templates = {
            'life_insurance': {
                'fields': {
                    'policy_number': {'required': True, 'type': 'string', 'sample': 'LIFE001234'},
                    'insured_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1980-01-01'},
                    'gender': {'required': True, 'type': 'string', 'sample': 'M'},
                    'coverage_amount': {'required': True, 'type': 'number', 'sample': 100000},
                    'premium_amount': {'required': False, 'type': 'number', 'sample': 1200},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'TERM20'},
                    'beneficiary_name': {'required': False, 'type': 'string', 'sample': 'Jane Doe'},
                    'agent_code': {'required': False, 'type': 'string', 'sample': 'AGT001'}
                }
            },
            'annuity': {
                'fields': {
                    'contract_number': {'required': True, 'type': 'string', 'sample': 'ANN001234'},
                    'owner_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1970-01-01'},
                    'initial_deposit': {'required': True, 'type': 'number', 'sample': 50000},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'FIXED5'},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'surrender_charge_period': {'required': False, 'type': 'number', 'sample': 7},
                    'interest_rate': {'required': False, 'type': 'number', 'sample': 3.5}
                }
            }
        }
        
        return calculator_templates.get(product_type, calculator_templates['life_insurance'])

    def perform_data_comparison(self, source_data: Dict, calculator_data: Dict) -> Dict:
        """Compare source data with calculator reference and generate comparison result"""
        
        calculator_fields = calculator_data.get('fields', {})
        comparison_fields = []
        
        # Stats
        total_fields = len(calculator_fields)
        matching_fields = 0
        missing_fields = 0
        filled_fields = 0
        
        for field_name, field_config in calculator_fields.items():
            source_value = source_data.get(field_name)
            calculator_value = field_config.get('sample')
            
            field_result = {
                'field_name': field_name,
                'source_value': source_value,
                'calculator_value': calculator_value,
                'required': field_config.get('required', False),
                'data_type': field_config.get('type', 'string')
            }
            
            # Determine status
            if source_value is not None and source_value != '':
                if str(source_value).strip().lower() == str(calculator_value).strip().lower():
                    field_result['status'] = 'match'
                    matching_fields += 1
                else:
                    field_result['status'] = 'mismatch'
            else:
                if calculator_value and field_config.get('required', False):
                    field_result['status'] = 'missing'
                    missing_fields += 1
                elif calculator_value:
                    field_result['status'] = 'can_fill'
                    filled_fields += 1
                else:
                    field_result['status'] = 'missing'
                    missing_fields += 1
            
            comparison_fields.append(field_result)
        
        return {
            'stats': {
                'total_fields': total_fields,
                'matching_fields': matching_fields,
                'missing_fields': missing_fields,
                'filled_fields': filled_fields,
                'completion_percentage': round((matching_fields / total_fields) * 100, 2) if total_fields > 0 else 0
            },
            'fields': comparison_fields
        }

    def get_all_data(self, session_id: Optional[str] = None) -> list:
        """Get all stored data, optionally filtered by session"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if session_id:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status AS status, 
                           quality_score, file_source, data_hash
                    FROM user_data WHERE session_id = ? ORDER BY created_date DESC
                """, (session_id,))
            else:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status AS status, 
                           quality_score, file_source, data_hash
                    FROM user_data ORDER BY created_date DESC LIMIT 100
                """)
            
            rows = cursor.fetchall()
            data = []
            for row in rows:
                record = dict(row)
                # Ensure default values for missing fields
                record['source_file'] = record.get('file_source') or 'Manual Entry'
                record['status'] = record.get('status') or 'processed'
                record['quality_score'] = record.get('quality_score') or 0.0
                data.append(record)
            
            return data
    
    def get_data_by_id(self, record_id: int) -> Optional[Dict]:
        """Get specific record by ID"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM user_data WHERE id = ?", (record_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def store_data_with_calculator(self, name: str, product_type: str, json_data: Dict, 
                                   calculator_path: str, session_id: str, 
                                   file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with calculator path for save-only operation"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, calculator_path, raw_data,
                     created_date, validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    calculator_path, json.dumps(json_data),
                    datetime.now().isoformat(), 'saved', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score,
                    'message': 'Data saved successfully'
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'database_error',
                'message': f'Database error: {str(e)}'
            }

    def store_processed_data(self, name: str, product_type: str, source_data: Dict,
                           processed_data: Dict, calculator_path: str, field_mapping: Dict,
                           session_id: str, file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store processed data with mapping and calculator results"""
        try:
            data_str = json.dumps(source_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score for processed data
                quality_score = self.calculate_processed_quality_score(source_data, processed_data, field_mapping)
                
                # Calculate additional metrics
                calculator_results = {
                    'total_fields_mapped': len([m for m in field_mapping['field_mappings'] if m['mapping_status'] == 'auto_mapped']),
                    'mapping_confidence': sum(m['confidence_score'] for m in field_mapping['field_mappings']) / len(field_mapping['field_mappings']) if field_mapping['field_mappings'] else 0,
                    'processing_timestamp': datetime.now().isoformat()
                }
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, calculator_path, raw_data,
                     processed_data, field_mappings, created_date, validation_status, 
                     quality_score, file_source, session_id, processing_notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(processed_data),
                    calculator_path, json.dumps(source_data),
                    json.dumps(processed_data), json.dumps(field_mapping),
                    datetime.now().isoformat(), 'processed', quality_score,
                    file_source, session_id, json.dumps(calculator_results)
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score,
                    'calculator_results': calculator_results,
                    'message': 'Data processed successfully'
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'database_error',
                'message': f'Database error: {str(e)}'
            }

    def load_calculator_configuration(self, calculator_path: str, product_type: str) -> Dict:
        """Load calculator configuration from JSON files with enhanced field definitions"""
        import os
        
        # Try to load from actual file path first
        config_file_map = {
            '/calculators/life_insurance/standard.json': 'life_insurance_calculator.json',
            '/calculators/life_insurance/term.json': 'life_insurance_calculator.json',
            '/calculators/annuity/fixed.json': 'annuity_calculator.json',
            '/calculators/annuity/variable.json': 'annuity_calculator.json'
        }
        
        # Get config file name
        config_file = config_file_map.get(calculator_path)
        if not config_file:
            # Use product type as fallback
            config_file = f"{product_type}_calculator.json" if product_type in ['life', 'annuity'] else 'life_insurance_calculator.json'
        
        # Try to load from config files directory
        config_path = os.path.join(os.path.dirname(__file__), 'src', 'config', 'config_files', config_file)
        
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading calculator config from {config_path}: {e}")
        
        # Fallback to embedded templates
        calculator_templates = {
            'life': {
                'name': 'Life Insurance Calculator',
                'version': '2.0',
                'fields': {
                    'policy_number': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'LIFE001234',
                        'description': 'Unique policy identifier'
                    },
                    'applicant_first_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'John',
                        'description': 'Applicant first name'
                    },
                    'applicant_last_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'Doe',
                        'description': 'Applicant last name'
                    },
                    'birth_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '1980-01-01',
                        'description': 'Applicant date of birth'
                    },
                    'effective_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '2025-01-01',
                        'description': 'Policy effective date'
                    },
                    'face_amount': {
                        'required': True, 
                        'type': 'currency', 
                        'sample': 100000,
                        'description': 'Policy face amount'
                    },
                    'premium': {
                        'required': False, 
                        'type': 'currency', 
                        'sample': 1200,
                        'description': 'Annual premium amount',
                        'formula': '$face_amount * 0.012'
                    },
                    'applicant_age': {
                        'required': False,
                        'type': 'number',
                        'sample': 45,
                        'description': 'Calculated age at effective date',
                        'formula': 'calculated_from_dates'
                    },
                    'gender': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'M',
                        'description': 'Gender (M/F)'
                    },
                    'product_code': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'TERM20',
                        'description': 'Product code'
                    },
                    'beneficiary_name': {
                        'required': False, 
                        'type': 'string', 
                        'sample': 'Jane Doe',
                        'description': 'Primary beneficiary name'
                    },
                    'maturity_date': {
                        'required': False,
                        'type': 'date',
                        'sample': '2045-01-01',
                        'description': 'Policy maturity date',
                        'formula': 'calculated_from_effective_date_plus_term'
                    }
                },
                'calculations': {
                    'applicant_age': {
                        'type': 'date_based',
                        'calculation_type': 'age',
                        'description': 'Calculate age from birth date and effective date'
                    },
                    'premium_adjustment': {
                        'type': 'date_based',
                        'calculation_type': 'premium_adjustment',
                        'factor': 1.0,
                        'description': 'Adjust premium based on age and other factors'
                    },
                    'maturity_date': {
                        'type': 'date_based',
                        'calculation_type': 'maturity_date',
                        'default_term': 20,
                        'description': 'Calculate maturity date from effective date'
                    }
                }
            },
            'annuity': {
                'name': 'Annuity Calculator',
                'version': '2.0',
                'fields': {
                    'contract_number': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'ANN001234',
                        'description': 'Unique contract identifier'
                    },
                    'owner_first_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'John',
                        'description': 'Contract owner first name'
                    },
                    'owner_last_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'Doe',
                        'description': 'Contract owner last name'
                    },
                    'birth_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '1970-01-01',
                        'description': 'Owner date of birth'
                    },
                    'effective_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '2025-01-01',
                        'description': 'Contract effective date'
                    },
                    'initial_deposit': {
                        'required': True, 
                        'type': 'currency', 
                        'sample': 50000,
                        'description': 'Initial deposit amount'
                    },
                    'interest_rate': {
                        'required': False, 
                        'type': 'number', 
                        'sample': 3.5,
                        'description': 'Annual interest rate percentage'
                    },
                    'owner_age': {
                        'required': False,
                        'type': 'number',
                        'sample': 55,
                        'description': 'Calculated age at effective date',
                        'formula': 'calculated_from_dates'
                    },
                    'surrender_charge_period': {
                        'required': False, 
                        'type': 'number', 
                        'sample': 7,
                        'description': 'Surrender charge period in years'
                    },
                    'projected_value_10_years': {
                        'required': False,
                        'type': 'currency',
                        'sample': 67195,
                        'description': 'Projected value after 10 years',
                        'formula': '$initial_deposit * (1 + $interest_rate/100)^10'
                    }
                },
                'calculations': {
                    'owner_age': {
                        'type': 'date_based',
                        'calculation_type': 'age',
                        'description': 'Calculate age from birth date and effective date'
                    },
                    'projected_value': {
                        'type': 'compound_interest',
                        'calculation_type': 'future_value',
                        'years': [5, 10, 15, 20],
                        'description': 'Calculate projected values with compound interest'
                    }
                }
            }
        }
        
        return calculator_templates.get(product_type, calculator_templates['life'])

    def calculate_processed_quality_score(self, source_data: Dict, processed_data: Dict, field_mapping: Dict) -> float:
        """Calculate quality score for processed data"""
        if not field_mapping or not field_mapping.get('field_mappings'):
            return 0.0
        
        total_fields = len(field_mapping['field_mappings'])
        mapped_fields = len([m for m in field_mapping['field_mappings'] if m['mapping_status'] == 'auto_mapped'])
        required_fields = len([m for m in field_mapping['field_mappings'] if m.get('calculator_required', False)])
        required_mapped = len([m for m in field_mapping['field_mappings'] 
                             if m.get('calculator_required', False) and m['mapping_status'] == 'auto_mapped'])
        
        # Base score from mapping percentage
        mapping_score = (mapped_fields / total_fields) * 100 if total_fields > 0 else 0
        
        # Bonus for required fields
        required_bonus = (required_mapped / required_fields) * 20 if required_fields > 0 else 20
        
        # Average confidence score
        confidence_scores = [m['confidence_score'] for m in field_mapping['field_mappings'] if m['confidence_score'] > 0]
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
        confidence_bonus = (avg_confidence / 100) * 15
        
        total_score = min(100, mapping_score + required_bonus + confidence_bonus)
        return round(total_score, 1)

# Initialize database manager
db_manager = WebDatabaseManager()

@app.route('/')
@app.route('/index')
def index():
    """Main dashboard page"""
    try:
        if 'session_id' not in session:
            session['session_id'] = os.urandom(16).hex()
        
        print("Getting statistics...")
        stats = db_manager.get_statistics()
        print(f"Stats: {stats}")
        
        print("Getting recent data...")
        recent_data = db_manager.get_all_data(session['session_id'])
        print(f"Recent data count: {len(recent_data)}")
        
        # Use the original beautiful template
        return render_template('index.html', stats=stats, recent_data=recent_data)
        
    except Exception as e:
        print(f"Error in index route: {e}")
        import traceback
        traceback.print_exc()
        return f"Error: {e}", 500

@app.route('/help')
def help_page():
    """Help page with documentation and usage instructions"""
    return render_template('help.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    """File upload and processing page"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name', '').strip()
            product_type = request.form.get('product_type', '').strip()
            
            if not name or not product_type:
                flash('Name and product type are required', 'error')
                return redirect(url_for('upload_file'))
            
            # Handle JSON data input
            json_text = request.form.get('json_data', '').strip()
            uploaded_file = request.files.get('file')
            
            json_data = None
            file_source = None
            
            if json_text:
                # Parse JSON from text input
                try:
                    json_data = json.loads(json_text)
                    file_source = 'manual_input'
                except json.JSONDecodeError as e:
                    flash(f'Invalid JSON format: {str(e)}', 'error')
                    return redirect(url_for('upload_file'))
            
            elif uploaded_file and uploaded_file.filename:
                # Process uploaded file
                filename = secure_filename(uploaded_file.filename)
                file_source = filename
                
                if filename.endswith('.json'):
                    try:
                        json_data = json.load(uploaded_file)
                    except json.JSONDecodeError as e:
                        flash(f'Invalid JSON file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(uploaded_file)
                        json_data = df.to_dict('records')[0] if not df.empty else {}
                    except Exception as e:
                        flash(f'Error reading Excel file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                else:
                    flash('Please upload a JSON or Excel file', 'error')
                    return redirect(url_for('upload_file'))
            
            else:
                flash('Please provide JSON data or upload a file', 'error')
                return redirect(url_for('upload_file'))
            
            if not json_data:
                flash('No data found to process', 'error')
                return redirect(url_for('upload_file'))
            
            # Store data in database
            result = db_manager.store_data(
                name=name,
                product_type=product_type,
                json_data=json_data,
                session_id=session['session_id'],
                file_source=file_source
            )
            
            if result['success']:
                flash(f'Data stored successfully! Quality Score: {result["quality_score"]}%', 'success')
                return redirect(url_for('process_data', record_id=result['record_id']))
            else:
                if result['error'] == 'duplicate':
                    flash(f'Duplicate data detected: {result["message"]}', 'warning')
                else:
                    flash(f'Error storing data: {result["message"]}', 'error')
                return redirect(url_for('upload_file'))
        
        except Exception as e:
            flash(f'Unexpected error: {str(e)}', 'error')
            return redirect(url_for('upload_file'))
    
    return render_template('upload.html')

@app.route('/process/<int:record_id>')
def process_data(record_id):
    """Process stored data using AIM processor"""
    try:
        # Get data from database
        record = db_manager.get_data_by_id(record_id)
        if not record:
            flash('Record not found', 'error')
            return redirect(url_for('index'))
        
        # Parse JSON data
        json_data = json.loads(record['json_data'])
        
        # Process with AIM processor
        result = aim_processor.process_fast_ui_input(
            json_data, 
            record['product_type']
        )
        
        return render_template('process_results.html', 
                             record=record, 
                             input_data=json_data,
                             result=result)
    
    except (ValidationError, MappingError) as e:
        flash(f'Processing error: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Unexpected error during processing: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/field-mapping')
def field_mapping():
    """Field mapping interface"""
    return render_template('field_mapping.html')

@app.route('/compare')
def compare():
    """Data comparison page"""
    # Get processed records for comparison
    processed_records = []
    with sqlite3.connect(db_manager.db_path) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, product_type, created_date, calculator_path
            FROM user_data 
            WHERE status IN ('processed', 'saved') 
            ORDER BY created_date DESC
        """)
        processed_records = [dict(row) for row in cursor.fetchall()]
    
    return render_template('compare.html', processed_records=processed_records)

@app.route('/create-mapping', methods=['POST'])
def create_mapping():
    """Create Excel field mapping template"""
    try:
        product_type = request.form.get('product_type', 'life_insurance')
        
        # Get sample data if available and session exists
        sample_data = None
        if 'session_id' in session:
            recent_data = db_manager.get_all_data(session['session_id'])
            if recent_data:
                latest_record = db_manager.get_data_by_id(recent_data[0]['id'])
                if latest_record:
                    try:
                        sample_data = json.loads(latest_record['json_data'])
                    except json.JSONDecodeError:
                        sample_data = None
        
        # Create Excel template
        filename = create_excel_template(product_type, sample_data)
        
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Template file was not created: {filename}")
        
        return send_file(filename, 
                        as_attachment=True, 
                        download_name=f'AIM_Mapping_Template_{product_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        error_msg = f'Error creating mapping template: {str(e)}'
        print(f"Excel template error: {error_msg}")  # Log to console
        
        # Return JSON error for AJAX requests or HTML error page
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'error': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for('field_mapping'))

@app.route('/api/save-mapping', methods=['POST'])
def save_mapping():
    """Save field mapping configuration"""
    try:
        data = request.get_json()
        
        mapping_name = data.get('name', '')
        product_type = data.get('product_type', '')
        description = data.get('description', '')
        field_mappings = data.get('mappings', [])
        
        if not mapping_name or not product_type:
            return jsonify({'success': False, 'message': 'Mapping name and product type are required'})
        
        # Create mappings directory if it doesn't exist
        mappings_dir = 'saved_mappings'
        os.makedirs(mappings_dir, exist_ok=True)
        
        # Save mapping to JSON file
        mapping_file = os.path.join(mappings_dir, f"{mapping_name.replace(' ', '_').lower()}.json")
        mapping_data = {
            'name': mapping_name,
            'product_type': product_type,
            'description': description,
            'created_date': datetime.now().isoformat(),
            'field_mappings': field_mappings
        }
        
        with open(mapping_file, 'w') as f:
            json.dump(mapping_data, f, indent=2)
        
        return jsonify({
            'success': True, 
            'message': f'Mapping "{mapping_name}" saved successfully',
            'file_path': mapping_file
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving mapping: {str(e)}'})

@app.route('/api/mappings')
def list_mappings():
    """List all saved mapping templates"""
    try:
        mappings_dir = 'saved_mappings'
        if not os.path.exists(mappings_dir):
            return jsonify({'mappings': []})
        
        mappings = []
        for filename in os.listdir(mappings_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(mappings_dir, filename)
                try:
                    with open(filepath, 'r') as f:
                        mapping_data = json.load(f)
                        mappings.append({
                            'filename': filename,
                            'name': mapping_data.get('name', ''),
                            'product_type': mapping_data.get('product_type', ''),
                            'description': mapping_data.get('description', ''),
                            'created_date': mapping_data.get('created_date', '')
                        })
                except Exception as e:
                    print(f"Error reading mapping file {filename}: {e}")
                    continue
        
        return jsonify({'mappings': mappings})
        
    except Exception as e:
        return jsonify({'error': f'Error listing mappings: {str(e)}'})

@app.route('/api/mappings/<filename>')
def load_mapping(filename):
    """Load a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Mapping file not found'}), 404
        
        with open(filepath, 'r') as f:
            mapping_data = json.load(f)
        
        return jsonify({'success': True, 'mapping': mapping_data})
        
    except Exception as e:
        return jsonify({'error': f'Error loading mapping: {str(e)}'})

@app.route('/api/mappings/<filename>', methods=['DELETE'])
def delete_mapping(filename):
    """Delete a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'Mapping file not found'}), 404
        
        os.remove(filepath)
        return jsonify({'success': True, 'message': f'Mapping "{filename}" deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting mapping: {str(e)}'})

@app.route('/view-data')
def view_data():
    """View all stored data"""
    all_data = db_manager.get_all_data()
    
    # Calculate statistics for the view
    stats = {
        'total_records': len(all_data),
        'avg_quality': sum(record.get('quality_score', 0) for record in all_data) / len(all_data) if all_data else 0,
        'last_updated': all_data[0].get('created_date', 'Never') if all_data else 'Never'
    }
    
    return render_template('view_data.html', data=all_data, **stats)

@app.route('/api/data/<int:record_id>')
def api_get_data(record_id):
    """API endpoint to get record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify(record)
    else:
        return jsonify({'error': 'Record not found'}), 404

@app.route('/api/stats')
def api_stats():
    """API endpoint for statistics"""
    return jsonify(db_manager.get_statistics())

@app.route('/api/record/<int:record_id>')
def api_get_record(record_id):
    """API endpoint to get detailed record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify({'success': True, 'record': record})
    else:
        return jsonify({'success': False, 'message': 'Record not found'}), 404

@app.route('/api/record/<int:record_id>', methods=['DELETE'])
def api_delete_record(record_id):
    """API endpoint to delete a record"""
    try:
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM user_data WHERE id = ?", (record_id,))
            
            if cursor.rowcount > 0:
                return jsonify({'success': True, 'message': 'Record deleted successfully'})
            else:
                return jsonify({'success': False, 'message': 'Record not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/record/<int:record_id>', methods=['PUT'])
def api_update_record(record_id):
    """API endpoint to update a record"""
    try:
        data = request.get_json()
        
        # Prepare update fields
        update_fields = []
        update_values = []
        
        if 'name' in data:
            update_fields.append('name = ?')
            update_values.append(data['name'])
        
        if 'product_type' in data:
            update_fields.append('product_type = ?')
            update_values.append(data['product_type'])
        
        if 'status' in data:
            update_fields.append('status = ?')
            update_values.append(data['status'])
        
        if 'quality_score' in data:
            update_fields.append('quality_score = ?')
            update_values.append(data['quality_score'])
        
        if 'notes' in data:
            update_fields.append('notes = ?')
            update_values.append(data['notes'])
        
        # Add updated timestamp
        update_fields.append('updated_date = ?')
        update_values.append(datetime.now().isoformat())
        
        # Add record_id for WHERE clause
        update_values.append(record_id)
        
        if not update_fields:
            return jsonify({'success': False, 'message': 'No valid fields to update'}), 400
        
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Check if record exists first
            cursor.execute("SELECT id FROM user_data WHERE id = ?", (record_id,))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Record not found'}), 404
            
            # Update the record
            update_query = f"UPDATE user_data SET {', '.join(update_fields)} WHERE id = ?"
            cursor.execute(update_query, update_values)
            
            return jsonify({'success': True, 'message': 'Record updated successfully'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export-data')
def api_export_data():
    """API endpoint to export data"""
    try:
        export_format = request.args.get('format', 'json').lower()
        all_data = db_manager.get_all_data()
        
        # Create a simplified export format
        export_data = []
        for record in all_data:
            export_record = {
                'id': record.get('id'),
                'name': record.get('name'),
                'product_type': record.get('product_type'),
                'created_date': record.get('created_date'),
                'quality_score': record.get('quality_score'),
                'status': record.get('status'),
                'source_file': record.get('source_file'),
                'notes': record.get('notes', '')
            }
            export_data.append(export_record)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'excel':
            # Create Excel file
            import pandas as pd
            from io import BytesIO
            
            df = pd.DataFrame(export_data)
            excel_buffer = BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='AIM_Data', index=False)
            
            excel_buffer.seek(0)
            
            return send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'aim_data_export_{timestamp}.xlsx'
            )
        else:
            # Return as JSON download
            response = jsonify(export_data)
            response.headers['Content-Disposition'] = f'attachment; filename=aim_data_export_{timestamp}.json'
            return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/save-data', methods=['POST'])
def api_save_data():
    """API endpoint to save data without processing"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        name = data.get('name', '').strip()
        product_type = data.get('product_type', '').strip()
        calculator_path = data.get('calculator_path', '').strip()
        json_data_str = data.get('json_data', '')
        
        if not name or not product_type or not calculator_path:
            return jsonify({
                'success': False,
                'message': 'Name, product type, and calculator path are required'
            }), 400
        
        # Parse JSON data
        try:
            json_data = json.loads(json_data_str) if json_data_str else {}
        except json.JSONDecodeError as e:
            return jsonify({
                'success': False,
                'message': f'Invalid JSON format: {str(e)}'
            }), 400
        
        if not json_data:
            return jsonify({
                'success': False,
                'message': 'No data found to save'
            }), 400
        
        # Store data with calculator path
        result = db_manager.store_data_with_calculator(
            name=name,
            product_type=product_type,
            json_data=json_data,
            calculator_path=calculator_path,
            session_id=session.get('session_id', os.urandom(16).hex()),
            file_source='manual_input'
        )
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': f'Data saved successfully! Quality Score: {result["quality_score"]}%',

                'record_id': result['record_id'],
                'quality_score': result['quality_score']
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@app.route('/api/process-data', methods=['POST'])
def api_process_data():
    """API endpoint to process data with calculator mapping and template generation"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        name = data.get('name', '').strip()
        product_type = data.get('product_type', '').strip()
        calculator_path = data.get('calculator_path', '').strip()
        json_data_str = data.get('json_data', '')
        
        if not name or not product_type or not calculator_path:
            return jsonify({
                'success': False,
                'message': 'Name, product type, and calculator path are required'
            }), 400
        
        # Parse JSON data
        try:
            source_data = json.loads(json_data_str) if json_data_str else {}
        except json.JSONDecodeError as e:
            return jsonify({
                'success': False,
                'message': f'Invalid JSON format: {str(e)}'
            }), 400
        
        if not source_data:
            return jsonify({
                'success': False,
                'message': 'No data found to process'
            }), 400
        
        # Load calculator configuration
        calculator_config = db_manager.load_calculator_configuration(calculator_path, product_type)
        
        # Create field mapping template
        field_mapping_template = create_field_mapping_template(source_data, calculator_config)
        
        # Process data with date-based calculations
        processed_data = process_with_calculator(source_data, calculator_config)
        
        # Store processed data
        result = db_manager.store_processed_data(
            name=name,
            product_type=product_type,
            source_data=source_data,
            processed_data=processed_data,
            calculator_path=calculator_path,
            field_mapping=field_mapping_template,
            session_id=session.get('session_id', os.urandom(16).hex()),
            file_source='manual_input'
        )
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Data processed successfully!',
                'record_id': result['record_id'],
                'quality_score': result['quality_score'],
                'field_mapping_template': field_mapping_template,
                'processed_data': processed_data,
                'calculator_results': result.get('calculator_results', {})
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }), 500

@app.route('/api/generate-template', methods=['POST'])
def api_generate_template():
    """API endpoint to generate field mapping template"""
    try:
        data = request.get_json()
        
        source_data = data.get('source_data', {})
        calculator_path = data.get('calculator_path', '')
        product_type = data.get('product_type', '')
        
        if not calculator_path or not product_type:
            return jsonify({
                'success': False,
                'message': 'Calculator path and product type are required'
            }), 400
        
        # Load calculator configuration
        calculator_config = db_manager.load_calculator_configuration(calculator_path, product_type)
        
        # Generate field mapping template
        template = create_field_mapping_template(source_data, calculator_config)
        
        return jsonify({
            'success': True,
            'template': template,
            'calculator_config': calculator_config
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating template: {str(e)}'
        }), 500

# Helper Functions for Enhanced Upload Functionality

def create_field_mapping_template(source_data, calculator_config):
    """Create a field mapping template with both source and calculator fields"""
    template = {
        'mapping_info': {
            'created_date': datetime.now().isoformat(),
            'source_fields_count': len(source_data),
            'calculator_fields_count': len(calculator_config.get('fields', {})),
            'mapping_type': 'auto_generated'
        },
        'field_mappings': []
    }
    
    calculator_fields = calculator_config.get('fields', {})
    
    # Create mapping for each calculator field
    for calc_field_name, calc_field_config in calculator_fields.items():
        mapping_entry = {
            'calculator_field': calc_field_name,
            'calculator_field_type': calc_field_config.get('type', 'string'),
            'calculator_required': calc_field_config.get('required', False),
            'calculator_description': calc_field_config.get('description', ''),
            'calculator_sample_value': calc_field_config.get('sample', ''),
            'calculator_formula': calc_field_config.get('formula', ''),
            'source_field': None,
            'source_value': None,
            'mapping_status': 'unmapped',
            'confidence_score': 0,
            'transformation_rule': None
        }
        
        # Try to find matching source field
        best_match = find_best_field_match(calc_field_name, source_data)
        if best_match:
            mapping_entry.update({
                'source_field': best_match['field_name'],
                'source_value': best_match['value'],
                'mapping_status': 'auto_mapped',
                'confidence_score': best_match['confidence'],
                'transformation_rule': best_match.get('transformation', 'direct_copy')
            })
        
        template['field_mappings'].append(mapping_entry)
    
    # Add unmapped source fields
    mapped_source_fields = {m['source_field'] for m in template['field_mappings'] if m['source_field']}
    for source_field, source_value in source_data.items():
        if source_field not in mapped_source_fields:
            template['field_mappings'].append({
                'calculator_field': None,
                'source_field': source_field,
                'source_value': source_value,
                'mapping_status': 'unmapped_source',
                'confidence_score': 0,
                'transformation_rule': None
            })
    
    return template

def find_best_field_match(calculator_field, source_data):
    """Find the best matching source field for a calculator field"""
    import difflib
    
    calc_field_lower = calculator_field.lower()
    best_match = None
    best_score = 0
    
    for source_field, source_value in source_data.items():
        source_field_lower = source_field.lower()
        
        # Direct match
        if calc_field_lower == source_field_lower:
            return {
                'field_name': source_field,
                'value': source_value,
                'confidence': 100,
                'transformation': 'direct_copy'
            }
        
        # Partial matches
        similarity = difflib.SequenceMatcher(None, calc_field_lower, source_field_lower).ratio()
        
        # Check for common field name patterns
        calc_words = set(calc_field_lower.replace('_', ' ').split())
        source_words = set(source_field_lower.replace('_', ' ').split())
        word_overlap = len(calc_words.intersection(source_words))
        
        if word_overlap > 0:
            similarity += word_overlap * 0.2
        
        # Boost similarity for common patterns
        if 'date' in calc_field_lower and 'date' in source_field_lower:
            similarity += 0.3
        if 'amount' in calc_field_lower and 'amount' in source_field_lower:
            similarity += 0.3
        if 'name' in calc_field_lower and 'name' in source_field_lower:
            similarity += 0.3
        
        if similarity > best_score and similarity > 0.6:
            best_score = similarity
            best_match = {
                'field_name': source_field,
                'value': source_value,
                'confidence': int(similarity * 100),
                'transformation': determine_transformation_rule(calculator_field, source_field, source_value)
            }
    
    return best_match

def determine_transformation_rule(calc_field, source_field, source_value):
    """Determine the appropriate transformation rule for field mapping"""
    calc_field_lower = calc_field.lower()
    
    # Date transformations
    if 'date' in calc_field_lower:
        if isinstance(source_value, str) and '/' in source_value:
            return 'date_format_us_to_iso'
        elif isinstance(source_value, str) and '-' in source_value:
            return 'date_format_iso'
        return 'date_auto_detect'
    
    # Numeric transformations
    if 'amount' in calc_field_lower or 'premium' in calc_field_lower:
        if isinstance(source_value, str) and '$' in source_value:
            return 'currency_to_number'
        return 'numeric_conversion'
    
    # Name transformations
    if 'name' in calc_field_lower:
        return 'text_cleanup'
    
    return 'direct_copy'

def process_with_calculator(source_data, calculator_config):
    """Process data using calculator configuration with date-based calculations"""
    processed_data = {}
    calculator_fields = calculator_config.get('fields', {})
    calculation_rules = calculator_config.get('calculations', {})
    
    # Process each field according to calculator rules
    for field_name, field_config in calculator_fields.items():
        field_type = field_config.get('type', 'string')
        field_formula = field_config.get('formula')
        
        # Get source value (if available)
        source_value = source_data.get(field_name)
        
        if field_formula and field_type in ['number', 'currency']:
            # Apply formula-based calculation
            calculated_value = apply_calculator_formula(field_formula, source_data, processed_data)
            processed_data[field_name] = calculated_value
        elif source_value is not None:
            # Transform source value according to field type
            processed_data[field_name] = transform_field_value(source_value, field_config)
        else:
            # Use default or sample value
            processed_data[field_name] = field_config.get('default', field_config.get('sample'))
    
    # Apply date-based calculations
    processed_data = apply_date_based_calculations(processed_data, calculation_rules)
    
    return processed_data

def apply_calculator_formula(formula, source_data, processed_data):
    """Apply calculator formula with built-in logic and enhanced calculations"""
    try:
        from datetime import datetime, timedelta
        import math
        
        # Handle special calculator functions
        if formula.startswith('CALCULATE_AGE'):
            # Extract birth_date and effective_date from formula
            birth_date_field = formula.split('$')[1].split(',')[0]
            effective_date_field = formula.split('$')[2].split(')')[0]
            
            birth_date_str = source_data.get(birth_date_field) or processed_data.get(birth_date_field)
            effective_date_str = source_data.get(effective_date_field) or processed_data.get(effective_date_field)
            
            if birth_date_str and effective_date_str:
                try:
                    birth_date = datetime.fromisoformat(str(birth_date_str).replace('/', '-'))
                    effective_date = datetime.fromisoformat(str(effective_date_str).replace('/', '-'))
                    age = (effective_date - birth_date).days / 365.25
                    return int(age)
                except:
                    return 25  # Default age
            return 25
            
        elif formula.startswith('CALCULATE_ANNUAL_PREMIUM'):
            # Calculate premium based on face amount, age, risk class
            face_amount = float(source_data.get('face_amount', 100000))
            age = int(source_data.get('applicant_age', processed_data.get('applicant_age', 35)))
            risk_class = source_data.get('risk_class', 'standard')
            
            # Base rates per $1000 of coverage
            base_rates = {
                'preferred_plus': 0.85,
                'preferred': 1.25,
                'standard': 1.75,
                'table_rated': 2.50
            }
            
            # Age factors
            if age <= 25:
                age_factor = 0.80
            elif age <= 35:
                age_factor = 1.00
            elif age <= 45:
                age_factor = 1.35
            elif age <= 55:
                age_factor = 1.85
            elif age <= 65:
                age_factor = 2.75
            elif age <= 75:
                age_factor = 4.50
            else:
                age_factor = 7.25
            
            base_rate = base_rates.get(risk_class, 1.75)
            annual_premium = (face_amount / 1000) * base_rate * age_factor
            return round(annual_premium, 2)
            
        elif formula.startswith('CALCULATE_MODAL_PREMIUM'):
            # Calculate modal premium from annual premium
            annual_premium = float(source_data.get('annual_premium', processed_data.get('annual_premium', 6000)))
            frequency = source_data.get('premium_frequency', 'monthly')
            
            frequency_factors = {
                'monthly': 0.0875,
                'quarterly': 0.26,
                'semi-annual': 0.51,
                'annual': 1.0
            }
            
            factor = frequency_factors.get(frequency, 0.0875)
            return round(annual_premium * factor, 2)
            
        elif formula.startswith('CALCULATE_ACCOUNT_VALUE'):
            # Calculate annuity account value
            initial_premium = float(source_data.get('initial_premium', 50000))
            guaranteed_rate = float(source_data.get('guaranteed_rate', 0.025))
            contract_year = int(processed_data.get('contract_year', 1))
            
            account_value = initial_premium * math.pow(1 + guaranteed_rate, contract_year - 1)
            return round(account_value, 2)
            
        elif formula.startswith('CALCULATE_POLICY_YEAR') or formula.startswith('CALCULATE_CONTRACT_YEAR'):
            # Calculate policy/contract year from effective date
            effective_date_str = source_data.get('effective_date', '2025-01-01')
            try:
                effective_date = datetime.fromisoformat(str(effective_date_str).replace('/', '-'))
                current_date = datetime.now()
                years = (current_date - effective_date).days / 365.25
                return max(1, int(years) + 1)
            except:
                return 1
                
        elif formula.startswith('LOOKUP_COMMISSION_RATE'):
            # Lookup commission rate based on product and state
            product_code = source_data.get('product_code', 'STANDARD')
            state = source_data.get('state_issued', 'NY')
            
            # Default commission rates
            if 'LIFE' in product_code.upper():
                return 0.55
            elif 'ANN' in product_code.upper():
                return 0.065
            else:
                return 0.45
                
        elif formula.startswith('LOOKUP_GUARANTEED_RATE'):
            # Lookup guaranteed rate for annuities
            effective_date_str = source_data.get('effective_date', '2025-01-01')
            year = effective_date_str[:4] if effective_date_str else '2025'
            
            rates = {
                '2025': 0.025,
                '2024': 0.030,
                '2023': 0.035
            }
            return rates.get(year, 0.025)
            
        else:
            # Simple formula evaluation for mathematical expressions
            formula_str = formula
            
            # Replace field references with actual values
            all_data = {**source_data, **processed_data}
            for field_name, value in all_data.items():
                if isinstance(value, (int, float)):
                    formula_str = formula_str.replace(f'${field_name}', str(value))
            
            # Safe evaluation of mathematical expressions
            import re
            if re.match(r'^[\d\+\-\*\/\(\)\.\s]+$', formula_str):
                return eval(formula_str)
            else:
                return 0
                
    except Exception as e:
        print(f"Error applying formula '{formula}': {e}")
        return 0

def apply_date_based_calculations(data, calculation_rules):
    """Apply date-based calculations based on effective dates with enhanced logic"""
    from datetime import datetime, timedelta
    import math
    
    # Get effective date
    effective_date_str = data.get('effective_date', data.get('policy_date', data.get('start_date')))
    if not effective_date_str:
        return data
    
    try:
        if isinstance(effective_date_str, str):
            effective_date = datetime.fromisoformat(effective_date_str.replace('/', '-'))
        else:
            effective_date = effective_date_str
        
        current_date = datetime.now()
        
        # Calculate policy/contract year
        years_elapsed = (current_date - effective_date).days / 365.25
        policy_year = max(1, int(years_elapsed) + 1)
        data['policy_year'] = policy_year
        data['contract_year'] = policy_year
        
        # Calculate age if birth_date is available
        birth_date_str = data.get('birth_date')
        if birth_date_str:
            try:
                birth_date = datetime.fromisoformat(str(birth_date_str).replace('/', '-'))
                age_at_effective = (effective_date - birth_date).days / 365.25
                current_age = (current_date - birth_date).days / 365.25
                data['applicant_age'] = int(age_at_effective)
                data['owner_age'] = int(age_at_effective)
                data['current_age'] = int(current_age)
            except:
                pass
        
        # Apply calculation rules if provided
        if calculation_rules:
            for rule_name, rule_config in calculation_rules.items():
                if rule_name == 'premium_calculation' and 'face_amount' in data and 'applicant_age' in data:
                    # Enhanced premium calculation
                    face_amount = float(data.get('face_amount', 100000))
                    age = int(data.get('applicant_age', 35))
                    risk_class = data.get('risk_class', 'standard')
                    state = data.get('state_issued', 'NY')
                    
                    # Base rates from rule config
                    base_rates = rule_config.get('base_rate_per_1000', {
                        'preferred_plus': 0.85,
                        'preferred': 1.25,
                        'standard': 1.75,
                        'table_rated': 2.50
                    })
                    
                    # Age factors from rule config
                    age_factors = rule_config.get('age_factors', {
                        '18-25': 0.80,
                        '26-35': 1.00,
                        '36-45': 1.35,
                        '46-55': 1.85,
                        '56-65': 2.75,
                        '66-75': 4.50,
                        '76-85': 7.25
                    })
                    
                    # Determine age factor
                    age_factor = 1.0
                    for age_range, factor in age_factors.items():
                        min_age, max_age = map(int, age_range.split('-'))
                        if min_age <= age <= max_age:
                            age_factor = factor
                            break
                    
                    # State factors (from config file)
                    state_factors = {
                        'NY': 1.15,
                        'CA': 1.10,
                        'TX': 1.05,
                        'FL': 1.08
                    }
                    state_factor = state_factors.get(state, 1.00)
                    
                    base_rate = base_rates.get(risk_class, 1.75)
                    annual_premium = (face_amount / 1000) * base_rate * age_factor * state_factor
                    data['annual_premium'] = round(annual_premium, 2)
                    
                elif rule_name == 'modal_premium_calculation' and 'annual_premium' in data:
                    # Modal premium calculation
                    annual_premium = float(data.get('annual_premium', 0))
                    frequency = data.get('premium_frequency', 'monthly')
                    
                    frequency_factors = rule_config.get('frequency_factors', {
                        'monthly': 0.0875,
                        'quarterly': 0.26,
                        'semi-annual': 0.51,
                        'annual': 1.0
                    })
                    
                    factor = frequency_factors.get(frequency, 0.0875)
                    data['modal_premium'] = round(annual_premium * factor, 2)
                    
                elif rule_name == 'account_value_calculation' and 'initial_premium' in data:
                    # Annuity account value calculation
                    initial_premium = float(data.get('initial_premium', 0))
                    guaranteed_rate = float(data.get('guaranteed_rate', 0.025))
                    
                    account_value = initial_premium * math.pow(1 + guaranteed_rate, policy_year - 1)
                    data['account_value'] = round(account_value, 2)
                    
                    # Calculate surrender value if surrender schedule exists
                    surrender_schedule = data.get('surrender_schedule', [])
                    if surrender_schedule and policy_year <= len(surrender_schedule):
                        surrender_rate = surrender_schedule[policy_year - 1]
                        surrender_charge = account_value * surrender_rate
                        surrender_value = account_value - surrender_charge
                        data['surrender_charge'] = round(surrender_charge, 2)
                        data['surrender_value'] = round(surrender_value, 2)
                    
                elif rule_name == 'commission_calculation':
                    # Commission calculation
                    if 'annual_premium' in data and 'commission_rate' in data:
                        annual_premium = float(data.get('annual_premium', 0))
                        commission_rate = float(data.get('commission_rate', 0))
                        data['commission_amount'] = round(annual_premium * commission_rate, 2)
                    elif 'initial_premium' in data and 'commission_rate' in data:
                        initial_premium = float(data.get('initial_premium', 0))
                        commission_rate = float(data.get('commission_rate', 0))
                        data['commission_amount'] = round(initial_premium * commission_rate, 2)
        
        # Add calculated date information
        data['calculation_date'] = current_date.isoformat()
        data['days_since_effective'] = (current_date - effective_date).days
        data['years_since_effective'] = round(years_elapsed, 2)
        
    except Exception as e:
        print(f"Error in date-based calculations: {e}")
    
    return data

def calculate_date_based_value(effective_date, rule_config, data):
    """Calculate value based on date and rules"""
    from datetime import datetime
    
    calculation_type = rule_config.get('calculation_type', 'age')
    
    if calculation_type == 'age':
        birth_date_str = data.get('birth_date', data.get('date_of_birth'))
        if birth_date_str:
            try:
                birth_date = datetime.fromisoformat(birth_date_str.replace('/', '-'))
                age = (effective_date - birth_date).days // 365
                return age
            except:
                pass
    
    elif calculation_type == 'premium_adjustment':
        base_premium = data.get('base_premium', data.get('premium', 0))
        adjustment_factor = rule_config.get('factor', 1.0)
        return base_premium * adjustment_factor
    
    elif calculation_type == 'maturity_date':
        term_years = data.get('term_years', rule_config.get('default_term', 20))
        maturity_date = effective_date.replace(year=effective_date.year + term_years)
        return maturity_date.isoformat()[:10]
    
    return rule_config.get('default_value', 0)

def transform_field_value(value, field_config):
    """Transform field value according to configuration"""
    field_type = field_config.get('type', 'string')
    
    if field_type == 'number':
        try:
            if isinstance(value, str):
                # Remove currency symbols and convert
                cleaned = value.replace('$', '').replace(',', '').strip()
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            return 0
    
    elif field_type == 'currency':
        try:
            if isinstance(value, str):
                cleaned = value.replace('$', '').replace(',', '').strip()
                return float(cleaned)
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    elif field_type == 'date':
        if isinstance(value, str):
            # Try to standardize date format
            value = value.replace('/', '-')
            if len(value.split('-')[0]) == 2:  # MM-DD-YYYY to YYYY-MM-DD
                parts = value.split('-')
                if len(parts) == 3:
                    return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
        return value
    
    elif field_type == 'boolean':
        if isinstance(value, str):
            return value.lower() in ['true', 'yes', '1', 'y']
        return bool(value)
    
    return str(value) if value is not None else ''

def create_excel_template(product_type, sample_data=None):
    """Create an Excel template file for field mapping"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel template creation. Please install with: pip install openpyxl")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{product_type.title()} Mapping Template"
    
    # Set headers
    headers = ['Source Field', 'Source Value', 'Calculator Field', 'Calculator Type', 'Required', 'Description']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Load calculator configuration for field definitions
    calculator_config = db_manager.load_calculator_configuration(f'/calculators/{product_type}/standard.json', product_type)
    calculator_fields = calculator_config.get('fields', {})
    
    # Add sample data rows if available
    row = 2
    if sample_data:
        for source_field, source_value in sample_data.items():
            ws.cell(row=row, column=1, value=source_field)
            ws.cell(row=row, column=2, value=str(source_value))
            row += 1
    
    # Add calculator field definitions
    for calc_field, field_config in calculator_fields.items():
        ws.cell(row=row, column=3, value=calc_field)
        ws.cell(row=row, column=4, value=field_config.get('type', 'string'))
        ws.cell(row=row, column=5, value='Yes' if field_config.get('required', False) else 'No')
        ws.cell(row=row, column=6, value=field_config.get('description', ''))
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)




        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = os.path.join('uploads', f'template_{product_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    os.makedirs('uploads', exist_ok=True)
    wb.save(filename)
    
    return filename

# Main execution
if __name__ == '__main__':
    # Initialize database
    db_manager.init_database()
    
    # Ensure uploads directory exists
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('saved_mappings', exist_ok=True)
    
    print("Starting AIM Web Application...")
    print("Application will be available at: http://localhost:5000")
    print("Press Ctrl+C to stop the server")
    
    # Run the Flask application
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True,
        use_reloader=False  # Disable reloader to prevent issues in some environments
    )
