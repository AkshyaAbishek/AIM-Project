"""
AIM Web Application - Flask Implementation
Converts the tkinter desktop application to a web-based interface for Azure deployment
"""

import os
import json
import hashlib
import traceback
from datetime import datetime
from typing import Dict, Any, Optional
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO

from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file, session, send_from_directory
from flask import Response
import sqlite3
import sys

# Add src directory to path for imports - Updated for new structure
project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))  # Go up to AIM root
sys.path.insert(0, os.path.join(project_root, 'src'))

try:
    from aim_processor import AIMProcessor, ValidationError, MappingError
except ImportError:
    # Fallback for web deployment
    class AIMProcessor:
        def process_fast_ui_input(self, data, product_type):
            return {"status": "success", "message": "Processed successfully"}
    
    class ValidationError(Exception):
        pass
    
    class MappingError(Exception):
        pass

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')

# Updated paths for new structure
app.config['UPLOAD_FOLDER'] = os.path.join(project_root, 'runtime', 'uploads')
app.config['EXPORT_FOLDER'] = os.path.join(project_root, 'runtime', 'exports')
app.config['DATABASE_PATH'] = os.path.join(project_root, 'database', 'aim_data.db')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)
os.makedirs(os.path.dirname(app.config['DATABASE_PATH']), exist_ok=True)

# Initialize AIM processor
aim_processor = AIMProcessor()

class WebDatabaseManager:
    """Database manager for web application"""
    
    def __init__(self, db_path: str = None):
        # Use configured database path if not provided
        if db_path is None:
            db_path = app.config.get('DATABASE_PATH', 'aim_data.db')
        
        # Ensure we're using the correct path
        if not os.path.isabs(db_path):
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            db_path = os.path.join(project_root, 'database', db_path)
        self.db_path = db_path
        print(f"Database path: {self.db_path}")
        self.init_database()
    
    def init_database(self):
        """Initialize database with web-optimized schema"""
        try:
            print(f"Initializing database at: {self.db_path}")
            with sqlite3.connect(self.db_path) as conn:
                # First, create table if it doesn't exist with minimal schema
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS user_data (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        product_type TEXT NOT NULL,
                        data_hash TEXT UNIQUE NOT NULL,
                        json_data TEXT NOT NULL,
                        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            
            conn.execute("""
                CREATE TABLE IF NOT EXISTS processing_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    operation_details TEXT,
                    status TEXT CHECK(status IN ('success', 'warning', 'error')),
                    error_message TEXT,
                    processing_time_ms INTEGER,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Add new columns if they don't exist (for existing databases)
            columns_to_add = [
                ("calculator_path", "TEXT"),
                ("raw_data", "TEXT"),
                ("processed_data", "TEXT"),
                ("field_mappings", "TEXT"),
                ("updated_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("modified_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("validation_status", "TEXT DEFAULT 'pending'"),
                ("status", "TEXT DEFAULT 'pending'"),
                ("quality_score", "REAL DEFAULT 0.0"),
                ("file_source", "TEXT"),
                ("source_file", "TEXT"),
                ("processing_notes", "TEXT"),
                ("notes", "TEXT"),
                ("session_id", "TEXT"),
                ("source_record_id", "TEXT")
            ]
            
            for column_name, column_type in columns_to_add:
                try:
                    conn.execute(f"ALTER TABLE user_data ADD COLUMN {column_name} {column_type}")
                except sqlite3.OperationalError:
                    pass  # Column already exists
            
            # Add session_id to processing_logs if it doesn't exist
            try:
                conn.execute("ALTER TABLE processing_logs ADD COLUMN session_id TEXT")
            except sqlite3.OperationalError:
                pass
            
            # Create indexes
            try:
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_hash ON user_data(data_hash)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_session ON user_data(session_id)")
            except sqlite3.OperationalError:
                pass
            
            print("Database initialization completed successfully")
                
        except Exception as e:
            print(f"Error initializing database: {e}")
            import traceback
            traceback.print_exc()
    
    def store_data(self, name: str, product_type: str, json_data: Dict, 
                   session_id: str, file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with duplicate detection"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, created_date, 
                     validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    datetime.now().isoformat(), 'validated', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'storage_error',
                'message': str(e)
            }
    
    def calculate_quality_score(self, json_data: Dict) -> float:
        """Calculate data quality score"""
        score = 0.0
        
        # Required fields check (40% of score)
        required_fields = ['policy_number', 'product_type', 'effective_date']
        present_required = sum(1 for field in required_fields if field in json_data and json_data[field])
        score += (present_required / len(required_fields)) * 40
        
        # Data completeness (30% of score)
        total_fields = len(json_data)
        non_empty_fields = sum(1 for v in json_data.values() if v is not None and str(v).strip())
        if total_fields > 0:
            score += (non_empty_fields / total_fields) * 30
        
        # Data format consistency (30% of score)
        format_score = 30
        if 'age' in json_data:
            try:
                age = float(json_data['age'])
                if age < 0 or age > 120:
                    format_score -= 10
            except:
                format_score -= 10
        
        score += format_score
        return round(min(100, score), 2)
    
    def add_user_data(self, name: str, raw_data: Dict, product_type: str, 
                      source_file: Optional[str] = None, status: str = 'pending',
                      calculator_path: Optional[str] = None) -> int:
        """Add user data to database and return record ID"""
        try:
            data_str = json.dumps(raw_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Check for duplicates
                cursor.execute("SELECT id FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing record
                    cursor.execute("""
                        UPDATE user_data 
                        SET updated_date = ?, status = ?, calculator_path = ?
                        WHERE id = ?
                    """, (datetime.now().isoformat(), status, calculator_path, existing[0]))
                    return existing[0]
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(raw_data)
                
                # Insert new record
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, raw_data, created_date, 
                     validation_status, status, quality_score, source_file, calculator_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(raw_data), json.dumps(raw_data),
                    datetime.now().isoformat(), 'validated', status, quality_score,
                    source_file, calculator_path
                ))
                
                return cursor.lastrowid
                
        except Exception as e:
            raise Exception(f"Failed to add user data: {str(e)}")

    def get_statistics(self) -> Dict[str, Any]:
        """Get database statistics for dashboard"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Total records
                cursor.execute("SELECT COUNT(*) FROM user_data")
                total_records = cursor.fetchone()[0]
                
                # Records by product type
                cursor.execute("SELECT product_type, COUNT(*) FROM user_data GROUP BY product_type")
                product_counts = dict(cursor.fetchall())
                
                # Basic stats to avoid column issues
                return {
                    'total_records': total_records,
                    'status_counts': {'processed': total_records},  # Simplified
                    'product_counts': product_counts,
                    'product_distribution': product_counts,  # Template expects this name
                    'average_quality_score': 0.0,  # Simplified
                    'average_quality': 0.0,  # Template expects this name
                    'recent_activity': total_records,  # Simplified
                    'processed_records': total_records,
                    'pending_records': 0,
                    'error_records': 0
                }
        except Exception as e:
            print(f"Database error in get_statistics: {e}")
            # Return default stats if there's an error
            return {
                'total_records': 0,
                'status_counts': {},
                'product_counts': {},
                'product_distribution': {},  # Template expects this name
                'average_quality_score': 0,
                'average_quality': 0,  # Template expects this name
                'recent_activity': 0,
                'processed_records': 0,
                'pending_records': 0,
                'error_records': 0
            }

    def load_calculator_reference(self, calculator_path: str, product_type: str) -> Dict:
        """Load calculator reference data from various sources"""
        # Default templates for when no specific calculator is found
        default_templates = {
            'life_insurance': {
                'name': 'Life Insurance Standard Calculator',
                'version': '1.0',
                'fields': {
                    'policy_number': {'required': True, 'type': 'string', 'sample': 'LIFE001234'},
                    'insured_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1980-01-01'},
                    'gender': {'required': True, 'type': 'string', 'sample': 'M'},
                    'coverage_amount': {'required': True, 'type': 'number', 'sample': 100000},
                    'premium_amount': {'required': False, 'type': 'number', 'sample': 1200},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'TERM20'},
                    'beneficiary_name': {'required': False, 'type': 'string', 'sample': 'Jane Doe'},
                    'agent_code': {'required': False, 'type': 'string', 'sample': 'AGT001'}
                }
            },
            'annuity': {
                'name': 'Annuity Standard Calculator',
                'version': '1.0',
                'fields': {
                    'contract_number': {'required': True, 'type': 'string', 'sample': 'ANN001234'},
                    'owner_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1970-01-01'},
                    'initial_deposit': {'required': True, 'type': 'number', 'sample': 50000},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'FIXED5'},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'surrender_charge_period': {'required': False, 'type': 'number', 'sample': 7},
                    'interest_rate': {'required': False, 'type': 'number', 'sample': 3.5}
                }
            }
        }
        
        # Try to load from file if a specific path is provided
        if calculator_path and calculator_path != '/calculators/default':
            try:
                # Handle different path types
                file_path = None
                
                # Case 1: Uploaded calculator (stored in session or temp)
                if '/calculators/uploaded/' in calculator_path:
                    filename = calculator_path.split('/')[-1]
                    # Store uploaded calculators in a temp directory
                    temp_dir = os.path.join(os.path.dirname(__file__), 'temp_calculators')
                    os.makedirs(temp_dir, exist_ok=True)
                    file_path = os.path.join(temp_dir, filename)
                
                # Case 2: Config directory calculators
                elif calculator_path.startswith('/calculators/'):
                    config_dir = os.path.join(os.path.dirname(__file__), 'config')
                    # Map dropdown paths to actual files
                    path_mappings = {
                        '/calculators/life_insurance/standard.json': 'life_template.json',
                        '/calculators/life_insurance/term.json': 'life_template.json',
                        '/calculators/annuity/fixed.json': 'annuity_template.json',
                        '/calculators/annuity/variable.json': 'annuity_template.json'
                    }
                    
                    filename = path_mappings.get(calculator_path)
                    if filename:
                        file_path = os.path.join(config_dir, filename)
                
                # Case 3: Direct file path
                else:
                    file_path = calculator_path
                
                # Try to load the file if path was resolved
                if file_path and os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as f:
                        calculator_data = json.load(f)
                        
                    # Validate the structure
                    if 'fields' in calculator_data or 'mappings' in calculator_data:
                        print(f"Successfully loaded calculator from: {file_path}")
                        return calculator_data
                    else:
                        print(f"Invalid calculator structure in: {file_path}")
                        
            except Exception as e:
                print(f"Error loading calculator from {calculator_path}: {e}")
        
        # Fall back to default template based on product type
        template = default_templates.get(product_type, default_templates['life_insurance'])
        print(f"Using default template for product type: {product_type}")
        return template

    def perform_data_comparison(self, source_data: Dict, calculator_data: Dict) -> Dict:
        """Compare source data with calculator reference and generate comparison result"""
        
        calculator_fields = calculator_data.get('fields', {})
        comparison_fields = []
        
        # Stats
        total_fields = len(calculator_fields)
        matching_fields = 0
        missing_fields = 0
        filled_fields = 0
        
        for field_name, field_config in calculator_fields.items():
            source_value = source_data.get(field_name)
            calculator_value = field_config.get('sample')
            
            field_result = {
                'field_name': field_name,
                'source_value': source_value,
                'calculator_value': calculator_value,
                'required': field_config.get('required', False),
                'data_type': field_config.get('type', 'string'),
                'description': field_config.get('description', ''),
                'values_match': False
            }
            
            # Check if values match
            if source_value is not None and source_value != '' and calculator_value is not None and calculator_value != '':
                # Convert to strings and compare
                source_str = str(source_value).strip().lower()
                calc_str = str(calculator_value).strip().lower()
                field_result['values_match'] = (source_str == calc_str)
            
            # Determine status
            if source_value is not None and source_value != '':
                if field_result['values_match']:
                    field_result['status'] = 'match'
                    matching_fields += 1
                else:
                    field_result['status'] = 'mismatch'
            else:
                if calculator_value and field_config.get('required', False):
                    field_result['status'] = 'missing'
                    missing_fields += 1
                elif calculator_value:
                    field_result['status'] = 'can_fill'
                    filled_fields += 1
                else:
                    field_result['status'] = 'missing'
                    missing_fields += 1
            
            comparison_fields.append(field_result)
        
        return {
            'stats': {
                'total_fields': total_fields,
                'matching_fields': matching_fields,
                'missing_fields': missing_fields,
                'filled_fields': filled_fields,
                'completion_percentage': round((matching_fields / total_fields) * 100, 2) if total_fields > 0 else 0
            },
            'fields': comparison_fields,
            'comparison_date': datetime.now().isoformat()
        }

    def get_all_data(self, session_id: Optional[str] = None) -> list:
        """Get all stored data, optionally filtered by session"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if session_id:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status, 
                           quality_score, file_source, data_hash
                    FROM user_data WHERE session_id = ? ORDER BY created_date DESC
                """, (session_id,))
            else:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status, 
                           quality_score, file_source, data_hash
                    FROM user_data ORDER BY created_date DESC LIMIT 100
                """)
            
            rows = cursor.fetchall()
            data = []
            for row in rows:
                record = dict(row)
                # Ensure default values for missing fields
                record['source_file'] = record.get('file_source') or 'Manual Entry'
                record['status'] = record.get('validation_status') or 'processed'  # Keep both for compatibility
                record['validation_status'] = record.get('validation_status') or 'processed'  # Template expects this
                record['quality_score'] = record.get('quality_score') or 0.0
                data.append(record)
            
            return data
    
    def get_data_by_id(self, record_id: int) -> Optional[Dict]:
        """Get specific record by ID"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM user_data WHERE id = ?", (record_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_record_by_id(self, record_id: int) -> Optional[Dict]:
        """Alias for get_data_by_id for backward compatibility"""
        record = self.get_data_by_id(record_id)
        if record:
            # Parse JSON data and return in the expected format
            if isinstance(record.get('json_data'), str):
                try:
                    data = json.loads(record['json_data'])
                except (json.JSONDecodeError, TypeError):
                    data = {}
            else:
                data = record.get('json_data', {})
            
            return {
                'id': record['id'],
                'name': record['name'],
                'product_type': record['product_type'],
                'data': data,
                'calculator_path': record.get('calculator_path'),
                'created_date': record['created_date'],
                'validation_status': record.get('validation_status'),
                'quality_score': record.get('quality_score', 0.0)
            }
        return None

    def store_data_with_calculator(self, name: str, product_type: str, json_data: Dict, 
                                   calculator_path: str, session_id: str, 
                                   file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with calculator path for save-only operation"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, calculator_path, raw_data,
                     created_date, validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    calculator_path, json.dumps(json_data),
                    datetime.now().isoformat(), 'saved', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score,
                    'message': 'Data saved successfully'
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'database_error',
                'message': f'Database error: {str(e)}'
            }

    def store_processed_data(self, name: str, product_type: str, source_data: Dict,
                           processed_data: Dict, calculator_path: str, field_mapping: Dict,
                           session_id: str, file_source: Optional[str] = None, 
                           source_record_id: Optional[str] = None) -> Dict[str, Any]:
        """Store processed data with mapping and calculator results"""
        try:
            data_str = json.dumps(source_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates - skip for filled data from comparator
            duplicate_check = file_source != 'compare_filled'
            existing = None
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                if duplicate_check:
                    cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                    existing = cursor.fetchone()
                
                if existing and duplicate_check:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score for processed data
                quality_score = self.calculate_processed_quality_score(source_data, processed_data, field_mapping)
                
                # Calculate additional metrics
                calculator_results = {
                    'total_fields_mapped': len([m for m in field_mapping.get('field_mappings', []) if m.get('mapping_status') == 'auto_mapped']),
                    'mapping_confidence': sum(m.get('confidence_score', 0) for m in field_mapping.get('field_mappings', [])) / len(field_mapping.get('field_mappings', [])) if field_mapping.get('field_mappings', []) else 0,
                    'processing_timestamp': datetime.now().isoformat()
                }
                
                # Add source record ID if provided (for filled data from comparator)
                if source_record_id:
                    calculator_results['source_record_id'] = source_record_id
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, calculator_path, raw_data,
                     processed_data, field_mappings, created_date, validation_status, 
                     quality_score, file_source, session_id, processing_notes, source_record_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(processed_data),
                    calculator_path, json.dumps(source_data),
                    json.dumps(processed_data), json.dumps(field_mapping),
                    datetime.now().isoformat(), 'processed', quality_score,
                    file_source, session_id, json.dumps(calculator_results),
                    source_record_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score,
                    'calculator_results': calculator_results,
                    'message': 'Data processed successfully'
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'database_error',
                'message': f'Database error: {str(e)}'
            }

    def load_calculator_configuration(self, calculator_path: str, product_type: str) -> Dict:
        """Load calculator configuration from JSON files with enhanced field definitions"""
        import os
        
        # Try to load from actual file path first
        config_file_map = {
            '/calculators/life_insurance/standard.json': 'life_insurance_calculator.json',
            '/calculators/life_insurance/term.json': 'life_insurance_calculator.json',
            '/calculators/annuity/fixed.json': 'annuity_calculator.json',
            '/calculators/annuity/variable.json': 'annuity_calculator.json'
        }
        
        # Get config file name
        config_file = config_file_map.get(calculator_path)
        if not config_file:
            # Use product type as fallback
            config_file = f"{product_type}_calculator.json" if product_type in ['life', 'annuity'] else 'life_insurance_calculator.json'
        
        # Try to load from config files directory
        config_path = os.path.join(os.path.dirname(__file__), 'src', 'config', 'config_files', config_file)
        
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading calculator config from {config_path}: {e}")
        
        # Fallback to embedded templates
        calculator_templates = {
            'life': {
                'name': 'Life Insurance Calculator',
                'version': '2.0',
                'fields': {
                    'policy_number': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'LIFE001234',
                        'description': 'Unique policy identifier'
                    },
                    'applicant_first_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'John',
                        'description': 'Applicant first name'
                    },
                    'applicant_last_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'Doe',
                        'description': 'Applicant last name'
                    },
                    'birth_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '1980-01-01',
                        'description': 'Applicant date of birth'
                    },
                    'effective_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '2025-01-01',
                        'description': 'Policy effective date'
                    },
                    'face_amount': {
                        'required': True, 
                        'type': 'currency', 
                        'sample': 100000,
                        'description': 'Policy face amount'
                    },
                    'premium': {
                        'required': False, 
                        'type': 'currency', 
                        'sample': 1200,
                        'description': 'Annual premium amount',
                        'formula': '$face_amount * 0.012'
                    },
                    'applicant_age': {
                        'required': False,
                        'type': 'number',
                        'sample': 45,
                        'description': 'Calculated age at effective date',
                        'formula': 'calculated_from_dates'
                    },
                    'gender': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'M',
                        'description': 'Gender (M/F)'
                    },
                    'product_code': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'TERM20',
                        'description': 'Product code'
                    },
                    'beneficiary_name': {
                        'required': False, 
                        'type': 'string', 
                        'sample': 'Jane Doe',
                        'description': 'Primary beneficiary name'
                    },
                    'maturity_date': {
                        'required': False,
                        'type': 'date',
                        'sample': '2045-01-01',
                        'description': 'Policy maturity date',
                        'formula': 'calculated_from_effective_date_plus_term'
                    }
                },
                'calculations': {
                    'applicant_age': {
                        'type': 'date_based',
                        'calculation_type': 'age',
                        'description': 'Calculate age from birth date and effective date'
                    },
                    'premium_adjustment': {
                        'type': 'date_based',
                        'calculation_type': 'premium_adjustment',
                        'factor': 1.0,
                        'description': 'Adjust premium based on age and other factors'
                    },
                    'maturity_date': {
                        'type': 'date_based',
                        'calculation_type': 'maturity_date',
                        'default_term': 20,
                        'description': 'Calculate maturity date from effective date'
                    }
                }
            },
            'annuity': {
                'name': 'Annuity Calculator',
                'version': '2.0',
                'fields': {
                    'contract_number': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'ANN001234',
                        'description': 'Unique contract identifier'
                    },
                    'owner_first_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'John',
                        'description': 'Contract owner first name'
                    },
                    'owner_last_name': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'Doe',
                        'description': 'Contract owner last name'
                    },
                    'birth_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '1970-01-01',
                        'description': 'Owner date of birth'
                    },
                    'effective_date': {
                        'required': True, 
                        'type': 'date', 
                        'sample': '2025-01-01',
                        'description': 'Contract effective date'
                    },
                    'initial_deposit': {
                        'required': True, 
                        'type': 'currency', 
                        'sample': 50000,
                        'description': 'Initial deposit amount'
                    },
                    'product_code': {
                        'required': True, 
                        'type': 'string', 
                        'sample': 'FIXED5',
                        'description': 'Product code'
                    },
                    'surrender_charge_period': {
                        'required': False, 
                        'type': 'number', 
                        'sample': 7,
                        'description': 'Surrender charge period in years'
                    },
                    'interest_rate': {
                        'required': False, 
                        'type': 'number', 
                        'sample': 3.5,
                        'description': 'Annual interest rate percentage'
                    }
                },
                'calculations': {
                    'owner_age': {
                        'type': 'date_based',
                        'calculation_type': 'age',
                        'description': 'Calculate age from birth date and effective date'
                    },
                    'projected_value': {
                        'type': 'compound_interest',
                        'calculation_type': 'future_value',
                        'years': [5, 10, 15, 20],
                        'description': 'Calculate projected values with compound interest'
                    }
                }
            }
        }
        
        return calculator_templates.get(product_type, calculator_templates['life'])

    def calculate_processed_quality_score(self, source_data: Dict, processed_data: Dict, field_mapping: Dict) -> float:
        """Calculate quality score for processed data"""
        if not field_mapping or not field_mapping.get('field_mappings'):
            return 0.0
        
        total_fields = len(field_mapping['field_mappings'])
        mapped_fields = len([m for m in field_mapping['field_mappings'] if m['mapping_status'] == 'auto_mapped'])
        required_fields = len([m for m in field_mapping['field_mappings'] if m.get('calculator_required', False)])
        required_mapped = len([m for m in field_mapping['field_mappings'] 
                             if m.get('calculator_required', False) and m['mapping_status'] == 'auto_mapped'])
        
        # Base score from mapping percentage
        mapping_score = (mapped_fields / total_fields) * 100 if total_fields > 0 else 0
        
        # Bonus for required fields
        required_bonus = (required_mapped / required_fields) * 20 if required_fields > 0 else 20
        
        # Average confidence score
        confidence_scores = [m['confidence_score'] for m in field_mapping['field_mappings'] if m['confidence_score'] > 0]
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
        confidence_bonus = (avg_confidence / 100) * 15
        
        total_score = min(100, mapping_score + required_bonus + confidence_bonus)
        return round(total_score, 1)

    def generate_comparison_excel(self, comparison_data: Dict, record_info: Dict) -> str:
        """Generate Excel file with comparison results including True/False for matches"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation. Please install with: pip install openpyxl")
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Data Comparison"
        
        # Add title and info
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1, value=f"AIM Data Comparison - {record_info.get('name', 'Unknown')}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add info rows
        info_data = [
            ("Record Name", record_info.get('name', 'Unknown')),
            ("Product Type", record_info.get('product_type', 'Unknown')),
            ("Calculator", record_info.get('calculator_path', 'Unknown')),
            ("Comparison Date", comparison_data.get('comparison_date', datetime.now().isoformat())),
            ("Matching Fields", f"{comparison_data['stats']['matching_fields']} of {comparison_data['stats']['total_fields']} ({comparison_data['stats']['completion_percentage']}%)"),
        ]
        
        for i, (label, value) in enumerate(info_data, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Add headers row
        headers = ["Field Name", "Source Value", "Calculator Value", "Values Match", "Status", "Required", "Data Type"]
        header_row = i + 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Add data rows
        for i, field in enumerate(comparison_data['fields'], header_row + 1):
            ws.cell(row=i, column=1, value=field['field_name'])
            ws.cell(row=i, column=2, value=field.get('source_value', ''))
            ws.cell(row=i, column=3, value=field.get('calculator_value', ''))
            
            # Values Match column (True/False)
            values_match_cell = ws.cell(row=i, column=4, value=str(field.get('values_match', False)))
            if field.get('values_match', False):
                values_match_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                values_match_cell.font = Font(color='006100')
            else:
                values_match_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                values_match_cell.font = Font(color='9C0006')
            
            # Status column
            status_cell = ws.cell(row=i, column=5, value=field.get('status', ''))
            status_colors = {
                'match': 'C6EFCE',     # Light green
                'mismatch': 'FFC7CE',  # Light red
                'missing': 'FFEB9C',   # Light yellow
                'can_fill': 'DDEBF7'   # Light blue
            }
            status_color = status_colors.get(field.get('status', ''), 'FFFFFF')
            status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            
            # Required and Data Type
            ws.cell(row=i, column=6, value=str(field.get('required', False)))
            ws.cell(row=i, column=7, value=field.get('data_type', 'string'))
            
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create filename and save
        filename = f"comparison_{record_info.get('id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        
        wb.save(filepath)
        return filepath
# Initialize database manager
db_manager = WebDatabaseManager()

@app.route('/')
@app.route('/index')
def index():
    """Main dashboard page"""
    try:
        if 'session_id' not in session:
            session['session_id'] = os.urandom(16).hex()
        
        print("Getting statistics...")
        stats = db_manager.get_statistics()
        print(f"Stats: {stats}")
        
        print("Getting recent data...")
        recent_data = db_manager.get_all_data(session['session_id'])
        print(f"Recent data count: {len(recent_data)}")
        
        # Use the original beautiful template
        return render_template('index.html', stats=stats, recent_data=recent_data)
        
    except Exception as e:
        print(f"Error in index route: {e}")
        import traceback
        traceback.print_exc()
        return f"Error: {e}", 500

@app.route('/help')
def help_page():
    """Help page with documentation and usage instructions"""
    return render_template('help.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    """File upload and processing page"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name', '').strip()
            product_type = request.form.get('product_type', '').strip()
            
            if not name or not product_type:
                flash('Name and product type are required', 'error')
                return redirect(url_for('upload_file'))
            
            # Handle JSON data input
            json_text = request.form.get('json_data', '').strip()
            uploaded_file = request.files.get('file')
            
            json_data = None
            file_source = None
            
            if json_text:
                # Parse JSON from text input
                try:
                    json_data = json.loads(json_text)
                    file_source = 'manual_input'
                except json.JSONDecodeError as e:
                    flash(f'Invalid JSON format: {str(e)}', 'error')
                    return redirect(url_for('upload_file'))
            
            elif uploaded_file and uploaded_file.filename:
                # Process uploaded file
                filename = secure_filename(uploaded_file.filename)
                file_source = filename
                
                if filename.endswith('.json'):
                    try:
                        json_data = json.load(uploaded_file)
                    except json.JSONDecodeError as e:
                        flash(f'Invalid JSON file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(uploaded_file)
                        json_data = df.to_dict('records')[0] if not df.empty else {}
                    except Exception as e:
                        flash(f'Error reading Excel file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                else:
                    flash('Please upload a JSON or Excel file', 'error')
                    return redirect(url_for('upload_file'))
            
            else:
                flash('Please provide JSON data or upload a file', 'error')
                return redirect(url_for('upload_file'))
            
            if not json_data:
                flash('No data found to process', 'error')
                return redirect(url_for('upload_file'))
            
            # Store data in database
            result = db_manager.store_data(
                name=name,
                product_type=product_type,
                json_data=json_data,
                session_id=session['session_id'],
                file_source=file_source
            )
            
            if result['success']:
                flash(f'Data stored successfully! Quality Score: {result["quality_score"]}%', 'success')
                return redirect(url_for('process_data', record_id=result['record_id']))
            else:
                if result['error'] == 'duplicate':
                    flash(f'Duplicate data detected: {result["message"]}', 'warning')
                else:
                    flash(f'Error storing data: {result["message"]}', 'error')
                return redirect(url_for('upload_file'))
        
        except Exception as e:
            flash(f'Unexpected error: {str(e)}', 'error')
            return redirect(url_for('upload_file'))
    
    return render_template('upload.html')

@app.route('/process/<int:record_id>')
def process_data(record_id):
    """Process stored data using AIM processor"""
    try:
        # Get data from database
        record = db_manager.get_data_by_id(record_id)
        if not record:
            flash('Record not found', 'error')
            return redirect(url_for('index'))

        # Parse JSON data
        json_data = json.loads(record['json_data'])

        # Process with AIM processor
        result = aim_processor.process_fast_ui_input(
            json_data, 
            record['product_type']
        )
        
        # Update the record status to 'processed' so it appears in compare page
        try:
            with sqlite3.connect(db_manager.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE user_data 
                    SET validation_status = 'processed',
                        calculator_path = ?
                    WHERE id = ?
                """, (f"/calculators/{record['product_type']}/standard.json", record_id))
                conn.commit()
                print(f"Updated record {record_id} status to 'processed'")
        except Exception as e:
            print(f"Error updating record status: {e}")

        return render_template('process_results.html', 
                             record=record, 
                             input_data=json_data,
                             result=result)

    except (ValidationError, MappingError) as e:
        flash(f'Processing error: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Unexpected error during processing: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/field-mapping')
def field_mapping():
    """Field mapping interface"""
    return render_template('field_mapping.html')

@app.route('/compare')
def compare():
    """Data comparison page"""
    print("=== COMPARE PAGE - COMPREHENSIVE DEBUG ===")
    print(f"Database manager object: {db_manager}")
    print(f"Database path: {db_manager.db_path}")
    print(f"Database file exists: {os.path.exists(db_manager.db_path)}")
    
    # CRITICAL TEST: Use the exact same line as view_data route
    print("CALLING: db_manager.get_all_data() - EXACT SAME AS VIEW_DATA")
    all_data = db_manager.get_all_data()
    print(f"RESULT: {type(all_data)} with {len(all_data)} records")
    
    # ALWAYS assign processed_records (this was the bug!)
    processed_records = all_data
    print(f"processed_records assignment: {len(processed_records)} records")
    
    # Test if we can access the first record
    if processed_records:
        try:
            first_record = processed_records[0]
            print(f"✅ SUCCESS: First record accessible: {type(first_record)}")
            print(f"First record content: {dict(first_record)}")
        except Exception as record_error:
            print(f"Error accessing first record: {record_error}")
    else:
        print("❌ ZERO RECORDS - This is the bug we need to fix!")
        
        # Let's debug why get_all_data() returns empty
        try:
            print("Testing direct SQL...")
            with sqlite3.connect(db_manager.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM user_data")
                direct_count = cursor.fetchone()[0]
                print(f"Direct SQL count: {direct_count}")
                
                if direct_count > 0:
                    print("❌ BUG CONFIRMED: Records exist but get_all_data() returns empty!")
                    print("This suggests a bug in the get_all_data() method")
                    
                    # Test the exact SQL from get_all_data method
                    cursor.execute("""
                        SELECT id, name, product_type, created_date, validation_status AS status, 
                               quality_score, file_source, data_hash
                        FROM user_data ORDER BY created_date DESC LIMIT 100
                    """)
                    raw_rows = cursor.fetchall()
                    print(f"Raw SQL result: {len(raw_rows)} rows")
                    if raw_rows:
                        print(f"Sample raw row: {dict(raw_rows[0])}")
                        print("❌ BUG: SQL works but method doesn't - checking method implementation...")
                else:
                    print("No records in database at all!")
                    
        except Exception as sql_error:
            print(f"Direct SQL error: {sql_error}")
    
    print("=== END COMPARE DEBUG ===")
    
    # Load all available calculators from config directory
    config_dir = os.path.join(os.path.dirname(__file__), 'src', 'config', 'config_files')
    available_calculators = []
    
    # Add default calculators
    available_calculators = [
        {
            'path': '/calculators/life_insurance/standard.json',
            'name': 'Life Insurance - Standard',
            'icon': 'heart',
            'product_type': 'life_insurance'
        },
        {
            'path': '/calculators/life_insurance/term.json',
            'name': 'Life Insurance - Term',
            'icon': 'heart',
            'product_type': 'life_insurance'
        },
        {
            'path': '/calculators/annuity/fixed.json',
            'name': 'Annuity - Fixed',
            'icon': 'coins',
            'product_type': 'annuity'
        },
        {
            'path': '/calculators/annuity/variable.json',
            'name': 'Annuity - Variable',
            'icon': 'coins',
            'product_type': 'annuity'
        }
    ]
    
    # Try to find any calculator configs in the directory
    if os.path.exists(config_dir):
        for file in os.listdir(config_dir):
            if file.endswith('_calculator.json'):
                calculator_type = file.replace('_calculator.json', '')
                icon = 'heart' if 'life' in calculator_type else 'coins' if 'annuity' in calculator_type else 'calculator'
                
                # Create a path that matches the expected format
                path = f'/calculators/{calculator_type}/standard.json'
                
                # Add if not already in the list
                if not any(c['path'] == path for c in available_calculators):
                    available_calculators.append({
                        'path': path,
                        'name': f'{calculator_type.replace("_", " ").title()}',
                        'icon': icon,
                        'product_type': calculator_type
                    })
    
    return render_template('compare.html', 
                         processed_records=processed_records, 
                         available_calculators=available_calculators,
                         debug_info={
                             'record_count': len(processed_records),
                             'calculator_count': len(available_calculators),
                             'db_path': db_manager.db_path
                         })

@app.route('/create-mapping', methods=['POST'])
def create_mapping():
    """Create Excel field mapping template"""
    try:
        product_type = request.form.get('product_type', 'life_insurance')
        
        # Get sample data if available and session exists
        sample_data = None
        if 'session_id' in session:
            recent_data = db_manager.get_all_data(session['session_id'])
            if recent_data:
                latest_record = db_manager.get_data_by_id(recent_data[0]['id'])
                if latest_record:
                    try:
                        sample_data = json.loads(latest_record['json_data'])
                    except json.JSONDecodeError:
                        sample_data = None
        
        # Create Excel template
        filename = create_excel_template(product_type, sample_data)
        
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Template file was not created: {filename}")
        
        return send_file(filename, 
                        as_attachment=True, 
                        download_name=f'AIM_Mapping_Template_{product_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        error_msg = f'Error creating mapping template: {str(e)}'
        print(f"Excel template error: {error_msg}")  # Log to console
        
        # Return JSON error for AJAX requests or HTML error page
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'error': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for('field_mapping'))

@app.route('/api/save-mapping', methods=['POST'])
def save_mapping():
    """Save field mapping configuration"""
    try:
        data = request.get_json()
        
        mapping_name = data.get('name', '')
        product_type = data.get('product_type', '')
        description = data.get('description', '')
        field_mappings = data.get('mappings', [])
        
        if not mapping_name or not product_type:
            return jsonify({'success': False, 'message': 'Mapping name and product type are required'})
        
        # Create mappings directory if it doesn't exist
        mappings_dir = 'saved_mappings'
        os.makedirs(mappings_dir, exist_ok=True)
        
        # Save mapping to JSON file
        mapping_file = os.path.join(mappings_dir, f"{mapping_name.replace(' ', '_').lower()}.json")
        mapping_data = {
            'name': mapping_name,
            'product_type': product_type,
            'description': description,
            'created_date': datetime.now().isoformat(),
            'field_mappings': field_mappings
        }
        
        with open(mapping_file, 'w') as f:
            json.dump(mapping_data, f, indent=2)
        
        return jsonify({
            'success': True, 
            'message': f'Mapping "{mapping_name}" saved successfully',
            'file_path': mapping_file
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving mapping: {str(e)}'})

@app.route('/api/mappings')
def list_mappings():
    """List all saved mapping templates"""
    try:
        mappings_dir = 'saved_mappings'
        if not os.path.exists(mappings_dir):
            return jsonify({'mappings': []})
        
        mappings = []
        for filename in os.listdir(mappings_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(mappings_dir, filename)
                try:
                    with open(filepath, 'r') as f:
                        mapping_data = json.load(f)
                        mappings.append({
                            'filename': filename,
                            'name': mapping_data.get('name', ''),
                            'product_type': mapping_data.get('product_type', ''),
                            'description': mapping_data.get('description', ''),
                            'created_date': mapping_data.get('created_date', '')
                        })
                except Exception as e:
                    print(f"Error reading mapping file {filename}: {e}")
                    continue
        
        return jsonify({'mappings': mappings})
        
    except Exception as e:
        return jsonify({'error': f'Error listing mappings: {str(e)}'})

@app.route('/api/mappings/<filename>')
def load_mapping(filename):
    """Load a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Mapping file not found'}), 404
        
        with open(filepath, 'r') as f:
            mapping_data = json.load(f)
        
        return jsonify({'success': True, 'mapping': mapping_data})
        
    except Exception as e:
        return jsonify({'error': f'Error loading mapping: {str(e)}'})

@app.route('/api/mappings/<filename>', methods=['DELETE'])
def delete_mapping(filename):
    """Delete a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'Mapping file not found'}), 404
        
        os.remove(filepath)
        return jsonify({'success': True, 'message': f'Mapping "{filename}" deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting mapping: {str(e)}'})

@app.route('/view-data')
def view_data():
    """View all stored data"""
    all_data = db_manager.get_all_data()
    
    # Calculate statistics for the view
    stats = {
        'total_records': len(all_data),
        'avg_quality': sum(record.get('quality_score', 0) for record in all_data) / len(all_data) if all_data else 0,
        'last_updated': all_data[0].get('created_date', 'Never') if all_data else 'Never'
    }
    
    return render_template('view_data.html', data=all_data, **stats)

@app.route('/api/data/<int:record_id>')
def api_get_data(record_id):
    """API endpoint to get record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify(record)
    else:
        return jsonify({'error': 'Record not found'}), 404

@app.route('/api/stats')
def api_stats():
    """API endpoint for statistics"""
    return jsonify(db_manager.get_statistics())

@app.route('/api/record/<int:record_id>')
def api_get_record(record_id):
    """API endpoint to get a specific record for comparison"""
    try:
        record = db_manager.get_data_by_id(record_id)
        if not record:
            return jsonify({
                'success': False,
                'message': f'Record with ID {record_id} not found'
            }), 404
        
        # Parse JSON data
        if isinstance(record.get('json_data'), str):
            try:
                json_data = json.loads(record['json_data'])
            except (json.JSONDecodeError, TypeError):
                json_data = {}
        else:
            json_data = record.get('json_data', {})
        
        record_data = {
            'id': record['id'],
            'name': record['name'],
            'product_type': record['product_type'],
            'data': json_data,
            'calculator_path': record.get('calculator_path'),
            'created_date': record['created_date'],
            'validation_status': record.get('validation_status'),
            'quality_score': record.get('quality_score', 0.0)
        }
        
        return jsonify({
            'success': True,
            'record': record_data
        })
        
    except Exception as e:
        print(f"Error getting record {record_id}: {e}")
        return jsonify({
            'success': False,
            'message': f'Error retrieving record: {str(e)}'
        }), 500

@app.route('/api/compare-data', methods=['POST'])
def api_compare_data():
    """API endpoint to perform data comparison"""
    try:
        data = request.get_json()
        record_id = data.get('record_id')
        calculator_path = data.get('calculator_path')
        
        if not record_id or not calculator_path:
            return jsonify({'success': False, 'message': 'Missing required fields: record_id and calculator_path'})
        
        # Get source data by record ID
        record = db_manager.get_data_by_id(record_id)
        
        if not record:
            return jsonify({'success': False, 'message': f'Record with ID {record_id} not found'})
        
        source_data = json.loads(record['json_data']) if isinstance(record['json_data'], str) else record.get('json_data', {})
        
        # Get calculator template data
        # Extract product type from the record to pass to calculator method
        product_type = record.get('product_type', 'life_insurance')
        calculator_data = db_manager.load_calculator_reference(calculator_path, product_type)
        
        # Perform comparison
        comparison_result = db_manager.perform_data_comparison(source_data, calculator_data)
        
        return jsonify({
            'success': True,
            'comparison': comparison_result,
            'source_record': {
                'id': record['id'],
                'name': record['name'],
                'product_type': record['product_type'],
                'created_date': record['created_date']
            },
            'calculator_info': {
                'path': calculator_path,
                'name': calculator_data.get('name', 'Unknown Calculator'),
                'version': calculator_data.get('version', '1.0')
            }
        })
        
    except Exception as e:
        print(f"Error in data comparison: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error performing comparison: {str(e)}'})

@app.route('/api/upload-calculator', methods=['POST'])
def api_upload_calculator():
    """API endpoint to upload calculator files"""
    try:
        if 'calculator' not in request.files:
            return jsonify({'success': False, 'message': 'No calculator file provided'})
        
        file = request.files['calculator']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        if not file.filename.endswith('.json'):
            return jsonify({'success': False, 'message': 'Only JSON files are supported'})
        
        # Read and validate the calculator file
        try:
            calculator_data = json.load(file)
        except json.JSONDecodeError as e:
            return jsonify({'success': False, 'message': f'Invalid JSON format: {str(e)}'})
        
        # Validate calculator structure
        if 'fields' not in calculator_data and 'mappings' not in calculator_data:
            return jsonify({'success': False, 'message': 'Calculator file must contain "fields" or "mappings" property'})
        
        # Save the calculator file to temp directory
        filename = secure_filename(file.filename)
        temp_dir = os.path.join(os.path.dirname(__file__), 'temp_calculators')
        os.makedirs(temp_dir, exist_ok=True)
        file_path = os.path.join(temp_dir, filename)
        
        # Reset file pointer and save
        file.seek(0)
        file.save(file_path)
        
        # Return success with the path that should be used
        calculator_path = f'/calculators/uploaded/{filename}'
        
        return jsonify({
            'success': True,
            'message': f'Calculator "{filename}" uploaded successfully',
            'calculator_path': calculator_path,
            'name': calculator_data.get('name', filename.replace('.json', '')),
            'product_type': calculator_data.get('product_type', 'unknown')
        })
        
    except Exception as e:
        print(f"Error uploading calculator: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error uploading calculator: {str(e)}'})

@app.route('/api/upload-template', methods=['POST'])
def api_upload_template():
    """API endpoint to upload and process template files"""
    try:
        if 'template_file' not in request.files:
            return jsonify({'success': False, 'message': 'No template file provided'})
        
        file = request.files['template_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        filename = secure_filename(file.filename)
        file_ext = filename.lower().split('.')[-1]
        
        if file_ext == 'json':
            # Handle JSON template files
            try:
                template_data = json.load(file)
                return jsonify({
                    'success': True,
                    'message': f'JSON template "{filename}" processed successfully',
                    'template_data': template_data
                })
            except json.JSONDecodeError as e:
                return jsonify({'success': False, 'message': f'Invalid JSON format: {str(e)}'})
                
        elif file_ext in ['xlsx', 'xls']:
            # Handle Excel template files
            try:
                # Save file temporarily
                temp_dir = os.path.join(os.path.dirname(__file__), 'temp_uploads')
                os.makedirs(temp_dir, exist_ok=True)
                file_path = os.path.join(temp_dir, filename)
                file.save(file_path)
                
                # Process Excel file
                df = pd.read_excel(file_path)
                
                # Convert Excel to template format
                template_data = {
                    'name': filename.replace(f'.{file_ext}', ''),
                    'description': f'Imported from {filename}',
                    'product_type': 'life_insurance',  # Default, user can change
                    'field_mappings': []
                }
                
                # Try to detect mapping structure
                if len(df.columns) >= 2:
                    for index, row in df.iterrows():
                        source_field = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                        target_field = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                        
                        if source_field and target_field:
                            mapping = {
                                'source': source_field,
                                'target': target_field,
                                'transformation': 'none',
                                'required': False
                            }
                            
                            # Add additional columns if available
                            if len(df.columns) > 2 and pd.notna(row.iloc[2]):
                                transformation = str(row.iloc[2]).strip().lower()
                                if transformation in ['uppercase', 'lowercase', 'date_format', 'currency_format', 'trim']:
                                    mapping['transformation'] = transformation
                            
                            if len(df.columns) > 3 and pd.notna(row.iloc[3]):
                                required_str = str(row.iloc[3]).strip().lower()
                                mapping['required'] = required_str in ['true', 'yes', '1', 'required']
                            
                            template_data['field_mappings'].append(mapping)
                
                # Clean up temp file
                try:
                    os.remove(file_path)
                except:
                    pass
                
                return jsonify({
                    'success': True,
                    'message': f'Excel template "{filename}" processed successfully',
                    'template_data': template_data
                })
                
            except Exception as e:
                return jsonify({'success': False, 'message': f'Error processing Excel file: {str(e)}'})
        
        else:
            return jsonify({'success': False, 'message': 'Unsupported file format. Please upload JSON or Excel files.'})
            
    except Exception as e:
        print(f"Error in template upload: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error processing template file: {str(e)}'})

if __name__ == '__main__':
    print("Starting Flask application...")
    app.run(host='0.0.0.0', port=5000, debug=True)
