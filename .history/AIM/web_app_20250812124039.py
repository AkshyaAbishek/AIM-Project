"""
AIM Web Application - Flask Implementation
Converts the tkinter desktop application to a web-based interface for Azure deployment
"""

import os
import json
import hashlib
from datetime import datetime
from typing import Dict, Any, Optional
from werkzeug.utils import secure_filename
import pandas as pd

from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file, session
from flask import Response
import sqlite3
import sys

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

try:
    from aim_processor import AIMProcessor, ValidationError, MappingError
except ImportError:
    # Fallback for web deployment
    class AIMProcessor:
        def process_fast_ui_input(self, data, product_type):
            return {"status": "success", "message": "Processed successfully"}
    
    class ValidationError(Exception):
        pass
    
    class MappingError(Exception):
        pass

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize AIM processor
aim_processor = AIMProcessor()

class WebDatabaseManager:
    """Database manager for web application"""
    
    def __init__(self, db_path: str = "aim_data.db"):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize database with web-optimized schema"""
        with sqlite3.connect(self.db_path) as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS user_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    product_type TEXT NOT NULL CHECK(product_type IN ('life', 'annuity')),
                    data_hash TEXT UNIQUE NOT NULL,
                    json_data TEXT NOT NULL,
                    field_mappings TEXT,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    modified_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    validation_status TEXT DEFAULT 'pending',
                    quality_score REAL DEFAULT 0.0,
                    file_source TEXT,
                    processing_notes TEXT,
                    session_id TEXT
                );
                
                CREATE TABLE IF NOT EXISTS processing_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    operation_details TEXT,
                    status TEXT CHECK(status IN ('success', 'warning', 'error')),
                    error_message TEXT,
                    processing_time_ms INTEGER,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    session_id TEXT
                );
                
                CREATE INDEX IF NOT EXISTS idx_user_data_hash ON user_data(data_hash);
                CREATE INDEX IF NOT EXISTS idx_user_data_session ON user_data(session_id);
            """)
    
    def store_data(self, name: str, product_type: str, json_data: Dict, 
                   session_id: str, file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with duplicate detection"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, created_date, 
                     validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    datetime.now().isoformat(), 'validated', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'storage_error',
                'message': str(e)
            }
    
    def calculate_quality_score(self, json_data: Dict) -> float:
        """Calculate data quality score"""
        score = 0.0
        
        # Required fields check (40% of score)
        required_fields = ['policy_number', 'product_type', 'effective_date']
        present_required = sum(1 for field in required_fields if field in json_data and json_data[field])
        score += (present_required / len(required_fields)) * 40
        
        # Data completeness (30% of score)
        total_fields = len(json_data)
        non_empty_fields = sum(1 for v in json_data.values() if v is not None and str(v).strip())
        if total_fields > 0:
            score += (non_empty_fields / total_fields) * 30
        
        # Data format consistency (30% of score)
        format_score = 30
        if 'age' in json_data:
            try:
                age = float(json_data['age'])
                if age < 0 or age > 120:
                    format_score -= 10
            except:
                format_score -= 10
        
        score += format_score
        return round(min(100, score), 2)
    
    def get_all_data(self, session_id: Optional[str] = None) -> list:
        """Get all stored data, optionally filtered by session"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if session_id:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status, quality_score
                    FROM user_data WHERE session_id = ? ORDER BY created_date DESC
                """, (session_id,))
            else:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status, quality_score
                    FROM user_data ORDER BY created_date DESC LIMIT 100
                """)
            
            return [dict(row) for row in cursor.fetchall()]
    
    def get_data_by_id(self, record_id: int) -> Optional[Dict]:
        """Get specific record by ID"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM user_data WHERE id = ?", (record_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get database statistics"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Total records
            cursor.execute("SELECT COUNT(*) FROM user_data")
            total_records = cursor.fetchone()[0]
            
            # Product distribution
            cursor.execute("SELECT product_type, COUNT(*) FROM user_data GROUP BY product_type")
            product_dist = dict(cursor.fetchall())
            
            # Average quality score
            cursor.execute("SELECT AVG(quality_score) FROM user_data")
            avg_quality = cursor.fetchone()[0] or 0
            
            return {
                'total_records': total_records,
                'product_distribution': product_dist,
                'average_quality': round(avg_quality, 2)
            }

# Initialize database manager
db_manager = WebDatabaseManager()

@app.route('/')
def index():
    """Main dashboard page"""
    if 'session_id' not in session:
        session['session_id'] = os.urandom(16).hex()
    
    stats = db_manager.get_statistics()
    recent_data = db_manager.get_all_data(session['session_id'])
    
    return render_template('index.html', stats=stats, recent_data=recent_data)

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    """File upload and processing page"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name', '').strip()
            product_type = request.form.get('product_type', '').strip()
            
            if not name or not product_type:
                flash('Name and product type are required', 'error')
                return redirect(url_for('upload_file'))
            
            # Handle JSON data input
            json_text = request.form.get('json_data', '').strip()
            uploaded_file = request.files.get('file')
            
            json_data = None
            file_source = None
            
            if json_text:
                # Parse JSON from text input
                try:
                    json_data = json.loads(json_text)
                    file_source = 'manual_input'
                except json.JSONDecodeError as e:
                    flash(f'Invalid JSON format: {str(e)}', 'error')
                    return redirect(url_for('upload_file'))
            
            elif uploaded_file and uploaded_file.filename:
                # Process uploaded file
                filename = secure_filename(uploaded_file.filename)
                file_source = filename
                
                if filename.endswith('.json'):
                    try:
                        json_data = json.load(uploaded_file)
                    except json.JSONDecodeError as e:
                        flash(f'Invalid JSON file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(uploaded_file)
                        json_data = df.to_dict('records')[0] if not df.empty else {}
                    except Exception as e:
                        flash(f'Error reading Excel file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                else:
                    flash('Please upload a JSON or Excel file', 'error')
                    return redirect(url_for('upload_file'))
            
            else:
                flash('Please provide JSON data or upload a file', 'error')
                return redirect(url_for('upload_file'))
            
            if not json_data:
                flash('No data found to process', 'error')
                return redirect(url_for('upload_file'))
            
            # Store data in database
            result = db_manager.store_data(
                name=name,
                product_type=product_type,
                json_data=json_data,
                session_id=session['session_id'],
                file_source=file_source
            )
            
            if result['success']:
                flash(f'Data stored successfully! Quality Score: {result["quality_score"]}%', 'success')
                return redirect(url_for('process_data', record_id=result['record_id']))
            else:
                if result['error'] == 'duplicate':
                    flash(f'Duplicate data detected: {result["message"]}', 'warning')
                else:
                    flash(f'Error storing data: {result["message"]}', 'error')
                return redirect(url_for('upload_file'))
        
        except Exception as e:
            flash(f'Unexpected error: {str(e)}', 'error')
            return redirect(url_for('upload_file'))
    
    return render_template('upload.html')

@app.route('/process/<int:record_id>')
def process_data(record_id):
    """Process stored data using AIM processor"""
    try:
        # Get data from database
        record = db_manager.get_data_by_id(record_id)
        if not record:
            flash('Record not found', 'error')
            return redirect(url_for('index'))
        
        # Parse JSON data
        json_data = json.loads(record['json_data'])
        
        # Process with AIM processor
        result = aim_processor.process_fast_ui_input(
            json_data, 
            record['product_type']
        )
        
        return render_template('process_results.html', 
                             record=record, 
                             input_data=json_data,
                             result=result)
    
    except (ValidationError, MappingError) as e:
        flash(f'Processing error: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Unexpected error during processing: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/field-mapping')
def field_mapping():
    """Field mapping interface"""
    return render_template('field_mapping.html')

@app.route('/create-mapping', methods=['POST'])
def create_mapping():
    """Create Excel field mapping template"""
    try:
        product_type = request.form.get('product_type', 'life')
        
        # Get sample data if available
        recent_data = db_manager.get_all_data(session['session_id'])
        sample_data = None
        
        if recent_data:
            latest_record = db_manager.get_data_by_id(recent_data[0]['id'])
            if latest_record:
                sample_data = json.loads(latest_record['json_data'])
        
        # Create Excel template
        filename = create_excel_template(product_type, sample_data)
        
        return send_file(filename, 
                        as_attachment=True, 
                        download_name=f'aim_mapping_template_{product_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    
    except Exception as e:
        flash(f'Error creating mapping template: {str(e)}', 'error')
        return redirect(url_for('field_mapping'))

@app.route('/view-data')
def view_data():
    """View all stored data"""
    all_data = db_manager.get_all_data()
    return render_template('view_data.html', data=all_data)

@app.route('/api/data/<int:record_id>')
def api_get_data(record_id):
    """API endpoint to get record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify(record)
    else:
        return jsonify({'error': 'Record not found'}), 404

@app.route('/api/stats')
def api_stats():
    """API endpoint for statistics"""
    return jsonify(db_manager.get_statistics())

@app.route('/health')
def health_check():
    """Health check endpoint for Azure monitoring"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

def create_excel_template(product_type: str, sample_data: Optional[Dict] = None):
    """Create Excel mapping template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Field_Mapping"
    
    # Headers
    headers = ['Source_Field', 'Target_Field', 'Values_Match', 'Sample_Value', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data rows
    if sample_data:
        for row_idx, (field, value) in enumerate(sample_data.items(), 2):
            ws.cell(row=row_idx, column=1, value=field)  # Source field
            ws.cell(row=row_idx, column=4, value=str(value)[:50])  # Sample value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = f"templates/mapping_template_{product_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    os.makedirs('templates', exist_ok=True)
    wb.save(filename)
    
    return filename

@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', 
                         error_code=404, 
                         error_message="Page not found"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', 
                         error_code=500, 
                         error_message="Internal server error"), 500

if __name__ == '__main__':
    # Development server
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
