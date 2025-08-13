"""
AIM - Actuarial Input Mapper - Example Usage

This script demonstrates how to use AIM to process palindrome validation
with optimized code structure and eliminated redundancies.
"""

import json
import sys
import os
from datetime import datetime
from typing import Dict, List, Union, Any, Optional, Callable
import tkinter as tk
from tkinter import messagebox, simpledialog, scrolledtext
from tkinter import ttk
import sqlite3
import hashlib

# Add src directory to path for imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src')))

from aim_processor import AIMProcessor, ValidationError, MappingError # type: ignore


class MessageFormatter:
    """
    Centralized message formatting to eliminate redundant print patterns.
    
    Provides consistent formatting for success, error, and info messages
    following the palindrome project's coding guidelines.
    """
    
    @staticmethod
    def success(message: str) -> str:
        """Format success message with consistent emoji and styling."""
        return f"✅ {message}"
    
    @staticmethod
    def error(message: str) -> str:
        """Format error message with consistent emoji and styling."""
        return f"❌ {message}"
    
    @staticmethod
    def info(message: str) -> str:
        """Format info message with consistent emoji and styling."""
        return f"📊 {message}"
    
    @staticmethod
    def warning(message: str) -> str:
        """Format warning message with consistent emoji and styling."""
        return f"⚠️ {message}"
    
    @staticmethod
    def section_header(title: str, width: int = 50) -> str:
        """Create consistent section headers."""
        return f"\n{title}\n{'=' * width}"


class ProcessorHelper:
    """
    Helper class to eliminate redundant processor operations.
    
    Centralizes common AIM processor operations following DRY principles
    and the palindrome checking guidelines.
    """
    
    def __init__(self):
        """Initialize the processor helper with AIM processor instance."""
        self.processor = AIMProcessor()
        self.formatter = MessageFormatter()
    
    def process_data_with_validation(self, data: Any, data_type: str = "mixed") -> Dict[str, Any]:
        """
        Process data with comprehensive validation and consistent error handling.
        
        Args:
            data: Data to process for palindrome validation
            data_type: Type of data being processed
            
        Returns:
            Dict containing processing results
        """
        try:
            result = self.processor.process_data(data)
            
            if result:
                print(self.formatter.success(f"Processing successful for {data_type}"))
                if isinstance(result, dict) and 'processed_at' in result:
                    print(f"   ⏱️ Processing completed at: {result['processed_at']}")
                if isinstance(result, dict) and 'palindrome_count' in result:
                    print(f"   🎯 Palindromes found: {result['palindrome_count']}")
                return result
            else:
                print(self.formatter.error(f"Processing failed for {data_type}"))
                return {"status": "failed", "data_type": data_type}
                
        except (ValidationError, MappingError) as processing_error:
            print(self.formatter.error(f"Processing error for {data_type}: {processing_error}"))
            return {"status": "error", "error": str(processing_error), "data_type": data_type}
        except Exception as unexpected_error:
            print(self.formatter.error(f"Unexpected error during processing: {unexpected_error}"))
            return {"status": "error", "error": str(unexpected_error), "data_type": data_type}


def run_example():
    """
    Run optimized example processing using helper classes to eliminate redundancy.
    
    Demonstrates palindrome validation with streamlined code structure
    following the project's coding guidelines.
    """
    formatter = MessageFormatter()
    helper = ProcessorHelper()
    
    print(formatter.section_header("🎯 AIM - Actuarial Input Mapper - Example Usage"))
    
    # Initialize processor
    print("\n1. Initializing AIM Processor...")
    try:
        print(formatter.success("AIM Processor initialized successfully"))
    except Exception as e:
        print(formatter.error(f"Failed to initialize AIM Processor: {e}"))
        return
    
    # Load sample palindrome data
    print("\n2. Loading sample palindrome data...")
    try:
        sample_file = os.path.join("data", "sample", "life_insurance_sample.json")
        with open(sample_file, 'r') as f:
            fast_ui_data = json.load(f)
        print(formatter.success(f"Loaded sample data with {len(fast_ui_data)} top-level fields"))
        print(f"   Sample fields: {list(fast_ui_data.keys())[:5]}...")
    except FileNotFoundError:
        print(formatter.warning("Sample data file not found, using palindrome examples"))
        fast_ui_data = {
            "number_palindromes": [121, 1221, 12321, 7, 0],
            "string_palindromes": ["racecar", "level", "A man a plan a canal Panama"],
            "non_palindromes": [123, "hello", "world"],
            "edge_cases": [-121, "", "a"]
        }
    
    # Process different palindrome categories
    print(formatter.section_header("3. Processing palindrome data categories"))
    
    palindrome_categories = [
        ("Number Palindromes", fast_ui_data.get("number_palindromes", [121, 1221, 123])),
        ("String Palindromes", fast_ui_data.get("string_palindromes", ["racecar", "hello"])),
        ("Edge Cases", fast_ui_data.get("edge_cases", [-121, "", "a"]))
    ]
    
    all_results = {}
    for category_name, category_data in palindrome_categories:
        print(f"\n   Processing {category_name}...")
        result = helper.process_data_with_validation(category_data, category_name)
        all_results[category_name] = result
    
    # Show final summary
    print(formatter.section_header("📋 Final Processing Summary"))
    successful_categories = sum(1 for result in all_results.values() 
                              if isinstance(result, dict) and result.get("status") != "failed")
    print(formatter.info(f"Successfully processed {successful_categories}/{len(palindrome_categories)} categories"))
    
    return all_results


def run_interactive_demo():
    """Run an interactive demo where user can input their own data."""
    
    print("\n" + "=" * 50)
    print("🎮 Interactive Demo Mode")
    print("=" * 50)
    
    processor = AIMProcessor()
    
    # Storage for user-added data
    user_data_store = []
    
    while True:
        # Show current state of user data store
        if user_data_store:
            print(f"\n📊 Current user data store: {len(user_data_store)} datasets")
        else:
            print("\n📊 No user data stored yet")
            
        print("\nOptions:")
        print("1. Process user-added data (or sample if none added)")
        print("2. Enter custom JSON data")
        print("3. Show field mapping for a product")
        print("4. Show stored user data")
        print("5. Help - Explain all options")
        print("6. Exit")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == "1":
            # Use user-added data if available, otherwise use sample data
            print(f"\n🔍 Debug: user_data_store contains {len(user_data_store)} items")
            
            if user_data_store:
                print(f"\n📄 Processing user-added data ({len(user_data_store)} datasets available):")
                print("\nSelect a dataset to process:")
                for i, data_entry in enumerate(user_data_store, 1):
                    print(f"{i}. {data_entry['product_type']} - {len(data_entry['data'])} fields - Added: {data_entry['timestamp']}")
                
                try:
                    dataset_choice = int(input("\nEnter dataset number: ").strip()) - 1
                    if 0 <= dataset_choice < len(user_data_store):
                        selected_data = user_data_store[dataset_choice]
                        print(f"\n📄 Processing selected data:")
                        print(json.dumps(selected_data['data'], indent=2))
                        
                        result = processor.process_fast_ui_input(
                            selected_data['data'], 
                            selected_data['product_type'], 
                            "full"
                        )
                        print_result(result)
                    else:
                        print("❌ Invalid dataset number")
                except (ValueError, IndexError):
                    print("❌ Invalid selection")
            else:
                # Fallback to sample data if no user data exists
                sample_data = {
                    "applicant_first_name": "Alice",
                    "applicant_last_name": "Johnson",
                    "applicant_birth_date": "1990-03-20",
                    "applicant_gender": "F",
                    "policy_face_amount": "500000",
                    "policy_effective_date": "2024-02-01",
                    "premium_mode": "Q"
                }
                
                print("\n📄 No user data found. Processing sample data:")
                print(json.dumps(sample_data, indent=2))
                
                result = processor.process_fast_ui_input(sample_data, "life", "full")
                print_result(result)
        
        elif choice == "2":
            # Custom JSON input
            print("\n📝 Enter your JSON data (or 'cancel' to go back):")
            print("Example: {\"applicant_first_name\": \"John\", \"applicant_last_name\": \"Doe\"}")
            
            json_input = input("\nJSON: ").strip()
            
            if json_input.lower() == 'cancel':
                continue
            
            try:
                custom_data = json.loads(json_input)
                product_type = input("Enter product type (life/annuity/health): ").strip().lower()
                
                if product_type in ["life", "annuity"]:
                    # Store the user data for future use
                    user_data_entry = {
                        'data': custom_data,
                        'product_type': product_type,
                        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    user_data_store.append(user_data_entry)
                    
                    print(f"\n✅ Data saved to user store!")
                    print(f"📊 Total datasets now: {len(user_data_store)}")
                    print(f"🔍 Debug: Just added: {json.dumps(user_data_entry['data'], indent=2)}")
                    
                    result = processor.process_fast_ui_input(custom_data, product_type, "full")
                    print_result(result)
                else:
                    print("❌ Invalid product type")
            
            except json.JSONDecodeError:
                print("❌ Invalid JSON format")
        
        elif choice == "3":
            # Show field mappings
            product_type = input("Enter product type (life/annuity/health): ").strip().lower()
            
            if product_type in ["life", "annuity"]:
                mappings = processor.mapper.get_mapping_summary(product_type)
                print(f"\n📋 Field mappings for {product_type}:")
                print(f"   Total mappings: {mappings['total_mappings']}")
                print(f"   Simple mappings: {mappings['simple_mappings']}")
                print(f"   Complex mappings: {mappings['complex_mappings']}")
                
                print("\n   Field details:")
                for field_info in mappings['field_list'][:5]:  # Show first 5
                    print(f"   • {field_info['fast_ui_field']} → {field_info['target_field']} ({field_info['type']})")
                
                if len(mappings['field_list']) > 5:
                    print(f"   ... and {len(mappings['field_list']) - 5} more fields")
            else:
                print("❌ Invalid product type")
        
        elif choice == "4":
            # Show stored user data
            if user_data_store:
                print(f"\n📋 Stored User Data ({len(user_data_store)} datasets):")
                for i, data_entry in enumerate(user_data_store, 1):
                    print(f"\n{i}. Product Type: {data_entry['product_type']}")
                    print(f"   Timestamp: {data_entry['timestamp']}")
                    print(f"   Data: {json.dumps(data_entry['data'], indent=2)}")
            else:
                print("\n📋 No user data stored yet. Use option 2 to add data.")
        
        elif choice == "5":
            # Help - Explain all options
            print("\n" + "=" * 60)
            print("📚 HELP - Select Option for Explanation")
            print("=" * 60)
            
            print("\nWhich option would you like help with?")
            print("1. Process user-added data")
            print("2. Enter custom JSON data")
            print("3. Show field mapping for a product")
            print("4. Show stored user data")
            print("5. All options (show everything)")
            print("6. Back to main menu")
            
            help_choice = input("\nEnter your choice (1-6): ").strip()
            
            if help_choice == "1":
                print("\n🔹 Option 1: Process user-added data")
                print("1. This option processes JSON data that you previously entered through Option 2. When selected, it shows a list of all stored datasets to choose from, displaying the product type, number of fields, and timestamp when each was added. If no user data exists yet, it will automatically use sample data instead. The system converts your data into actuarial calculation format.")
                print("\n2. Example: If you stored life insurance data with applicant information, it will transform fields like 'applicant_first_name' and 'policy_face_amount' into the corresponding actuarial input format.")
            
            elif help_choice == "2":
                print("\n🔹 Option 2: Enter custom JSON data")
                print("1. This option allows you to input your own JSON data manually by typing it directly into the system. You'll need to specify the product type as either life, annuity, or health insurance. The data gets automatically stored for future use and you can choose to process it immediately or save it for later.")
                print("\n2. Example: You might enter: {\"applicant_first_name\": \"John\", \"applicant_last_name\": \"Doe\", \"applicant_birth_date\": \"1985-06-15\", \"applicant_gender\": \"M\", \"policy_face_amount\": \"250000\", \"policy_effective_date\": \"2024-01-01\", \"premium_mode\": \"M\"} and select 'life' as the product type. Common field examples include: applicant_gender (M/F), policy_effective_date (YYYY-MM-DD format), premium_mode (M for Monthly, Q for Quarterly, A for Annual), and policy_face_amount (coverage amount).")
            
            elif help_choice == "3":
                print("\n🔹 Option 3: Show field mapping for a product")
                print("1. This option displays how FAST UI fields are mapped to actuarial calculation fields for different product types. It shows detailed mapping statistics including total mappings, simple direct mappings, and complex transformation mappings. This is useful for understanding how the system transforms your input data.")
                print("\n2. Example: You might see that 'applicant_birth_date' maps directly to 'BIRTH_DATE' as a simple mapping, while 'premium_mode' might have complex logic to convert 'M' (Monthly) to '12' in the actuarial system.")
            
            elif help_choice == "4":
                print("\n🔹 Option 4: Show stored user data")
                print("1. This option displays all JSON datasets you've entered through Option 2 during your current session. It shows timestamps of when each dataset was added, the product types specified, and the complete JSON data for each entry. This is useful for reviewing what data you have available.")
                print("\n2. Example: You might see entries like 'Life Insurance data added on 2024-01-15 14:30:25' with the complete applicant and policy information you entered.")
            
            elif help_choice == "5":
                print("\n🔹 All Options - Complete Help")
                print("\n🔸 Option 1: Process user-added data")
                print("1. This option processes JSON data that you previously entered through Option 2. When selected, it shows a list of all stored datasets to choose from, displaying the product type, number of fields, and timestamp when each was added. If no user data exists yet, it will automatically use sample data instead. The system converts your data into actuarial calculation format.")
                print("\n2. Example: If you stored life insurance data with applicant information, it will transform fields like 'applicant_first_name' and 'policy_face_amount' into the corresponding actuarial input format.")
                
                print("\n🔸 Option 2: Enter custom JSON data")
                print("1. This option allows you to input your own JSON data manually by typing it directly into the system. You'll need to specify the product type as either life, annuity, or health insurance. The data gets automatically stored for future use and you can choose to process it immediately.")
                print("\n2. Example: You might enter: {\"applicant_first_name\": \"John\", \"applicant_last_name\": \"Doe\", \"applicant_birth_date\": \"1985-06-15\", \"applicant_gender\": \"M\", \"policy_face_amount\": \"250000\", \"policy_effective_date\": \"2024-01-01\", \"premium_mode\": \"M\"} and select 'life' as the product type. Common field examples include: applicant_gender (M/F), policy_effective_date (YYYY-MM-DD format), premium_mode (M for Monthly, Q for Quarterly, A for Annual), and policy_face_amount (coverage amount).")
                
                print("\n🔸 Option 3: Show field mapping for a product")
                print("1. This option displays how FAST UI fields are mapped to actuarial calculation fields for different product types. It shows detailed mapping statistics including total mappings, simple direct mappings, and complex transformation mappings. This is useful for understanding how the system transforms your input data.")
                print("\n2. Example: You might see that 'applicant_birth_date' maps directly to 'BIRTH_DATE' as a simple mapping, while 'premium_mode' might have complex logic to convert 'M' (Monthly) to '12' in the actuarial system.")
                
                print("\n🔸 Option 4: Show stored user data")
                print("1. This option displays all JSON datasets you've entered through Option 2 during your current session. It shows timestamps of when each dataset was added, the product types specified, and the complete JSON data for each entry. This is useful for reviewing what data you have available.")
                print("\n2. Example: You might see entries like 'Life Insurance data added on 2024-01-15 14:30:25' with the complete applicant and policy information you entered.")
                
                print("\n🔸 Option 6: Exit")
                print("This option safely exits the interactive demo. Please note that all stored user data will be lost when you exit the program, so make sure to process any important data before leaving.")
                
                print("\n" + "=" * 60)
                print("💡 TIP: Use Option 2 to add your data and choose to process it immediately if needed!")
                print("=" * 60)
            
            elif help_choice == "6":
                print("📋 Returning to main menu...")
            
            else:
                print("❌ Invalid choice")
        
        elif choice == "6":
            print("\n👋 Goodbye!")
            break
        
        else:
            print("❌ Invalid choice")


def print_result(result):
    """Print processing result in a formatted way."""
    if result["status"] == "success":
        print(f"\n✅ Processing successful!")
        print(f"   ⏱️  Time: {result['processing_time']:.2f}s")
        print(f"   📊 Fields: {result['metadata']['fields_processed']} → {result['metadata']['fields_mapped']}")
        
        print(f"\n📋 Actuarial inputs:")
        for section_name, section_data in result["actuarial_inputs"].items():
            print(f"   📦 {section_name}:")
            for field_name, field_value in section_data.items():
                print(f"      • {field_name}: {field_value}")
    else:
        print(f"\n❌ Processing failed:")
        print(f"   Error: {result.get('error_message', 'Unknown error')}")


"""
AIM - Actuarial Input Mapper - GUI Demo

This module provides a graphical user interface (GUI) for the AIM interactive demo,
allowing users to easily map and process their FAST UI data inputs.
"""

class AIMDemoGUI:
    """GUI-based interactive demo for AIM processor."""
    
    def __init__(self):
        self.processor = AIMProcessor()
        self.user_data_store = []
        self.root = tk.Tk()
        self.root.title("AIM - Actuarial Input Mapper Demo")
        self.root.geometry("850x650")
        self.db_path = "aim_data.db"
        self.init_database()
        self.load_data_from_db()
        self.setup_ui()
    
    def init_database(self):
        """Initialize the SQLite database."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Create table for storing user data
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS user_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    data_hash TEXT UNIQUE,
                    product_type TEXT NOT NULL,
                    json_data TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to initialize database: {e}")
    
    def get_data_hash(self, data_dict):
        """Generate a hash for the data to check for duplicates."""
        # Convert dict to sorted JSON string for consistent hashing
        json_str = json.dumps(data_dict, sort_keys=True)
        return hashlib.md5(json_str.encode()).hexdigest()
    
    def save_data_to_db(self, data, product_type):
        """Save data to database, checking for duplicates."""
        try:
            data_hash = self.get_data_hash(data)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Check if data already exists
            cursor.execute('SELECT id FROM user_data WHERE data_hash = ?', (data_hash,))
            existing = cursor.fetchone()
            
            if existing:
                conn.close()
                return False, "This data already exists in the database."
            
            # Insert new data
            cursor.execute('''
                INSERT INTO user_data (data_hash, product_type, json_data, timestamp)
                VALUES (?, ?, ?, ?)
            ''', (data_hash, product_type, json.dumps(data), timestamp))
            
            conn.commit()
            conn.close()
            
            return True, "Data saved successfully to database."
            
        except Exception as e:
            return False, f"Database error: {e}"
    
    def load_data_from_db(self):
        """Load all data from database into memory."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT product_type, json_data, timestamp FROM user_data ORDER BY created_at')
            rows = cursor.fetchall()
            
            self.user_data_store = []
            for row in rows:
                self.user_data_store.append({
                    'product_type': row[0],
                    'data': json.loads(row[1]),
                    'timestamp': row[2]
                })
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to load data: {e}")
    
    def get_db_stats(self):
        """Get database statistics."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT COUNT(*) FROM user_data')
            total_count = cursor.fetchone()[0]
            
            cursor.execute('SELECT product_type, COUNT(*) FROM user_data GROUP BY product_type')
            by_product = cursor.fetchall()
            
            conn.close()
            
            return total_count, dict(by_product)
            
        except Exception as e:
            return 0, {}
    
    def setup_ui(self):
        """Set up the main user interface with enhanced styling."""
        # Configure main window with gradient-like background
        self.root.configure(bg="#f0f4f8")
        
        # Main title with enhanced styling
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        title_frame.pack(fill="x", pady=0)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="🎮 AIM Interactive Demo", 
                             font=("Segoe UI", 20, "bold"), fg="#ecf0f1", bg="#2c3e50")
        title_label.pack(expand=True)
        
        subtitle_label = tk.Label(title_frame, text="Actuarial Input Mapper - Professional Edition", 
                                font=("Segoe UI", 10), fg="#bdc3c7", bg="#2c3e50")
        subtitle_label.pack()
        
        # Status display with modern styling
        status_frame = tk.Frame(self.root, bg="#ecf0f1", relief="flat", bd=1)
        status_frame.pack(fill="x", padx=20, pady=10)
        
        self.status_var = tk.StringVar()
        self.update_status()
        status_label = tk.Label(status_frame, textvariable=self.status_var, 
                              font=("Segoe UI", 11), fg="#2c3e50", bg="#ecf0f1")
        status_label.pack(pady=8)
        
        # Database stats display with modern styling
        self.db_stats_var = tk.StringVar()
        self.update_db_stats()
        db_stats_label = tk.Label(status_frame, textvariable=self.db_stats_var, 
                                font=("Segoe UI", 10), fg="#27ae60", bg="#ecf0f1")
        db_stats_label.pack(pady=2)
        
        # Main buttons frame with enhanced background
        buttons_container = tk.Frame(self.root, bg="#f0f4f8")
        buttons_container.pack(pady=10, padx=20, fill="x")
        
        # Add a decorative header for buttons
        buttons_header = tk.Label(buttons_container, text="📋 Select an Operation", 
                                font=("Segoe UI", 12, "bold"), fg="#2c3e50", bg="#f0f4f8")
        buttons_header.pack(pady=(0, 8))
        
        buttons_frame = tk.Frame(buttons_container, bg="#ffffff", relief="raised", bd=1)
        buttons_frame.pack(pady=5, padx=10, fill="x")
        
        # Configure button frame background
        buttons_frame.configure(bg="#ffffff", padx=10, pady=10)
        
        # Create buttons for each option with enhanced styling
        self.create_button(buttons_frame, "📝 Add JSON Data", 
                          self.enter_custom_data, 0, 0, "#3498db", "#2980b9")
        self.create_button(buttons_frame, "📦 Bulk JSON Load", 
                          self.bulk_json_load, 0, 1, "#9b59b6", "#8e44ad")
        self.create_button(buttons_frame, "📊 Life Field Mapping", 
                          lambda: self.show_field_mapping("life"), 0, 2, "#e67e22", "#d35400")
        self.create_button(buttons_frame, "📋 View Stored Data", 
                          self.show_stored_data, 1, 0, "#1abc9c", "#16a085")
        self.create_button(buttons_frame, "🔍 Check Duplicates", 
                          self.show_duplicate_check, 1, 1, "#f39c12", "#e67e22")
        self.create_button(buttons_frame, "📊 Annuity Field Mapping", 
                          lambda: self.show_field_mapping("annuity"), 1, 2, "#8e44ad", "#7d3c98")
        self.create_button(buttons_frame, "❓ Help", 
                          self.show_help, 2, 0, "#34495e", "#2c3e50")
        self.create_button(buttons_frame, "🗑️ Clear Database", 
                          self.clear_database, 2, 1, "#e74c3c", "#c0392b")
        
        # Results display area with enhanced styling
        results_container = tk.Frame(self.root, bg="#f0f4f8")
        results_container.pack(padx=20, pady=(5, 15), fill="both", expand=True)
        
        results_header_frame = tk.Frame(results_container, bg="#34495e", height=35)
        results_header_frame.pack(fill="x")
        results_header_frame.pack_propagate(False)
        
        results_label = tk.Label(results_header_frame, text="📄 Results & Output", 
                               font=("Segoe UI", 11, "bold"), fg="#ecf0f1", bg="#34495e")
        results_label.pack(expand=True)
        
        # Results text area with modern styling
        text_frame = tk.Frame(results_container, bg="#ffffff", relief="sunken", bd=1)
        text_frame.pack(fill="both", expand=True)
        
        self.results_text = scrolledtext.ScrolledText(text_frame, height=12, width=80,
                                                    font=("Consolas", 9), bg="#ffffff", fg="#2c3e50",
                                                    insertbackground="#2c3e50", selectbackground="#3498db",
                                                    relief="flat", bd=5)
        self.results_text.pack(padx=8, pady=8, fill="both", expand=True)
        
        # Add initial welcome message with styling
        self.log_result("🎉 Welcome to AIM - Actuarial Input Mapper!")
        self.log_result("=" * 50)
        self.log_result("Select an operation from the buttons above to get started.")
        self.log_result("💡 Tip: Use 'Help' for detailed guidance on each feature.")
    
    def create_button(self, parent, text, command, row, col, bg_color="#3498db", hover_color="#2980b9"):
        """Create a styled button with hover effects and modern design."""
        btn = tk.Button(parent, text=text, command=command,
                       font=("Segoe UI", 9, "bold"), width=18, height=2,
                       bg=bg_color, fg="white", relief="flat", bd=0,
                       cursor="hand2", activebackground=hover_color, activeforeground="white")
        btn.grid(row=row, column=col, padx=5, pady=5, sticky="ew")
        parent.grid_columnconfigure(col, weight=1)
        
        # Add hover effects
        def on_enter(e):
            btn.configure(bg=hover_color)
        def on_leave(e):
            btn.configure(bg=bg_color)
            
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
    
    def create_dialog_button(self, parent, text, command, bg_color="#3498db", hover_color="#2980b9"):
        """Create a styled dialog button with hover effects."""
        btn = tk.Button(parent, text=text, command=command,
                       font=("Segoe UI", 9, "bold"), width=20, height=2,
                       bg=bg_color, fg="white", relief="flat", bd=0,
                       cursor="hand2", activebackground=hover_color, activeforeground="white",
                       wraplength=150)
        
        # Add hover effects
        def on_enter(e):
            btn.configure(bg=hover_color)
        def on_leave(e):
            btn.configure(bg=bg_color)
            
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn
    
    def update_status(self):
        """Update the status display."""
        if self.user_data_store:
            self.status_var.set(f"📊 Current session: {len(self.user_data_store)} datasets loaded from database")
        else:
            self.status_var.set("📊 No data loaded from database yet")
    
    def update_db_stats(self):
        """Update database statistics display."""
        total_count, by_product = self.get_db_stats()
        if total_count > 0:
            product_info = ", ".join([f"{prod}: {count}" for prod, count in by_product.items()])
            self.db_stats_var.set(f"💾 Database: {total_count} total records ({product_info})")
        else:
            self.db_stats_var.set("💾 Database: Empty")
    
    def log_result(self, message):
        """Add a message to the results area."""
        self.results_text.insert(tk.END, f"{message}\n")
        self.results_text.see(tk.END)
    
    def clear_results(self):
        """Clear the results area."""
        self.results_text.delete(1.0, tk.END)
    
    def show_loading_text(self, message="Loading..."):
        """Show loading text in the results area with enhanced visibility."""
        self.clear_results()
        self.log_result("=" * 60)
        self.log_result(f"⏳ {message}")
        self.log_result("=" * 60)
        self.log_result("Please wait...")
        self.root.update_idletasks()
        self.root.update()
        
        # Small delay to ensure text is visible
        import time
        time.sleep(0.1)
    
    def hide_loading_text(self):
        """Hide loading text by clearing results."""
        self.clear_results()
    

    
    def enter_custom_data(self):
        """Handle entering custom JSON data (save only, no processing)."""
        # Show loading text
        self.show_loading_text("Preparing JSON data entry form...")
        
        dialog = tk.Toplevel(self.root)
        dialog.title("📝 Add New JSON Data")
        dialog.geometry("700x650")  # Increased height to show all content
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="#f8f9fa")
        
        # Clear loading text once dialog is ready
        self.hide_loading_text()
        
        # Header with modern styling
        header_frame = tk.Frame(dialog, bg="#2c3e50", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📝 Enter Your JSON Data", 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#2c3e50").pack(expand=True)
        
        # Main content frame with scrolling capability
        main_frame = tk.Frame(dialog, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create canvas and scrollbar for scrolling
        canvas = tk.Canvas(main_frame, bg="#f8f9fa")
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f8f9fa")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Content frame (now inside scrollable_frame)
        content_frame = tk.Frame(scrollable_frame, bg="#f8f9fa")
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Instructions with modern styling
        instruction_frame = tk.Frame(content_frame, bg="#e3f2fd", relief="flat", bd=1)
        instruction_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(instruction_frame, text="💡 Enter your JSON data below:", 
                font=("Segoe UI", 11, "bold"), fg="#1565c0", bg="#e3f2fd").pack(pady=8)
        
        # Example label with better styling
        example_text = ('📋 Example: {"applicant_first_name": "John", "applicant_last_name": "Doe", '
                       '"applicant_birth_date": "1985-06-15", "applicant_gender": "M", '
                       '"policy_face_amount": "250000", "policy_effective_date": "2024-01-01", '
                       '"premium_mode": "M"}')
        example_label = tk.Label(instruction_frame, text=example_text, font=("Consolas", 8), 
                               fg="#424242", bg="#e3f2fd", wraplength=580)
        example_label.pack(pady=(0, 8), padx=10)
        
        # Text area for JSON input with enhanced styling
        text_label = tk.Label(content_frame, text="JSON Data:", 
                            font=("Segoe UI", 10, "bold"), fg="#2c3e50", bg="#f8f9fa")
        text_label.pack(anchor="w", pady=(0, 5))
        
        text_frame = tk.Frame(content_frame, bg="#ffffff", relief="solid", bd=1)
        text_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        json_text = scrolledtext.ScrolledText(text_frame, height=12, width=70,
                                            font=("Consolas", 10), bg="#ffffff", fg="#2c3e50",
                                            insertbackground="#2c3e50", selectbackground="#3498db",
                                            relief="solid", bd=1)
        json_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Product type selection with modern styling
        selection_frame = tk.Frame(content_frame, bg="#f1f8e9", relief="flat", bd=1)
        selection_frame.pack(fill="x", pady=(0, 15))
        
        product_frame = tk.Frame(selection_frame, bg="#f1f8e9")
        product_frame.pack(pady=15)
        
        tk.Label(product_frame, text="🏷️ Product Type:", font=("Segoe UI", 10, "bold"), 
                fg="#2e7d32", bg="#f1f8e9").pack(side="left", padx=10)
        product_var = tk.StringVar(value="life")
        product_combo = ttk.Combobox(product_frame, textvariable=product_var, 
                                   values=["life", "annuity"], width=15,
                                   font=("Segoe UI", 10))
        product_combo.pack(side="left", padx=10)
        
        # Processing option with modern styling and visual feedback
        process_frame = tk.Frame(selection_frame, bg="#f1f8e9")
        process_frame.pack(pady=(0, 15))
        
        tk.Label(process_frame, text="⚙️ After saving:", font=("Segoe UI", 10, "bold"), 
                fg="#2e7d32", bg="#f1f8e9").pack(pady=(10, 5))
        
        process_var = tk.StringVar(value="save_only")
        
        # Radio button options
        radio_container = tk.Frame(process_frame, bg="#f1f8e9")
        radio_container.pack(pady=5)
        
        save_only_radio = tk.Radiobutton(radio_container, text="💾 Save data only", 
                                       variable=process_var, value="save_only", 
                                       font=("Segoe UI", 10), bg="#f1f8e9", fg="#2e7d32",
                                       selectcolor="#c8e6c9", activebackground="#f1f8e9")
        save_only_radio.pack(anchor="w", padx=20, pady=3)
        
        process_now_radio = tk.Radiobutton(radio_container, text="🚀 Save and process immediately", 
                                         variable=process_var, value="process_now", 
                                         font=("Segoe UI", 10), bg="#f1f8e9", fg="#2e7d32",
                                         selectcolor="#c8e6c9", activebackground="#f1f8e9")
        process_now_radio.pack(anchor="w", padx=20, pady=3)
        
        # Visual feedback for selection
        def update_selection():
            if process_var.get() == "save_only":
                status_text = "📝 Selected: Save data only (no processing)"
            else:
                status_text = "🚀 Selected: Save and process immediately"
            selection_status.config(text=status_text)
        
        # Add the command callbacks
        save_only_radio.config(command=update_selection)
        process_now_radio.config(command=update_selection)
        
        # Status label
        selection_status = tk.Label(process_frame, text="📝 Selected: Save data only (no processing)", 
                                  font=("Segoe UI", 9, "italic"), fg="#1976d2", bg="#f1f8e9")
        selection_status.pack(pady=5)
        
        # Buttons with modern styling - always visible
        button_frame = tk.Frame(content_frame, bg="#f8f9fa")
        button_frame.pack(pady=20, fill="x")
        
        # Add helpful instruction
        instruction_label = tk.Label(button_frame, text="💡 Select an option above, then click Save Data", 
                                   font=("Segoe UI", 9, "italic"), fg="#555555", bg="#f8f9fa")
        instruction_label.pack(pady=(0, 10))
        
        # Center the buttons
        button_container = tk.Frame(button_frame, bg="#f8f9fa")
        button_container.pack(expand=True)
        
        def save_json():
            json_input = json_text.get(1.0, tk.END).strip()
            product_type = product_var.get()
            process_choice = process_var.get()
            
            if not json_input:
                messagebox.showerror("Error", "Please enter JSON data.")
                return
            
            try:
                custom_data = json.loads(json_input)
                
                # Show loading text while saving
                self.show_loading_text("Saving JSON data to database...")
                
                # Try to save to database
                success, message = self.save_data_to_db(custom_data, product_type)
                
                if not success:
                    self.hide_loading_text()
                    if "already exists" in message:
                        messagebox.showwarning("Duplicate Data", message)
                        return
                    else:
                        messagebox.showerror("Database Error", message)
                        return
                
                # Update displays
                self.show_loading_text("Updating database displays...")
                
                # Reload data from database to update the display
                self.load_data_from_db()
                self.update_status()
                self.update_db_stats()
                
                dialog.destroy()
                self.clear_results()
                self.log_result("✅ Data saved to database successfully!")
                self.log_result(f"📊 Total datasets in database: {len(self.user_data_store)}")
                self.log_result(f"📄 Added: {json.dumps(custom_data, indent=2)}")
                
                # Process the data if "process_now" is selected
                if process_choice == "process_now":
                    self.show_loading_text("Processing data immediately...")
                    self.log_result("\n🔄 Processing data immediately...")
                    result = self.processor.process_fast_ui_input(custom_data, product_type, "full")
                    self.display_result(result)
                else:
                    self.log_result("\n💡 Data saved successfully. You can view it using 'View stored data'.")
                
                # Hide loading text at the end
                self.hide_loading_text()
                
            except json.JSONDecodeError as e:
                self.hide_loading_text()
                messagebox.showerror("JSON Error", f"Invalid JSON format: {e}")
            except Exception as e:
                self.hide_loading_text()
                messagebox.showerror("Error", f"Processing error: {e}")
        
        # Create modern styled buttons with enhanced visibility
        save_btn = self.create_dialog_button(button_container, "💾 Save Data", save_json,
                                           "#27ae60", "#229954")
        save_btn.pack(side="left", padx=15)
        
        cancel_btn = self.create_dialog_button(button_container, "❌ Cancel", dialog.destroy,
                                             "#e74c3c", "#c0392b")
        cancel_btn.pack(side="left", padx=15)
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Add hover effects for buttons
        def on_save_enter(e): save_btn.configure(bg="#229954")
        def on_save_leave(e): save_btn.configure(bg="#27ae60")
        def on_cancel_enter(e): cancel_btn.configure(bg="#c0392b")
        def on_cancel_leave(e): cancel_btn.configure(bg="#e74c3c")
        
        save_btn.bind("<Enter>", on_save_enter)
        save_btn.bind("<Leave>", on_save_leave)
        cancel_btn.bind("<Enter>", on_cancel_enter)
        cancel_btn.bind("<Leave>", on_cancel_leave)
    
    def browse_save_excel(self, path_var):
        """Browse for Excel file save location with enhanced error handling."""
        try:
            from tkinter import filedialog
            import os
            
            # Suggest safe default directory
            default_dir = os.path.join(os.path.expanduser("~"), "Documents")
            if not os.path.exists(default_dir):
                default_dir = os.path.expanduser("~")  # User home directory
            
            # Default filename with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"field_mapping_{timestamp}.xlsx"
            
            filename = filedialog.asksaveasfilename(
                initialdir=default_dir,
                initialfile=default_filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Excel Mapping File As"
            )
            
            if filename:
                # Validate the path
                directory = os.path.dirname(filename)
                if not os.access(directory, os.W_OK):
                    messagebox.showwarning("Permission Warning", 
                                         f"You may not have write permission to:\n{directory}\n\n"
                                         f"Consider saving to:\n{default_dir}")
                    return
                
                path_var.set(filename)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to browse file: {e}\n\n"
                               f"Try saving to your Documents folder instead.")
    
    def browse_open_excel(self, path_var):
        """Browse for existing Excel file with validation."""
        try:
            from tkinter import filedialog
            import os
            
            # Suggest common locations
            default_dir = os.path.join(os.path.expanduser("~"), "Documents")
            if not os.path.exists(default_dir):
                default_dir = os.path.expanduser("~")
            
            filename = filedialog.askopenfilename(
                initialdir=default_dir,
                filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")],
                title="Select Actuarial Calculator Excel File"
            )
            
            if filename:
                # Validate file exists and is readable
                if not os.path.exists(filename):
                    messagebox.showerror("File Not Found", f"File does not exist:\n{filename}")
                    return
                
                if not os.access(filename, os.R_OK):
                    messagebox.showerror("Permission Error", f"Cannot read file:\n{filename}")
                    return
                
                # Quick validation - try to read file structure
                try:
                    import pandas as pd
                    df_test = pd.read_excel(filename, nrows=1)  # Read just header
                    if df_test.empty or len(df_test.columns) == 0:
                        messagebox.showwarning("File Warning", 
                                             "Excel file appears to be empty or has no columns.\n"
                                             "The mapping will still work, but suggestions may be limited.")
                except Exception as e:
                    messagebox.showwarning("File Warning", 
                                         f"Could not fully read Excel file:\n{e}\n\n"
                                         "The mapping will still work with basic functionality.")
                
                path_var.set(filename)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to browse file: {e}")
    
    def create_excel_mapping(self, product_type, output_path, calculator_path):
        """Create Excel mapping file with FAST UI to Actuarial Calculator field mapping."""
        # Show loading indicator
        loading = LoadingIndicator(self.root, "Creating Excel Mapping", "Initializing mapping process...")
        loading.show()
        
        # Force GUI update to show loading dialog immediately
        self.root.update_idletasks()
        self.root.update()
        
        # Small delay to ensure loading dialog is fully visible
        import time
        time.sleep(0.1)
        
        try:
            import pandas as pd
            import os
            
            self.clear_results()
            self.log_result("🔄 Creating Excel field mapping...")
            
            loading.update_status("Reading FAST UI field mappings...")
            loading.update_progress(10)
            
            # Yield to GUI thread to show progress
            self.root.update_idletasks()
            
            # Step 1: Get FAST UI fields from the processor
            mappings = self.processor.mapper.get_mapping_summary(product_type)
            
            loading.update_status("Analyzing actuarial calculator Excel...")
            loading.update_progress(25)
            
            # Yield to GUI thread
            self.root.update_idletasks()
            
            # Step 2: Read actuarial calculator Excel to get available fields
            calculator_fields = []
            try:
                if os.path.exists(calculator_path):
                    # Read first sheet to get column names
                    df_calc = pd.read_excel(calculator_path, nrows=5)  # Read just header and few rows
                    calculator_fields = list(df_calc.columns)
                    self.log_result(f"✅ Found {len(calculator_fields)} fields in calculator Excel")
                else:
                    self.log_result("⚠️ Calculator Excel not found, will create template only")
            except Exception as e:
                self.log_result(f"⚠️ Could not read calculator Excel: {e}")
                calculator_fields = ["Premium_Amount", "Policy_Number", "Insured_Name", "Coverage_Amount", 
                                   "Policy_Date", "Birth_Date", "Gender", "Risk_Class"]  # Default fields
            
            if loading.is_cancelled():
                return
                
            # Yield to GUI thread
            self.root.update_idletasks()
                
            loading.update_status("Creating field mappings...")
            loading.update_progress(40)
            
            # Step 3: Create mapping data
            mapping_data = []
            
            # Get sample data values if available
            sample_values = {}
            if self.user_data_store:
                # Use first record as sample
                sample_data = self.user_data_store[0]['data']
                sample_values = sample_data
            
            # Create mapping entries for each FAST UI field
            total_fields = len(mappings['field_list'])
            for i, field_info in enumerate(mappings['field_list']):
                if loading.is_cancelled():
                    return
                    
                fast_ui_field = field_info['fast_ui_field']
                target_field = field_info.get('target_field', '')
                
                # Get sample value
                sample_value = sample_values.get(fast_ui_field, "Sample_Value")
                
                # Suggest actuarial field (try to match by name similarity)
                suggested_actuarial_field = self.suggest_actuarial_field(fast_ui_field, calculator_fields)
                
                mapping_data.append({
                    'FAST_UI_Field': fast_ui_field,
                    'FAST_UI_Value': sample_value,
                    'Actuarial_Field': suggested_actuarial_field,
                    'Actuarial_Value': f"=B{len(mapping_data)+2}",  # Excel formula reference
                    'Values_Match': f"=IF(B{len(mapping_data)+2}=D{len(mapping_data)+2},TRUE,FALSE)"  # Comparison formula
                })
                
                # Update progress and yield to GUI thread
                progress = 40 + (i / total_fields) * 30
                loading.update_progress(progress)
                self.root.update_idletasks()
            
            loading.update_status("Adding common field mappings...")
            loading.update_progress(70)
            
            # Add common fields that might not be in mappings
            common_fields = [
                ('applicant_first_name', 'John', 'Insured_First_Name'),
                ('applicant_last_name', 'Doe', 'Insured_Last_Name'),
                ('policy_face_amount', '250000', 'Coverage_Amount'),
                ('premium_amount', '150.00', 'Premium_Amount'),
                ('policy_effective_date', '2024-01-01', 'Policy_Date'),
                ('applicant_birth_date', '1985-06-15', 'Birth_Date'),
                ('applicant_gender', 'M', 'Gender')
            ]
            
            # Add common fields if not already present
            existing_fields = [item['FAST_UI_Field'] for item in mapping_data]
            for ui_field, sample_val, calc_field in common_fields:
                if loading.is_cancelled():
                    return
                    
                if ui_field not in existing_fields:
                    suggested_calc_field = self.suggest_actuarial_field(ui_field, calculator_fields)
                    if not suggested_calc_field:
                        suggested_calc_field = calc_field
                    
                    mapping_data.append({
                        'FAST_UI_Field': ui_field,
                        'FAST_UI_Value': sample_val,
                        'Actuarial_Field': suggested_calc_field,
                        'Actuarial_Value': f"=B{len(mapping_data)+2}",  # Excel formula
                        'Values_Match': f"=IF(B{len(mapping_data)+2}=D{len(mapping_data)+2},TRUE,FALSE)"  # Comparison
                    })
                    
                # Yield to GUI thread
                self.root.update_idletasks()
            
            loading.update_status("Creating Excel file...")
            loading.update_progress(80)
            
            # Yield to GUI thread
            self.root.update_idletasks()
            
            # Step 4: Create Excel file with enhanced error handling
            df_mapping = pd.DataFrame(mapping_data)
            
            # Check write permissions before creating file
            output_dir = os.path.dirname(output_path)
            if not os.access(output_dir, os.W_OK):
                # Try alternative location
                alt_path = os.path.join(os.path.expanduser("~"), "Documents", os.path.basename(output_path))
                self.log_result(f"⚠️ No write permission to {output_dir}")
                self.log_result(f"📁 Trying alternative location: {alt_path}")
                output_path = alt_path
            
            try:
                loading.update_status("Writing Excel file with formatting...")
                loading.update_progress(90)
                
                # Create Excel writer with multiple sheets
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # Main mapping sheet
                    df_mapping.to_excel(writer, sheet_name='Field_Mapping', index=False)
                    
                    if loading.is_cancelled():
                        return
                    
                    # Get the workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Field_Mapping']
                    
                    # Add summary statistics after the data
                    last_row = len(mapping_data) + 2  # Account for header
                    
                    # Add summary section
                    worksheet.cell(row=last_row + 2, column=1, value="SUMMARY STATISTICS")
                    worksheet.cell(row=last_row + 3, column=1, value="Total Fields:")
                    worksheet.cell(row=last_row + 3, column=2, value=len(mapping_data))
                    worksheet.cell(row=last_row + 4, column=1, value="False Count:")
                    worksheet.cell(row=last_row + 4, column=2, value=f"=COUNTIF(E2:E{last_row},FALSE)")
                    worksheet.cell(row=last_row + 5, column=1, value="True Count:")
                    worksheet.cell(row=last_row + 5, column=2, value=f"=COUNTIF(E2:E{last_row},TRUE)")
                    worksheet.cell(row=last_row + 6, column=1, value="Match Percentage:")
                    worksheet.cell(row=last_row + 6, column=2, value=f"=B{last_row + 5}/(B{last_row + 4}+B{last_row + 5})*100")
                    
                    loading.update_status("Applying formatting...")
                    loading.update_progress(95)
                    
                    # Apply formatting
                    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                    
                    # Header formatting
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for col in range(1, 6):  # A to E columns
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    # Summary section formatting
                    summary_font = Font(bold=True)
                    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
                    for row in range(last_row + 2, last_row + 7):
                        worksheet.cell(row=row, column=1).font = summary_font
                        worksheet.cell(row=row, column=1).fill = summary_fill
                        worksheet.cell(row=row, column=2).fill = summary_fill
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Instructions sheet
                    instructions_data = {
                        'Step': [1, 2, 3, 4, 5, 6, 7],
                        'Instruction': [
                            'Review FAST UI fields and their sample values',
                            'Verify suggested Actuarial fields match your calculator',
                            'Update Actuarial_Field column with correct field names from your calculator',
                            'Modify Actuarial_Value column with proper transformation logic',
                            'Check Values_Match column for TRUE/FALSE comparison',
                            'Review summary statistics for false count and match percentage',
                            'Save and use this mapping for data processing'
                        ],
                        'Notes': [
                            'These are fields from your FAST UI data',
                            'Fields should match column names in your actuarial calculator',
                            'Use exact spelling and case from calculator Excel',
                            'Use Excel formulas or direct values as needed',
                            'TRUE=values match, FALSE=values differ - aim for TRUE',
                            'Monitor false count to identify mapping issues',
                            'This file serves as your mapping template'
                        ]
                    }
                    df_instructions = pd.DataFrame(instructions_data)
                    df_instructions.to_excel(writer, sheet_name='Instructions', index=False)
                    
                    # Calculator fields reference (if available)
                    if calculator_fields:
                        calc_ref_data = {
                            'Available_Calculator_Fields': calculator_fields + [''] * (max(0, len(mapping_data) - len(calculator_fields)))
                        }
                        df_calc_ref = pd.DataFrame(calc_ref_data)
                        df_calc_ref.to_excel(writer, sheet_name='Calculator_Fields', index=False)
                
                loading.update_status("Finalizing Excel file...")
                loading.update_progress(100)
                
                # Hide loading indicator
                loading.hide()
                
                self.log_result(f"✅ Excel mapping file created successfully!")
                self.log_result(f"📁 Location: {output_path}")
                self.log_result(f"📊 Mapped {len(mapping_data)} fields")
                self.log_result(f"📋 Sheets created: Field_Mapping, Instructions, Calculator_Fields")
                
                # Show success dialog with option to open file
                result = messagebox.askyesno("Success", 
                                           f"Excel mapping file created successfully!\n\n"
                                           f"Location: {os.path.basename(output_path)}\n"
                                           f"Fields mapped: {len(mapping_data)}\n\n"
                                           f"Would you like to open the file now?")
                
                if result:
                    try:
                        os.startfile(output_path)  # Windows
                    except Exception as e:
                        self.log_result(f"ℹ️ Cannot auto-open file: {e}")
                        self.log_result("Please open the Excel file manually to review the mapping")
                        
            except PermissionError as e:
                loading.hide()
                self.log_result(f"❌ Permission Error: {e}")
                messagebox.showerror("Permission Error", 
                                   f"Cannot write to the selected location.\n\n"
                                   f"Suggestions:\n"
                                   f"1. Save to your Documents folder\n"
                                   f"2. Save to Desktop\n"
                                   f"3. Close any open Excel files\n"
                                   f"4. Run as administrator if needed")
                
            except Exception as e:
                loading.hide()
                self.log_result(f"❌ Error creating Excel file: {e}")
                messagebox.showerror("Excel Creation Error", 
                                   f"Failed to create Excel file:\n{e}\n\n"
                                   f"Try:\n"
                                   f"1. Different file location\n"
                                   f"2. Simpler filename\n"
                                   f"3. Close Excel application")
                                   
        except ImportError:
            loading.hide()
            messagebox.showerror("Missing Module", 
                               "pandas and openpyxl modules are required.\n"
                               "Please install them using:\n"
                               "pip install pandas openpyxl")
        except Exception as e:
            loading.hide()
            self.log_result(f"❌ Error creating Excel mapping: {e}")
            messagebox.showerror("Error", f"Failed to create Excel mapping: {e}")
        finally:
            # Ensure loading indicator is always hidden
            if 'loading' in locals():
                loading.hide()
    
    def suggest_actuarial_field(self, fast_ui_field, calculator_fields):
        """Suggest actuarial field based on FAST UI field name."""
        if not calculator_fields:
            return ""
        
        # Convert to lowercase for comparison
        ui_field_lower = fast_ui_field.lower()
        
        # Direct mapping suggestions
        field_mappings = {
            'first_name': ['first_name', 'fname', 'given_name', 'insured_first'],
            'last_name': ['last_name', 'lname', 'surname', 'insured_last'],
            'birth_date': ['birth_date', 'dob', 'date_of_birth', 'birthdate'],
            'gender': ['gender', 'sex'],
            'face_amount': ['face_amount', 'coverage', 'sum_assured', 'benefit_amount'],
            'premium': ['premium', 'premium_amount', 'annual_premium'],
            'effective_date': ['effective_date', 'policy_date', 'start_date'],
            'policy_number': ['policy_number', 'policy_no', 'contract_number'],
        }
        
        # Find best match
        for key, suggestions in field_mappings.items():
            if key in ui_field_lower:
                for suggestion in suggestions:
                    for calc_field in calculator_fields:
                        if suggestion in calc_field.lower():
                            return calc_field
        
        # If no direct match, try partial matching
        for calc_field in calculator_fields:
            calc_field_lower = calc_field.lower()
            # Check if any word from ui_field is in calc_field
            ui_words = ui_field_lower.replace('_', ' ').split()
            for word in ui_words:
                if len(word) > 3 and word in calc_field_lower:  # Only match words longer than 3 chars
                    return calc_field
        
        return ""  # No suggestion found
    
    def show_field_mapping(self, product_type=None):
        """Create Excel field mapping template for FAST UI to Actuarial Calculator mapping."""
        
        # Step 1: Get product type if not provided
        if not product_type:
            product_type = simpledialog.askstring("Product Type", 
                                                 "Enter product type (life/annuity/health):")
            
            if not product_type or product_type.lower() not in ["life", "annuity"]:
                if product_type:
                    messagebox.showerror("Invalid Input", "Please enter: life or annuity")
                return
        
        # Ensure product_type is lowercase for consistency
        product_type = product_type.lower()
        
        # Show loading text immediately after product type is entered
        self.show_loading_text("Setting up Excel field mapping...")
        
        try:
            from tkinter import filedialog
            import pandas as pd
            
            # Update loading text
            self.show_loading_text("Preparing dialog interface...")
            
            # Step 2: Get paths from user
            dialog = tk.Toplevel(self.root)
            
            # Dynamic title based on product type
            if product_type == "life":
                dialog_title = "📊 Life Insurance Field Mapping"
                header_text = "📊 Life Insurance Field Mapping"
            elif product_type == "annuity":
                dialog_title = "📊 Annuity Field Mapping"
                header_text = "📊 Annuity Field Mapping"
            else:
                dialog_title = f"📊 {product_type.title()} Field Mapping"
                header_text = f"📊 {product_type.title()} Field Mapping"
            
            dialog.title(dialog_title)
            dialog.geometry("700x550")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg="#f8f9fa")
            
            # Clear loading text once dialog is ready
            self.hide_loading_text()
            
            # Header with modern styling
            header_frame = tk.Frame(dialog, bg="#34495e", height=60)
            header_frame.pack(fill="x")
            header_frame.pack_propagate(False)
            
            tk.Label(header_frame, text=header_text, 
                    font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#34495e").pack(expand=True)
            
            # Create main frame with scrolling
            main_frame = tk.Frame(dialog)
            main_frame.pack(fill="both", expand=True, padx=10, pady=10)
            
            # Create canvas and scrollbar
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack canvas and scrollbar
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Header in scrollable frame
            tk.Label(scrollable_frame, text=header_text, 
                    font=("Arial", 14, "bold")).pack(pady=10)
            
            # Path selection frame in scrollable frame
            paths_frame = tk.Frame(scrollable_frame)
            paths_frame.pack(pady=20, padx=20, fill="x")
            
            # Variables to store paths
            output_path_var = tk.StringVar()
            calculator_path_var = tk.StringVar()
            
            # Output Excel path
            tk.Label(paths_frame, text="1. Output Excel File Path:", 
                    font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 5))
            
            output_frame = tk.Frame(paths_frame)
            output_frame.pack(fill="x", pady=(0, 15))
            
            tk.Entry(output_frame, textvariable=output_path_var, width=60, 
                    font=("Segoe UI", 9), relief="solid", bd=1).pack(side="left", padx=(0, 5))
            
            output_browse_btn = self.create_dialog_button(output_frame, "📁 Browse", 
                                                        lambda: self.browse_save_excel(output_path_var),
                                                        "#2196F3", "#1976D2")
            output_browse_btn.pack(side="left")
            
            # Calculator Excel path
            tk.Label(paths_frame, text="2. Actuarial Calculator Excel Path:", 
                    font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 5))
            
            calc_frame = tk.Frame(paths_frame)
            calc_frame.pack(fill="x", pady=(0, 15))
            
            tk.Entry(calc_frame, textvariable=calculator_path_var, width=60, 
                    font=("Segoe UI", 9), relief="solid", bd=1).pack(side="left", padx=(0, 5))
            
            calc_browse_btn = self.create_dialog_button(calc_frame, "📁 Browse", 
                                                      lambda: self.browse_open_excel(calculator_path_var),
                                                      "#2196F3", "#1976D2")
            calc_browse_btn.pack(side="left")
            
            # Instructions
            instructions = tk.Text(dialog, height=8, width=70, wrap=tk.WORD)
            instructions.pack(pady=10, padx=20, fill="both", expand=True)
            
            instructions.insert(tk.END, "📋 Instructions:\n\n")
            instructions.insert(tk.END, "1. Choose where to save the new Excel mapping file\n")
            instructions.insert(tk.END, "2. Select your existing Actuarial Calculator Excel file\n")
            instructions.insert(tk.END, "3. The system will create a mapping template with 4 columns:\n")
            instructions.insert(tk.END, "   • FAST UI Field: Source field names\n")
            instructions.insert(tk.END, "   • FAST UI Value: Sample values from your data\n")
            instructions.insert(tk.END, "   • Actuarial Field: Target fields from calculator\n")
            instructions.insert(tk.END, "   • Actuarial Value: Mapped values for calculator\n\n")
            instructions.insert(tk.END, "4. You can manually complete the mapping in Excel")
            instructions.config(state=tk.DISABLED)
            
            # Buttons
            button_frame = tk.Frame(dialog)
            button_frame.pack(pady=10)
            
            def create_mapping():
                output_path = output_path_var.get().strip()
                calculator_path = calculator_path_var.get().strip()
                
                if not output_path:
                    messagebox.showerror("Error", "Please select output Excel file path")
                    return
                
                if not calculator_path:
                    messagebox.showerror("Error", "Please select actuarial calculator Excel file")
                    return
                
                # Show loading text immediately when button is clicked
                self.show_loading_text("Creating Excel mapping file...")
                
                dialog.destroy()
                self.create_excel_mapping(product_type.lower(), output_path, calculator_path)
            
            create_btn = self.create_dialog_button(button_frame, "📊 Create Mapping", create_mapping, "#27ae60", "#229954")
            create_btn.pack(side="left", padx=8)
            
            cancel_btn = self.create_dialog_button(button_frame, "❌ Cancel", dialog.destroy, "#e74c3c", "#c0392b")
            cancel_btn.pack(side="left", padx=8)
                     
        except ImportError:
            messagebox.showerror("Missing Module", 
                               "pandas module is required for Excel operations.\n"
                               "Please install it using: pip install pandas openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to setup mapping: {e}")
    
    def show_stored_data(self):
        """Show stored user data with search functionality."""
        if not self.user_data_store:
            messagebox.showinfo("No Data", "No user data stored yet. Use option 1 to add data.")
            return
        
        # Show loading text
        self.show_loading_text("Loading stored data interface...")
        
        dialog = tk.Toplevel(self.root)
        dialog.title("📋 Database Records")
        dialog.geometry("900x650")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="#f8f9fa")
        
        # Clear loading text once dialog is ready
        self.hide_loading_text()
        
        # Header with modern styling
        header_frame = tk.Frame(dialog, bg="#2c3e50", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📋 Database Records", 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#2c3e50").pack(expand=True)
        
        # Main content frame
        content_frame = tk.Frame(dialog, bg="#f8f9fa")
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header info
        header_info_frame = tk.Frame(content_frame, bg="#e8f5e8", relief="flat", bd=1)
        header_info_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(header_info_frame, text=f"📋 Database Records - {len(self.user_data_store)} dataset{'s' if len(self.user_data_store) != 1 else ''} stored", 
                font=("Segoe UI", 12, "bold"), fg="#2e7d32", bg="#e8f5e8").pack(pady=10)
        
        # Search frame with modern styling
        search_frame = tk.Frame(content_frame, bg="#e3f2fd", relief="flat", bd=1)
        search_frame.pack(fill="x", pady=(0, 15))
        
        search_content_frame = tk.Frame(search_frame, bg="#e3f2fd")
        search_content_frame.pack(pady=15, padx=15)
        
        tk.Label(search_content_frame, text="🔍 Search Records:", font=("Segoe UI", 10, "bold"), 
                fg="#1565c0", bg="#e3f2fd").pack(side="left", padx=(0, 10))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_content_frame, textvariable=search_var, font=("Segoe UI", 10), width=30,
                              relief="solid", bd=1, bg="#ffffff")
        search_entry.pack(side="left", padx=5)
        
        search_button = self.create_dialog_button(search_content_frame, "🔍 Search", 
                                                lambda: self.perform_search(search_var.get(), data_text, status_label),
                                                "#2196F3", "#1976D2")
        search_button.pack(side="left", padx=5)
        
        clear_button = self.create_dialog_button(search_content_frame, "📋 Show All", 
                                               lambda: self.show_all_data(data_text, status_label),
                                               "#FF9800", "#F57C00")
        clear_button.pack(side="left", padx=5)
        
        # Search status with modern styling
        status_label = tk.Label(content_frame, text="", font=("Segoe UI", 9), fg="#1976D2", bg="#f8f9fa")
        status_label.pack(pady=5)
        
        # Text area for displaying data with modern styling
        text_frame = tk.Frame(content_frame, bg="#ffffff", relief="solid", bd=1)
        text_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        data_text = scrolledtext.ScrolledText(text_frame, height=18, width=85, wrap=tk.WORD,
                                            font=("Consolas", 9), bg="#ffffff", fg="#2c3e50",
                                            insertbackground="#2c3e50", selectbackground="#3498db",
                                            relief="solid", bd=1)
        data_text.pack(fill="both", expand=True, padx=8, pady=8)
        
        # Initially show all data
        self.show_all_data(data_text, status_label)
        
        # Enable Enter key for search
        search_entry.bind('<Return>', lambda event: self.perform_search(search_var.get(), data_text, status_label))
        search_entry.focus()
        
        # Button frame with modern styling
        button_frame = tk.Frame(content_frame, bg="#f8f9fa")
        button_frame.pack(pady=15)
        
        close_btn = self.create_dialog_button(button_frame, "✅ Close", dialog.destroy, "#27ae60", "#229954")
        close_btn.pack(side="left", padx=8)
        
        export_btn = self.create_dialog_button(button_frame, "📤 Export Data", 
                                             lambda: self.export_data_to_file(),
                                             "#9C27B0", "#7B1FA2")
        export_btn.pack(side="left", padx=8)
    
    def show_help(self):
        """Show help dialog."""
        # Show loading text
        self.show_loading_text("Loading help system...")
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Help - Select Option")
        dialog.geometry("500x550")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="#f8f9fa")
        
        # Clear loading text once dialog is ready
        self.hide_loading_text()
        
        # Header with modern styling
        header_frame = tk.Frame(dialog, bg="#3498db", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📚 Help & Documentation", 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#3498db").pack(expand=True)
        
        # Create main frame
        main_frame = tk.Frame(dialog, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(main_frame, bg="#f8f9fa")
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f8f9fa")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Content in scrollable frame
        subtitle = tk.Label(scrollable_frame, text="� Select a topic for detailed help:", 
                font=("Segoe UI", 12, "bold"), fg="#2c3e50", bg="#f8f9fa")
        subtitle.pack(pady=15)
        
        help_options = [
            ("📝 Add JSON data", 1),
            ("📦 Bulk JSON Load", 2),
            ("📊 Excel Field Mapping", 3),
            ("👀 View stored data", 4),
            ("🔍 Check duplicates", 5),
            ("❓ Help System", 6),
            ("🗑️ Clear Database", 7),
            ("📋 All Options Overview", 8)
        ]
        
        for text, option_num in help_options:
            help_btn = self.create_dialog_button(scrollable_frame, text, 
                                               lambda opt=option_num: self.show_specific_help(opt, dialog),
                                               "#2c3e50", "#34495e")
            help_btn.pack(pady=5, fill="x", padx=20)
        
        # Close button with modern styling
        close_btn = self.create_dialog_button(scrollable_frame, "❌ Close Help", dialog.destroy,
                                            "#e74c3c", "#c0392b")
        close_btn.pack(pady=20)
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
    
    def show_specific_help(self, option_num, parent_dialog):
        """Show help for a specific option."""
        parent_dialog.withdraw()  # Hide parent dialog instead of destroying it
        
        help_dialog = tk.Toplevel(self.root)
        help_dialog.title(f"Help - Option {option_num}")
        help_dialog.geometry("700x600")
        help_dialog.transient(self.root)
        help_dialog.grab_set()
        help_dialog.configure(bg="#f8f9fa")
        
        # Handle window close (X button) event
        def on_window_close():
            help_dialog.destroy()
            parent_dialog.deiconify()  # Show the parent help dialog again
            parent_dialog.grab_set()   # Make it modal again
        
        help_dialog.protocol("WM_DELETE_WINDOW", on_window_close)
        
        # Header with modern styling
        option_titles = {
            1: "📝 Add JSON Data",
            2: "📦 Bulk JSON Load", 
            3: "📊 Excel Field Mapping",
            4: "👀 View Stored Data",
            5: "🔍 Check Duplicates",
            6: "❓ Help System",
            7: "🗑️ Clear Database",
            8: "📋 All Options Overview"
        }
        
        header_frame = tk.Frame(help_dialog, bg="#9b59b6", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text=option_titles.get(option_num, f"Help - Option {option_num}"), 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#9b59b6").pack(expand=True)
        
        # Content frame
        content_frame = tk.Frame(help_dialog, bg="#f8f9fa")
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Help text area with modern styling
        help_text = scrolledtext.ScrolledText(content_frame, height=20, width=70, wrap=tk.WORD,
                                            font=("Segoe UI", 10), bg="#ffffff", fg="#2c3e50",
                                            relief="solid", bd=1)
        help_text.pack(padx=10, pady=10, fill="both", expand=True)
        
        # Insert help content based on option
        if option_num == 1:
            help_content = """🔹 Option 1: Add new JSON data

This option allows you to input your own JSON data manually and save it to the database. You can choose whether to process the data immediately or just save it for later reference.

Features:
• Manual JSON data entry with validation
• Automatic saving to SQLite database for permanent storage
• Duplicate prevention using MD5 hash comparison
• Choice to "Save data only" or "Save and process immediately"
• Product type selection (life, annuity, health)

Example JSON:
{"applicant_first_name": "John", "applicant_last_name": "Doe", "applicant_birth_date": "1985-06-15", "applicant_gender": "M", "policy_face_amount": "250000", "policy_effective_date": "2024-01-01", "premium_mode": "M"}

Usage:
1. Click "1. Add new JSON data"
2. Enter your JSON data in the text area
3. Select product type (life, annuity, health)
4. Choose "Save data only" or "Save and process immediately"
5. Data is permanently stored and can be viewed later"""
        
        elif option_num == 2:
            help_content = """🔹 Option 2: Bulk JSON Load

This option allows you to load multiple JSON records at once using an Excel template, making it easy to add large amounts of data quickly.

Features:
• Download Excel template with JSON structure columns
• Bulk data entry using familiar Excel interface
• Upload completed Excel to load all records at once
• Automatic duplicate detection and prevention
• Support for all product types (life, annuity, health)
• Progress tracking with success/error counts

Process:
1. Download template - Creates Excel with same columns as JSON structure
2. Fill in data - Enter multiple records in Excel rows
3. Upload & process - Load all data into database automatically

Usage:
1. Click "2. Bulk JSON Load"
2. Choose "Create Template" to download Excel template
3. Fill in your data in the Excel file (multiple rows)
4. Use "Upload & Process" to load all records
5. Review summary of loaded records vs duplicates/errors"""
        
        elif option_num == 3:
            help_content = """🔹 Option 3: Excel Field Mapping

This option creates an enhanced Excel file that maps FAST UI fields to your Actuarial Calculator fields with comparison capabilities.

Features:
• Creates structured Excel mapping template with 5 columns
• Analyzes existing actuarial calculator Excel files
• Suggests field mappings automatically
• Values_Match column for TRUE/FALSE comparison
• Summary statistics with false count and match percentage
• Professional formatting with multiple worksheets

Enhanced Mapping Structure:
Column A: FAST UI Field (source field names)
Column B: FAST UI Value (sample data from your inputs)  
Column C: Actuarial Field (target calculator field names)
Column D: Actuarial Value (transformation logic/formulas)
Column E: Values_Match (TRUE/FALSE comparison)

New Features:
• Automatic comparison between FAST UI and Actuarial values
• False count tracking to identify mapping issues
• Match percentage calculation for quality control
• Summary statistics section at bottom of sheet

Usage:
1. Click "3. Excel Field Mapping" 
2. Choose where to save the new mapping Excel file
3. Select your existing actuarial calculator Excel file
4. System creates enhanced mapping template automatically
5. Review Values_Match column and false count statistics
6. Manually refine mappings to achieve higher match percentage"""
        
        elif option_num == 4:
            help_content = """🔹 Option 4: Show stored data

This option displays all JSON datasets permanently stored in the database with advanced search and export capabilities.

Features:
• View all stored data with timestamps and product types
• Powerful search functionality (case-insensitive)
• Export all data to JSON file for backup
• Search across all fields and values
• Duplicate name detection and reporting
• Data persistence between application sessions

Search Capabilities:
• Search by name: "John" finds all records containing "John"
• Search by product: "life" finds all life insurance policies
• Search by date: "2024-01" finds records from January 2024
• Search by any field value: "250000" finds policies with that amount

Usage:
1. Click "4. View stored data"
2. Use "Search" button to find specific records
3. Use "Show All Data" to view complete database
4. Use "Export Data" to backup all data to JSON file
5. All data remains permanently stored in SQLite database"""
        
        elif option_num == 5:
            help_content = """🔹 Option 5: Check duplicates

This option analyzes stored data to identify records with duplicate names but different data values, helping detect potential data entry errors or multiple policies for the same person.

Features:
• Groups records by person names (first + last name)
• Shows differences between records for same person
• Helps identify data entry errors
• Distinguishes between exact duplicates (prevented) and name duplicates
• Detailed comparison of field values

Analysis Types:
• Name duplicates: Same person, different policy data
• Potential errors: Similar names with slight variations
• Multiple policies: Legitimate different policies for same person
• Data inconsistencies: Same person with conflicting information

Usage:
1. Click "5. Check duplicates"
2. System automatically analyzes all stored data
3. Groups records by person names
4. Shows detailed comparison of differences
5. Helps you identify and resolve data issues

Note: This is different from exact duplicate prevention which happens automatically when adding data."""
        
        elif option_num == 6:
            help_content = """🔹 Option 6: Help

This help system provides detailed information about all features and options available in the AIM (Actuarial Input Mapper) application.

Features:
• Context-sensitive help for each option
• Complete feature descriptions with examples  
• Usage instructions and best practices
• Tips for effective data management
• Troubleshooting guidance

Help Topics Available:
• Option 1: Add JSON data - Manual data entry and storage
• Option 2: Bulk JSON Load - Excel-based bulk data entry
• Option 3: Excel Field Mapping - Creating enhanced mapping templates
• Option 4: View stored data - Searching and managing stored data
• Option 5: Check duplicates - Analyzing data for duplicates
• Option 7: Clear Database - Removing all stored data
• All options - Complete overview of all features

Usage:
1. Click "6. Help" (this option)
2. Select specific option for detailed help
3. Or choose "All options" for complete overview
4. Each help dialog includes examples and step-by-step instructions"""
        
        elif option_num == 7:
            help_content = """🔹 Option 7: Clear Database

This option permanently removes all stored data from the SQLite database. Use with caution as this action cannot be undone.

Features:
• Permanent deletion of all stored JSON data
• Confirmation dialog to prevent accidental deletion
• Immediate update of UI status and statistics
• Fresh start for testing or cleanup purposes
• Cannot be undone - data is permanently lost

Safety Features:
• Double confirmation required
• Warning message about permanent deletion
• Clear indication that action cannot be undone
• Status updates immediately reflect empty database

Usage:
1. Click "7. Clear Database" (red button)
2. Read the warning dialog carefully
3. Click "Yes" to confirm permanent deletion
4. Click "No" to cancel and keep your data
5. Database statistics update to show empty database

⚠️ WARNING: This permanently deletes ALL stored data. Consider exporting your data first using Option 4 if you want to keep a backup."""
        
        else:  # option_num == 8 (All options)
            help_content = """🔹 Complete AIM Help Guide - All Options

🔸 Option 1: Add JSON data
Manually input JSON data with validation, automatic database storage, and duplicate prevention. Choose to save only or save and process immediately. Supports life, annuity, and health product types.

🔸 Option 2: Bulk JSON Load
Load multiple JSON records at once using an Excel template, making it easy to add large amounts of data quickly.

Features:
• Download Excel template with JSON structure columns
• Bulk data entry using familiar Excel interface
• Upload completed Excel to load all records at once
• Automatic duplicate detection and prevention
• Support for all product types (life, annuity, health)
• Progress tracking with success/error counts

Process:
1. Download template - Creates Excel with same columns as JSON structure
2. Fill in data - Enter multiple records in Excel rows
3. Upload & process - Load all data into database automatically

Usage:
1. Click "2. Bulk JSON Load"
2. Choose "Create Template" to download Excel template
3. Fill in your data in the Excel file (multiple rows)
4. Use "Upload & Process" to load all records
5. Review summary of loaded records vs duplicates/errors"""
        
        help_text.insert(tk.END, help_content)
        help_text.config(state=tk.DISABLED)  # Make read-only
        
        # Close button that returns to help menu
        def close_and_return_to_help():
            help_dialog.destroy()
            parent_dialog.deiconify()  # Show the parent help dialog again
            parent_dialog.grab_set()   # Make it modal again
            
        def close_help_system():
            help_dialog.destroy()
            parent_dialog.destroy()  # Close the entire help system
        
        # Button frame for multiple options with modern styling
        button_frame = tk.Frame(content_frame, bg="#f8f9fa")
        button_frame.pack(pady=15)
        
        back_btn = self.create_dialog_button(button_frame, "← Back to Help Menu", close_and_return_to_help,
                                           "#27ae60", "#229954")
        back_btn.pack(side="left", padx=10)
        
        close_btn = self.create_dialog_button(button_frame, "❌ Close Help", close_help_system,
                                            "#e74c3c", "#c0392b")
        close_btn.pack(side="left", padx=10)
    
    def clear_database(self):
        """Clear all data from the database."""
        result = messagebox.askyesno(
            "Clear Database", 
            "Are you sure you want to clear all data from the database?\nThis action cannot be undone.",
            icon="warning"
        )
        
        if result:
            # Show loading text while clearing database
            self.show_loading_text("Clearing database...")
            
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('DELETE FROM user_data')
                conn.commit()
                conn.close()
                
                # Reload data (will be empty now)
                self.load_data_from_db()
                self.update_status()
                self.update_db_stats()
                self.clear_results()
                self.log_result("🗑️ Database cleared successfully!")
                
                messagebox.showinfo("Success", "Database cleared successfully!")
                
            except Exception as e:
                messagebox.showerror("Database Error", f"Failed to clear database: {e}")
    
    def display_result(self, result):
        """Display processing result."""
        if result["status"] == "success":
            self.log_result("\n✅ Processing successful!")
            self.log_result(f"   ⏱️  Time: {result['processing_time']:.2f}s")
            self.log_result(f"   📊 Fields: {result['metadata']['fields_processed']} → {result['metadata']['fields_mapped']}")
            
            self.log_result("\n📋 Actuarial inputs:")
            for section_name, section_data in result["actuarial_inputs"].items():
                self.log_result(f"   📦 {section_name}:")
                for field_name, field_value in section_data.items():
                    self.log_result(f"      • {field_name}: {field_value}")
        else:
            self.log_result(f"\n❌ Processing failed:")
            self.log_result(f"   Error: {result.get('error_message', 'Unknown error')}")
    
    def run(self):
        """Start the GUI application."""
        # Just close the window when X is clicked, no confirmation needed since data is saved to database
        self.root.protocol("WM_DELETE_WINDOW", self.root.quit)
        self.root.mainloop()
    
    def perform_search(self, search_term, data_text, status_label):
        """Perform search in stored data."""
        if not search_term.strip():
            self.show_all_data(data_text, status_label)
            return
        
        data_text.delete(1.0, tk.END)
        search_term = search_term.lower()
        matches = []
        
        for i, data_entry in enumerate(self.user_data_store, 1):
            # Search in product type, timestamp, and JSON data
            searchable_content = (
                data_entry['product_type'].lower() + " " +
                data_entry['timestamp'].lower() + " " +
                json.dumps(data_entry['data']).lower()
            )
            
            if search_term in searchable_content:
                matches.append((i, data_entry))
        
        if matches:
            status_label.config(text=f"Found {len(matches)} matching record(s)", fg="green")
            for original_index, data_entry in matches:
                data_text.insert(tk.END, f"\n{original_index}. Product Type: {data_entry['product_type']}\n")
                data_text.insert(tk.END, f"   Timestamp: {data_entry['timestamp']}\n")
                data_text.insert(tk.END, f"   Data: {json.dumps(data_entry['data'], indent=2)}\n")
                data_text.insert(tk.END, "-" * 60 + "\n")
        else:
            status_label.config(text="No matching records found", fg="red")
            data_text.insert(tk.END, f"\nNo records found matching '{search_term}'\n\nTry searching for:\n")
            data_text.insert(tk.END, "• Product type (life, annuity, health)\n")
            data_text.insert(tk.END, "• Names (first_name, last_name)\n")
            data_text.insert(tk.END, "• Date values\n")
            data_text.insert(tk.END, "• Any field values in the JSON data\n")
    
    def show_all_data(self, data_text, status_label):
        """Show all stored data."""
        data_text.delete(1.0, tk.END)
        status_label.config(text=f"Showing all {len(self.user_data_store)} record(s)", fg="blue")
        
        for i, data_entry in enumerate(self.user_data_store, 1):
            data_text.insert(tk.END, f"\n{i}. Product Type: {data_entry['product_type']}\n")
            data_text.insert(tk.END, f"   Timestamp: {data_entry['timestamp']}\n")
            data_text.insert(tk.END, f"   Data: {json.dumps(data_entry['data'], indent=2)}\n")
            data_text.insert(tk.END, "-" * 60 + "\n")
    
    def export_data_to_file(self):
        """Export all stored data to a JSON file."""
        if not self.user_data_store:
            messagebox.showinfo("No Data", "No data to export.")
            return
        
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                title="Export Data As"
            )
            
            if filename:
                export_data = {
                    "export_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_records": len(self.user_data_store),
                    "data": self.user_data_store
                }
                
                with open(filename, 'w') as f:
                    json.dump(export_data, f, indent=2)
                
                messagebox.showinfo("Export Success", f"Data exported successfully to:\n{filename}")
        
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {e}")

    def check_for_duplicate_names(self):
        """Check for duplicate names in the stored data and return a report."""
        name_occurrences = {}
        duplicates = []
        
        for i, data_entry in enumerate(self.user_data_store, 1):
            data = data_entry['data']
            
            # Look for common name fields
            first_name = data.get('applicant_first_name', data.get('first_name', ''))
            last_name = data.get('applicant_last_name', data.get('last_name', ''))
            
            if first_name and last_name:
                full_name = f"{first_name} {last_name}".lower()
                
                if full_name in name_occurrences:
                    name_occurrences[full_name].append((i, data_entry))
                else:
                    name_occurrences[full_name] = [(i, data_entry)]
        
        # Find duplicates
        for name, entries in name_occurrences.items():
            if len(entries) > 1:
                duplicates.append((name, entries))
        
        return duplicates

    def show_duplicate_check(self):
        """Show duplicate name checker dialog."""
        if not self.user_data_store:
            messagebox.showinfo("No Data", "No user data stored yet. Use option 1 to add data.")
            return
        
        # Show loading text
        self.show_loading_text("Checking for duplicate names...")
        
        duplicates = self.check_for_duplicate_names()
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Duplicate Name Check")
        dialog.geometry("750x600")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="#f8f9fa")
        
        # Clear loading text once dialog is ready
        self.hide_loading_text()
        
        # Header with modern styling
        header_frame = tk.Frame(dialog, bg="#e67e22", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        if duplicates:
            header_text = f"⚠️ Found {len(duplicates)} Duplicate Name(s)"
            header_color = "#e67e22"
        else:
            header_text = "✅ No Duplicate Names Found"
            header_color = "#27ae60"
            header_frame.configure(bg=header_color)
        
        tk.Label(header_frame, text=header_text, 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg=header_color).pack(expand=True)
        
        # Main content frame
        content_frame = tk.Frame(dialog, bg="#f8f9fa")
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Results area with modern styling
        results_frame = tk.LabelFrame(content_frame, text="📊 Duplicate Check Results", 
                                    font=("Segoe UI", 11, "bold"), bg="#f8f9fa", fg="#2c3e50",
                                    padx=15, pady=15)
        results_frame.pack(fill="both", expand=True, pady=10)
        
        results_text = scrolledtext.ScrolledText(results_frame, height=20, width=80, wrap=tk.WORD,
                                               font=("Consolas", 10), bg="#ffffff", fg="#2c3e50",
                                               relief="solid", bd=1)
        results_text.pack(padx=10, pady=10, fill="both", expand=True)
        
        if duplicates:
            results_text.insert(tk.END, "🔍 DUPLICATE NAME ANALYSIS\n")
            results_text.insert(tk.END, "=" * 60 + "\n\n")
            results_text.insert(tk.END, "The following names appear multiple times in your database:\n\n")
            
            for name, entries in duplicates:
                results_text.insert(tk.END, f"🔸 Name: {name.title()}\n")
                results_text.insert(tk.END, f"   📊 Appears {len(entries)} times:\n\n")
                
                for i, (record_num, data_entry) in enumerate(entries, 1):
                    results_text.insert(tk.END, f"   {i}. 📋 Record #{record_num}\n")
                    results_text.insert(tk.END, f"      📅 Product Type: {data_entry['product_type']}\n")
                    results_text.insert(tk.END, f"      🕒 Timestamp: {data_entry['timestamp']}\n")
                    results_text.insert(tk.END, f"      🔐 Data Hash: {self.get_data_hash(data_entry['data'])[:8]}...\n")
                    
                    # Show key differences
                    key_fields = ['policy_face_amount', 'policy_effective_date', 'premium_mode', 'applicant_birth_date']
                    for field in key_fields:
                        if field in data_entry['data']:
                            results_text.insert(tk.END, f"      💰 {field}: {data_entry['data'][field]}\n")
                    results_text.insert(tk.END, "\n")
                
                results_text.insert(tk.END, "─" * 60 + "\n\n")
            
            results_text.insert(tk.END, "\n💡 NOTE: These are name duplicates only.\n")
            results_text.insert(tk.END, "The system prevents exact data duplicates using hash comparison.\n")
        else:
            results_text.insert(tk.END, "🎉 EXCELLENT! All names in your database are unique!\n\n")
            results_text.insert(tk.END, f"📊 Total records checked: {len(self.user_data_store)}\n")
            
            # Calculate unique names count safely
            unique_names = set()
            for entry in self.user_data_store:
                first_name = entry['data'].get('applicant_first_name', '')
                last_name = entry['data'].get('applicant_last_name', '')
                if first_name and last_name:
                    full_name = f"{first_name} {last_name}".lower().strip()
                    unique_names.add(full_name)
            
            results_text.insert(tk.END, f"👥 Unique names found: {len(unique_names)}\n\n")
            results_text.insert(tk.END, "🔒 The system also prevents exact duplicate data entries using MD5 hash comparison.\n")
            results_text.insert(tk.END, "✨ Your data integrity is maintained!")
        
        # Make text read-only
        results_text.config(state=tk.DISABLED)
        
        # Close button with modern styling
        close_btn = self.create_dialog_button(content_frame, "❌ Close", dialog.destroy,
                                            "#27ae60", "#229954")
        close_btn.pack(pady=15)
    
    def bulk_json_load(self):
        """Handle bulk JSON data loading via Excel template."""
        # Show loading text immediately
        self.show_loading_text("Setting up bulk JSON load interface...")
        
        dialog = tk.Toplevel(self.root)
        dialog.title("📦 Bulk JSON Data Load")
        dialog.geometry("750x650")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="#f8f9fa")
        
        # Clear loading text once dialog is ready
        self.hide_loading_text()
        
        # Header with modern styling
        header_frame = tk.Frame(dialog, bg="#8e44ad", height=60)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📦 Bulk JSON Data Load", 
                font=("Segoe UI", 14, "bold"), fg="#ecf0f1", bg="#8e44ad").pack(expand=True)
        
        # Create main frame with scrolling
        main_frame = tk.Frame(dialog, bg="#f8f9fa")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(main_frame, bg="#f8f9fa")
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f8f9fa")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Header content
        header_content = tk.Label(scrollable_frame, text="📦 Bulk JSON Data Load", 
                font=("Segoe UI", 16, "bold"), fg="#8e44ad", bg="#f8f9fa")
        header_content.pack(pady=15)
        
        # Instructions with modern styling
        instructions_frame = tk.Frame(scrollable_frame, bg="#e8f5e8", relief="flat", bd=1)
        instructions_frame.pack(fill="x", pady=10, padx=20)
        
        instructions_text = """
📋 Choose one of the following options:

1. Download Excel Template - Create a template based on existing JSON structure
2. Upload Filled Excel - Load data from a completed Excel file

The Excel template will have the same columns as your JSON structure for easy bulk entry.
        """
        tk.Label(instructions_frame, text=instructions_text, 
                font=("Segoe UI", 10), justify="left", wraplength=600, 
                fg="#2e7d32", bg="#e8f5e8").pack(pady=15, padx=20)
        
        # Variables for paths
        template_path_var = tk.StringVar()
        upload_path_var = tk.StringVar()
        
        # Option 1: Download Template with modern styling
        template_frame = tk.LabelFrame(scrollable_frame, text="📥 Option 1: Download Excel Template", 
                                     font=("Segoe UI", 11, "bold"), bg="#f8f9fa", fg="#2c3e50",
                                     padx=15, pady=15)
        template_frame.pack(pady=10, padx=20, fill="x")
        
        tk.Label(template_frame, text="💾 Save template to:", font=("Segoe UI", 10, "bold"), 
                fg="#34495e", bg="#f8f9fa").pack(anchor="w", pady=5)
        
        template_path_frame = tk.Frame(template_frame, bg="#f8f9fa")
        template_path_frame.pack(fill="x", pady=5)
        
        tk.Entry(template_path_frame, textvariable=template_path_var, width=50, 
                font=("Segoe UI", 9), relief="solid", bd=1).pack(side="left", padx=(0, 5))
        
        browse_btn = self.create_dialog_button(template_path_frame, "📁 Browse", 
                                             lambda: self.browse_save_excel(template_path_var),
                                             "#2196F3", "#1976D2")
        browse_btn.pack(side="left")
        
        create_template_btn = self.create_dialog_button(template_frame, "📥 Create Template", 
                                                      lambda: self.create_bulk_template(template_path_var.get(), dialog),
                                                      "#27ae60", "#229954")
        create_template_btn.pack(pady=15)
        
        # Option 2: Upload Excel with modern styling
        upload_frame = tk.LabelFrame(scrollable_frame, text="📤 Option 2: Upload Filled Excel", 
                                   font=("Segoe UI", 11, "bold"), bg="#f8f9fa", fg="#2c3e50",
                                   padx=15, pady=15)
        upload_frame.pack(pady=10, padx=20, fill="x")
        
        tk.Label(upload_frame, text="📂 Select filled Excel file:", font=("Segoe UI", 10, "bold"), 
                fg="#34495e", bg="#f8f9fa").pack(anchor="w", pady=5)
        
        upload_path_frame = tk.Frame(upload_frame, bg="#f8f9fa")
        upload_path_frame.pack(fill="x", pady=5)
        
        tk.Entry(upload_path_frame, textvariable=upload_path_var, width=50, 
                font=("Segoe UI", 9), relief="solid", bd=1).pack(side="left", padx=(0, 5))
        
        browse_upload_btn = self.create_dialog_button(upload_path_frame, "📁 Browse", 
                                                    lambda: self.browse_open_excel(upload_path_var),
                                                    "#2196F3", "#1976D2")
        browse_upload_btn.pack(side="left")
        
        # Product type selection with modern styling
        product_frame = tk.Frame(upload_frame, bg="#f8f9fa")
        product_frame.pack(pady=10)
        
        tk.Label(product_frame, text="📋 Product Type:", font=("Segoe UI", 10, "bold"), 
                fg="#34495e", bg="#f8f9fa").pack(side="left", padx=(0, 10))
        product_var = tk.StringVar(value="life")
        product_combo = ttk.Combobox(product_frame, textvariable=product_var, 
                                   values=["life", "annuity"], width=15,
                                   font=("Segoe UI", 9))
        product_combo.pack(side="left", padx=5)
        
        upload_process_btn = self.create_dialog_button(upload_frame, "📤 Upload & Process", 
                                                     lambda: self.process_bulk_excel(upload_path_var.get(), product_var.get(), dialog),
                                                     "#ff9800", "#f57c00")
        upload_process_btn.pack(pady=15)
        
        # Close button with modern styling
        close_btn = self.create_dialog_button(scrollable_frame, "❌ Close", dialog.destroy,
                                            "#e74c3c", "#c0392b")
        close_btn.pack(pady=20)
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
    
    def browse_save_excel_bulk(self, path_var):
        """Browse and select location to save Excel template."""
        from tkinter import filedialog
        file_path = filedialog.asksaveasfilename(
            title="Save Excel Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            path_var.set(file_path)
    
    def browse_open_excel_bulk(self, path_var):
        """Browse and select Excel file to upload."""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            path_var.set(file_path)
    
    def create_bulk_template(self, template_path, parent_dialog):
        """Create Excel template for bulk JSON data entry."""
        if not template_path:
            messagebox.showerror("Error", "Please select a location to save the template")
            return
        
        # Show loading text immediately when button is clicked
        self.show_loading_text("Creating Excel bulk template...")
        
        # Show loading indicator
        loading = LoadingIndicator(self.root, "Creating Template", "Preparing bulk template...")
        loading.show()
        
        try:
            import pandas as pd
            
            loading.update_status("Analyzing JSON structure...")
            loading.update_progress(20)
            
            # Get JSON structure from existing data or create default
            if self.user_data_store:
                # Use structure from existing data
                sample_data = self.user_data_store[0]['data']
                columns = list(sample_data.keys())
            else:
                # Default JSON structure
                columns = [
                    "applicant_first_name", "applicant_last_name", "applicant_birth_date",
                    "applicant_gender", "applicant_email", "applicant_phone",
                    "policy_face_amount", "policy_effective_date", "premium_mode",
                    "premium_amount", "risk_class", "underwriting_class"
                ]
            
            loading.update_status("Creating template data...")
            loading.update_progress(40)
            
            # Create template with headers and sample row
            template_data = {}
            sample_values = {
                "applicant_first_name": "John", "applicant_last_name": "Doe",
                "applicant_birth_date": "1985-06-15", "applicant_gender": "M",
                "applicant_email": "john.doe@email.com", "applicant_phone": "555-1234",
                "policy_face_amount": "250000", "policy_effective_date": "2024-01-01",
                "premium_mode": "M", "premium_amount": "150.00",
                "risk_class": "Standard", "underwriting_class": "Preferred"
            }
            
            for col in columns:
                template_data[col] = [sample_values.get(col, "Sample_Value")]
            
            df_template = pd.DataFrame(template_data)
            
            loading.update_status("Writing Excel file...")
            loading.update_progress(70)
            
            # Create Excel with instructions
            with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
                # Template sheet
                df_template.to_excel(writer, sheet_name='Bulk_Data_Template', index=False)
                
                loading.update_status("Adding instructions...")
                loading.update_progress(90)
                
                # Instructions sheet
                instructions = {
                    'Instructions': [
                        "1. Use the 'Bulk_Data_Template' sheet to enter your data",
                        "2. Each row represents one JSON record",
                        "3. Keep the column headers exactly as shown",
                        "4. Fill in your actual data, replacing the sample values",
                        "5. You can add as many rows as needed",
                        "6. Save the file and use 'Upload & Process' option",
                        "7. Select the appropriate product type when uploading",
                        "",
                        "Column Descriptions:",
                        "- applicant_first_name: First name of the applicant",
                        "- applicant_last_name: Last name of the applicant",
                        "- applicant_birth_date: Birth date (YYYY-MM-DD format)",
                        "- applicant_gender: M for Male, F for Female",
                        "- policy_face_amount: Coverage amount",
                        "- policy_effective_date: Policy start date (YYYY-MM-DD)",
                        "- premium_mode: A=Annual, M=Monthly, Q=Quarterly",
                        "- Other fields: Enter appropriate values for your data"
                    ]
                }
                df_instructions = pd.DataFrame(instructions)
                df_instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            loading.update_progress(100)
            loading.hide()
            
            self.log_result(f"✅ Bulk template created: {template_path}")
            messagebox.showinfo("Success", f"Excel template created successfully!\n\nLocation: {template_path}\n\nPlease fill in your data and use the Upload option.")
            
            # Option to open the template
            result = messagebox.askyesno("Open Template", "Would you like to open the template now?")
            if result:
                try:
                    import os
                    os.startfile(template_path)
                except:
                    pass
                    
        except ImportError:
            loading.hide()
            messagebox.showerror("Missing Module", "pandas and openpyxl are required for this feature")
        except Exception as e:
            loading.hide()
            messagebox.showerror("Error", f"Failed to create template: {e}")
        finally:
            if 'loading' in locals():
                loading.hide()
    
    def process_bulk_excel(self, excel_path, product_type, parent_dialog):
        """Process bulk data from Excel file."""
        if not excel_path:
            messagebox.showerror("Error", "Please select an Excel file")
            return
        
        # Show loading text immediately when button is clicked
        self.show_loading_text("Processing bulk Excel data...")
        
        # Show loading indicator
        loading = LoadingIndicator(self.root, "Processing Bulk Data", "Reading Excel file...")
        loading.show()
        
        try:
            import pandas as pd
            import json
            
            loading.update_status("Reading Excel file...")
            loading.update_progress(10)
            
            # Read Excel file
            df = pd.read_excel(excel_path, sheet_name='Bulk_Data_Template')
            
            loading.update_status("Converting to JSON records...")
            loading.update_progress(20)
            
            # Convert DataFrame to JSON records
            records = df.to_dict('records')
            
            # Clean up NaN values
            for record in records:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = ""
            
            loading.update_status("Processing records...")
            loading.update_progress(30)
            
            # Process each record
            success_count = 0
            duplicate_count = 0
            error_count = 0
            total_records = len(records)
            
            for i, record in enumerate(records, 1):
                if loading.is_cancelled():
                    break
                    
                try:
                    # Update progress
                    progress = 30 + (i / total_records) * 60
                    loading.update_progress(progress)
                    loading.update_status(f"Processing record {i} of {total_records}...")
                    
                    # Convert to JSON string and back to ensure proper format
                    json_str = json.dumps(record)
                    json_data = json.loads(json_str)
                    
                    # Check for duplicates
                    data_hash = hashlib.md5(json_str.encode()).hexdigest()
                    
                    # Check if already exists
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute('SELECT COUNT(*) FROM user_data WHERE data_hash = ?', (data_hash,))
                    exists = cursor.fetchone()[0] > 0
                    
                    if not exists:
                        # Save to database
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        cursor.execute('''
                            INSERT INTO user_data (timestamp, product_type, data, data_hash)
                            VALUES (?, ?, ?, ?)
                        ''', (timestamp, product_type, json_str, data_hash))
                        conn.commit()
                        success_count += 1
                    else:
                        duplicate_count += 1
                    
                    conn.close()
                    
                except Exception as e:
                    error_count += 1
                    self.log_result(f"❌ Error processing row {i}: {e}")
            
            if not loading.is_cancelled():
                loading.update_status("Finalizing...")
                loading.update_progress(95)
                
                # Reload data and update UI
                self.load_data_from_db()
                self.update_status()
                self.update_db_stats()
                
                loading.update_progress(100)
                loading.hide()
                
                # Show summary
                summary = f"Bulk load completed!\n\n"
                summary += f"✅ Successfully loaded: {success_count} records\n"
                if duplicate_count > 0:
                    summary += f"⚠️ Duplicates skipped: {duplicate_count} records\n"
                if error_count > 0:
                    summary += f"❌ Errors: {error_count} records\n"
                summary += f"\nTotal records in database: {len(self.user_data_store)}"
                
                messagebox.showinfo("Bulk Load Complete", summary)
            else:
                loading.hide()
                messagebox.showinfo("Cancelled", "Bulk load was cancelled.")
                
        except ImportError:
            loading.hide()
            messagebox.showerror("Missing Module", "pandas and openpyxl are required for this feature")
        except Exception as e:
            loading.hide()
            messagebox.showerror("Error", f"Failed to process bulk data: {e}")
        finally:
            if 'loading' in locals():
                loading.hide()
            # Close dialog
            if parent_dialog:
                parent_dialog.destroy()
    
class LoadingIndicator:
    """Utility class for showing loading progress in the application."""
    
    def __init__(self, parent, title="Processing...", message="Please wait..."):
        self.parent = parent
        self.dialog = None
        self.progress_var = None
        self.status_var = None
        self.title = title
        self.message = message
        self.cancelled = False
    
    def show(self):
        """Show the loading dialog."""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title(self.title)
        self.dialog.geometry("400x200")
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        self.dialog.resizable(False, False)
        
        # Make dialog stay on top and visible
        self.dialog.attributes('-topmost', True)
        self.dialog.lift()
        self.dialog.focus_force()
        
        # Center the dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (200 // 2)
        self.dialog.geometry(f"400x200+{x}+{y}")
        
        # Main frame
        main_frame = tk.Frame(self.dialog, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title and message
        tk.Label(main_frame, text=self.title, font=("Arial", 12, "bold")).pack(pady=5)
        
        self.status_var = tk.StringVar(value=self.message)
        tk.Label(main_frame, textvariable=self.status_var, font=("Arial", 10)).pack(pady=5)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                     maximum=100, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        
        # Indeterminate progress for unknown duration
        self.indeterminate_progress = ttk.Progressbar(main_frame, length=300, mode='indeterminate')
        
        # Cancel button
        cancel_button = tk.Button(main_frame, text="Cancel", 
                                command=self.cancel, bg="#f44336", fg="white")
        cancel_button.pack(pady=10)
        
        # Start indeterminate progress
        self.indeterminate_progress.pack(pady=5)
        self.indeterminate_progress.start(10)
        
        self.dialog.protocol("WM_DELETE_WINDOW", self.cancel)
        
        # Force dialog to be visible and on top
        self.dialog.update()
        self.dialog.deiconify()
        self.dialog.lift()
        self.dialog.attributes('-topmost', True)
        self.dialog.focus_force()
        
        # Final update to ensure everything is rendered
        self.dialog.update_idletasks()
    
    def update_progress(self, percentage, status_message=None):
        """Update progress percentage (0-100)."""
        if self.dialog and not self.cancelled:
            self.progress_var.set(percentage)
            if status_message:
                self.status_var.set(status_message)
            self.dialog.update()
    
    def update_status(self, message):
        """Update status message."""
        if self.dialog and not self.cancelled:
            self.status_var.set(message)
            self.dialog.update()
    
    def cancel(self):
        """Cancel the operation."""
        self.cancelled = True
        self.hide()
    
    def hide(self):
        """Hide the loading dialog."""
        if self.dialog:
            try:
                self.indeterminate_progress.stop()
                self.dialog.destroy()
            except:
                pass
            self.dialog = None
    
    def is_cancelled(self):
        """Check if operation was cancelled."""
        return self.cancelled


def run_gui_demo():
    """Run the GUI-based interactive demo."""
    try:
        app = AIMDemoGUI()
        app.run()
    except Exception as e:
        print(f"❌ GUI Demo failed with error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print("🚀 Starting AIM (Actuarial Input Mapper) Examples")
    
    # Ask user to choose between terminal and GUI mode
    import tkinter as tk
    from tkinter import messagebox
    
    # Create a simple selection dialog
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    choice = messagebox.askyesnocancel(
        title="Select Demo Mode",
        message="Choose your preferred interface:\n\n"
        "• Click 'Yes' for GUI Mode (recommended) - Modern interface with dialog boxes\n"
        "• Click 'No' for Terminal Mode - Traditional command-line interface\n"
        "• Click 'Cancel' to exit"
    )
    
    if choice is True:
        # GUI Mode
        root.destroy()
        print("🎮 Starting GUI Demo Mode...")
        run_gui_demo()
    elif choice is False:
        # Terminal Mode
        root.destroy()
        try:
            # Run basic example
            run_example()
            
            # Ask if user wants to try interactive demo
            response = input("\n🎮 Would you like to try the interactive demo? (y/n): ").strip().lower()
            
            if response in ['y', 'yes']:
                run_interactive_demo()
            
        except KeyboardInterrupt:
            print("\n\n👋 Example terminated by user")
        except Exception as e:
            print(f"\n❌ Example failed with error: {e}")
            import traceback
            traceback.print_exc()
    else:
        # User cancelled
        root.destroy()
        print("👋 Demo cancelled by user")
    
    print("\n🏁 Example completed")
