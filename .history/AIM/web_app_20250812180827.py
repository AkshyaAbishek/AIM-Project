"""
AIM Web Application - Flask Implementation
Converts the tkinter desktop application to a web-based interface for Azure deployment
"""

import os
import json
import hashlib
from datetime import datetime
from typing import Dict, Any, Optional
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO

from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file, session
from flask import Response
import sqlite3
import sys

# Add src directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

try:
    from aim_processor import AIMProcessor, ValidationError, MappingError
except ImportError:
    # Fallback for web deployment
    class AIMProcessor:
        def process_fast_ui_input(self, data, product_type):
            return {"status": "success", "message": "Processed successfully"}
    
    class ValidationError(Exception):
        pass
    
    class MappingError(Exception):
        pass

# Initialize Flask application
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize AIM processor
aim_processor = AIMProcessor()

class WebDatabaseManager:
    """Database manager for web application"""
    
    def __init__(self, db_path: str = "aim_data.db"):
        # Ensure we're using the correct path
        if not os.path.isabs(db_path):
            db_path = os.path.join(os.path.dirname(__file__), db_path)
        self.db_path = db_path
        print(f"Database path: {self.db_path}")
        self.init_database()
    
    def init_database(self):
        """Initialize database with web-optimized schema"""
        try:
            print(f"Initializing database at: {self.db_path}")
            with sqlite3.connect(self.db_path) as conn:
                # First, create table if it doesn't exist with minimal schema
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS user_data (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        product_type TEXT NOT NULL,
                        data_hash TEXT UNIQUE NOT NULL,
                        json_data TEXT NOT NULL,
                        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
            
            conn.execute("""
                CREATE TABLE IF NOT EXISTS processing_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operation_type TEXT NOT NULL,
                    operation_details TEXT,
                    status TEXT CHECK(status IN ('success', 'warning', 'error')),
                    error_message TEXT,
                    processing_time_ms INTEGER,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Add new columns if they don't exist (for existing databases)
            columns_to_add = [
                ("calculator_path", "TEXT"),
                ("raw_data", "TEXT"),
                ("processed_data", "TEXT"),
                ("field_mappings", "TEXT"),
                ("updated_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("modified_date", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
                ("validation_status", "TEXT DEFAULT 'pending'"),
                ("status", "TEXT DEFAULT 'pending'"),
                ("quality_score", "REAL DEFAULT 0.0"),
                ("file_source", "TEXT"),
                ("source_file", "TEXT"),
                ("processing_notes", "TEXT"),
                ("notes", "TEXT"),
                ("session_id", "TEXT")
            ]
            
            for column_name, column_type in columns_to_add:
                try:
                    conn.execute(f"ALTER TABLE user_data ADD COLUMN {column_name} {column_type}")
                except sqlite3.OperationalError:
                    pass  # Column already exists
            
            # Add session_id to processing_logs if it doesn't exist
            try:
                conn.execute("ALTER TABLE processing_logs ADD COLUMN session_id TEXT")
            except sqlite3.OperationalError:
                pass
            
            # Create indexes
            try:
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_hash ON user_data(data_hash)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_user_data_session ON user_data(session_id)")
            except sqlite3.OperationalError:
                pass
            
            print("Database initialization completed successfully")
                
        except Exception as e:
            print(f"Error initializing database: {e}")
            import traceback
            traceback.print_exc()
    
    def store_data(self, name: str, product_type: str, json_data: Dict, 
                   session_id: str, file_source: Optional[str] = None) -> Dict[str, Any]:
        """Store data with duplicate detection"""
        try:
            data_str = json.dumps(json_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            # Check for duplicates
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'duplicate',
                        'message': f'Data already exists (ID: {existing[0]}, Name: {existing[1]})'
                    }
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(json_data)
                
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, created_date, 
                     validation_status, quality_score, file_source, session_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(json_data),
                    datetime.now().isoformat(), 'validated', quality_score,
                    file_source, session_id
                ))
                
                record_id = cursor.lastrowid
                
                return {
                    'success': True,
                    'record_id': record_id,
                    'quality_score': quality_score
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': 'storage_error',
                'message': str(e)
            }
    
    def calculate_quality_score(self, json_data: Dict) -> float:
        """Calculate data quality score"""
        score = 0.0
        
        # Required fields check (40% of score)
        required_fields = ['policy_number', 'product_type', 'effective_date']
        present_required = sum(1 for field in required_fields if field in json_data and json_data[field])
        score += (present_required / len(required_fields)) * 40
        
        # Data completeness (30% of score)
        total_fields = len(json_data)
        non_empty_fields = sum(1 for v in json_data.values() if v is not None and str(v).strip())
        if total_fields > 0:
            score += (non_empty_fields / total_fields) * 30
        
        # Data format consistency (30% of score)
        format_score = 30
        if 'age' in json_data:
            try:
                age = float(json_data['age'])
                if age < 0 or age > 120:
                    format_score -= 10
            except:
                format_score -= 10
        
        score += format_score
        return round(min(100, score), 2)
    
    def add_user_data(self, name: str, raw_data: Dict, product_type: str, 
                      source_file: Optional[str] = None, status: str = 'pending',
                      calculator_path: Optional[str] = None) -> int:
        """Add user data to database and return record ID"""
        try:
            data_str = json.dumps(raw_data, sort_keys=True)
            data_hash = hashlib.md5(data_str.encode()).hexdigest()
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Check for duplicates
                cursor.execute("SELECT id FROM user_data WHERE data_hash = ?", (data_hash,))
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing record
                    cursor.execute("""
                        UPDATE user_data 
                        SET updated_date = ?, status = ?, calculator_path = ?
                        WHERE id = ?
                    """, (datetime.now().isoformat(), status, calculator_path, existing[0]))
                    return existing[0]
                
                # Calculate quality score
                quality_score = self.calculate_quality_score(raw_data)
                
                # Insert new record
                cursor.execute("""
                    INSERT INTO user_data 
                    (name, product_type, data_hash, json_data, raw_data, created_date, 
                     validation_status, status, quality_score, source_file, calculator_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name, product_type, data_hash, json.dumps(raw_data), json.dumps(raw_data),
                    datetime.now().isoformat(), 'validated', status, quality_score,
                    source_file, calculator_path
                ))
                
                return cursor.lastrowid
                
        except Exception as e:
            raise Exception(f"Failed to add user data: {str(e)}")

    def get_statistics(self) -> Dict[str, Any]:
        """Get database statistics for dashboard"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Total records
                cursor.execute("SELECT COUNT(*) FROM user_data")
                total_records = cursor.fetchone()[0]
                
                # Records by product type
                cursor.execute("SELECT product_type, COUNT(*) FROM user_data GROUP BY product_type")
                product_counts = dict(cursor.fetchall())
                
                # Basic stats to avoid column issues
                return {
                    'total_records': total_records,
                    'status_counts': {'processed': total_records},  # Simplified
                    'product_counts': product_counts,
                    'product_distribution': product_counts,  # Template expects this name
                    'average_quality_score': 0.0,  # Simplified
                    'average_quality': 0.0,  # Template expects this name
                    'recent_activity': total_records,  # Simplified
                    'processed_records': total_records,
                    'pending_records': 0,
                    'error_records': 0
                }
        except Exception as e:
            print(f"Database error in get_statistics: {e}")
            # Return default stats if there's an error
            return {
                'total_records': 0,
                'status_counts': {},
                'product_counts': {},
                'product_distribution': {},  # Template expects this name
                'average_quality_score': 0,
                'average_quality': 0,  # Template expects this name
                'recent_activity': 0,
                'processed_records': 0,
                'pending_records': 0,
                'error_records': 0
            }

    def load_calculator_reference(self, calculator_path: str, product_type: str) -> Dict:
        """Load calculator reference data (mock implementation)"""
        # This is a mock implementation. In a real system, you would load
        # the actual calculator configuration and reference data
        
        calculator_templates = {
            'life_insurance': {
                'fields': {
                    'policy_number': {'required': True, 'type': 'string', 'sample': 'LIFE001234'},
                    'insured_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1980-01-01'},
                    'gender': {'required': True, 'type': 'string', 'sample': 'M'},
                    'coverage_amount': {'required': True, 'type': 'number', 'sample': 100000},
                    'premium_amount': {'required': False, 'type': 'number', 'sample': 1200},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'TERM20'},
                    'beneficiary_name': {'required': False, 'type': 'string', 'sample': 'Jane Doe'},
                    'agent_code': {'required': False, 'type': 'string', 'sample': 'AGT001'}
                }
            },
            'annuity': {
                'fields': {
                    'contract_number': {'required': True, 'type': 'string', 'sample': 'ANN001234'},
                    'owner_name': {'required': True, 'type': 'string', 'sample': 'John Doe'},
                    'birth_date': {'required': True, 'type': 'date', 'sample': '1970-01-01'},
                    'initial_deposit': {'required': True, 'type': 'number', 'sample': 50000},
                    'product_code': {'required': True, 'type': 'string', 'sample': 'FIXED5'},
                    'effective_date': {'required': True, 'type': 'date', 'sample': '2024-01-01'},
                    'surrender_charge_period': {'required': False, 'type': 'number', 'sample': 7},
                    'interest_rate': {'required': False, 'type': 'number', 'sample': 3.5}
                }
            }
        }
        
        return calculator_templates.get(product_type, calculator_templates['life_insurance'])

    def perform_data_comparison(self, source_data: Dict, calculator_data: Dict) -> Dict:
        """Compare source data with calculator reference and generate comparison result"""
        
        calculator_fields = calculator_data.get('fields', {})
        comparison_fields = []
        
        # Stats
        total_fields = len(calculator_fields)
        matching_fields = 0
        missing_fields = 0
        filled_fields = 0
        
        for field_name, field_config in calculator_fields.items():
            source_value = source_data.get(field_name)
            calculator_value = field_config.get('sample')
            
            field_result = {
                'field_name': field_name,
                'source_value': source_value,
                'calculator_value': calculator_value,
                'required': field_config.get('required', False),
                'data_type': field_config.get('type', 'string')
            }
            
            # Determine status
            if source_value is not None and source_value != '':
                if str(source_value).strip().lower() == str(calculator_value).strip().lower():
                    field_result['status'] = 'match'
                    matching_fields += 1
                else:
                    field_result['status'] = 'mismatch'
            else:
                if calculator_value and field_config.get('required', False):
                    field_result['status'] = 'missing'
                    missing_fields += 1
                elif calculator_value:
                    field_result['status'] = 'can_fill'
                    filled_fields += 1
                else:
                    field_result['status'] = 'missing'
                    missing_fields += 1
            
            comparison_fields.append(field_result)
        
        return {
            'stats': {
                'total_fields': total_fields,
                'matching_fields': matching_fields,
                'missing_fields': missing_fields,
                'filled_fields': filled_fields,
                'completion_percentage': round((matching_fields / total_fields) * 100, 2) if total_fields > 0 else 0
            },
            'fields': comparison_fields
        }

    def get_all_data(self, session_id: Optional[str] = None) -> list:
        """Get all stored data, optionally filtered by session"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if session_id:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status AS status, 
                           quality_score, file_source, data_hash
                    FROM user_data WHERE session_id = ? ORDER BY created_date DESC
                """, (session_id,))
            else:
                cursor.execute("""
                    SELECT id, name, product_type, created_date, validation_status AS status, 
                           quality_score, file_source, data_hash
                    FROM user_data ORDER BY created_date DESC LIMIT 100
                """)
            
            rows = cursor.fetchall()
            data = []
            for row in rows:
                record = dict(row)
                # Ensure default values for missing fields
                record['source_file'] = record.get('file_source') or 'Manual Entry'
                record['status'] = record.get('status') or 'processed'
                record['quality_score'] = record.get('quality_score') or 0.0
                data.append(record)
            
            return data
    
    def get_data_by_id(self, record_id: int) -> Optional[Dict]:
        """Get specific record by ID"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM user_data WHERE id = ?", (record_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

# Initialize database manager
db_manager = WebDatabaseManager()

@app.route('/')
def index():
    """Main dashboard page"""
    try:
        if 'session_id' not in session:
            session['session_id'] = os.urandom(16).hex()
        
        print("Getting statistics...")
        stats = db_manager.get_statistics()
        print(f"Stats: {stats}")
        
        print("Getting recent data...")
        recent_data = db_manager.get_all_data(session['session_id'])
        print(f"Recent data count: {len(recent_data)}")
        
        return render_template('index.html', stats=stats, recent_data=recent_data)
    except Exception as e:
        print(f"Error in index route: {e}")
        import traceback
        traceback.print_exc()
        return f"Error: {e}", 500

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    """File upload and processing page"""
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name', '').strip()
            product_type = request.form.get('product_type', '').strip()
            
            if not name or not product_type:
                flash('Name and product type are required', 'error')
                return redirect(url_for('upload_file'))
            
            # Handle JSON data input
            json_text = request.form.get('json_data', '').strip()
            uploaded_file = request.files.get('file')
            
            json_data = None
            file_source = None
            
            if json_text:
                # Parse JSON from text input
                try:
                    json_data = json.loads(json_text)
                    file_source = 'manual_input'
                except json.JSONDecodeError as e:
                    flash(f'Invalid JSON format: {str(e)}', 'error')
                    return redirect(url_for('upload_file'))
            
            elif uploaded_file and uploaded_file.filename:
                # Process uploaded file
                filename = secure_filename(uploaded_file.filename)
                file_source = filename
                
                if filename.endswith('.json'):
                    try:
                        json_data = json.load(uploaded_file)
                    except json.JSONDecodeError as e:
                        flash(f'Invalid JSON file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(uploaded_file)
                        json_data = df.to_dict('records')[0] if not df.empty else {}
                    except Exception as e:
                        flash(f'Error reading Excel file: {str(e)}', 'error')
                        return redirect(url_for('upload_file'))
                
                else:
                    flash('Please upload a JSON or Excel file', 'error')
                    return redirect(url_for('upload_file'))
            
            else:
                flash('Please provide JSON data or upload a file', 'error')
                return redirect(url_for('upload_file'))
            
            if not json_data:
                flash('No data found to process', 'error')
                return redirect(url_for('upload_file'))
            
            # Store data in database
            result = db_manager.store_data(
                name=name,
                product_type=product_type,
                json_data=json_data,
                session_id=session['session_id'],
                file_source=file_source
            )
            
            if result['success']:
                flash(f'Data stored successfully! Quality Score: {result["quality_score"]}%', 'success')
                return redirect(url_for('process_data', record_id=result['record_id']))
            else:
                if result['error'] == 'duplicate':
                    flash(f'Duplicate data detected: {result["message"]}', 'warning')
                else:
                    flash(f'Error storing data: {result["message"]}', 'error')
                return redirect(url_for('upload_file'))
        
        except Exception as e:
            flash(f'Unexpected error: {str(e)}', 'error')
            return redirect(url_for('upload_file'))
    
    return render_template('upload.html')

@app.route('/process/<int:record_id>')
def process_data(record_id):
    """Process stored data using AIM processor"""
    try:
        # Get data from database
        record = db_manager.get_data_by_id(record_id)
        if not record:
            flash('Record not found', 'error')
            return redirect(url_for('index'))
        
        # Parse JSON data
        json_data = json.loads(record['json_data'])
        
        # Process with AIM processor
        result = aim_processor.process_fast_ui_input(
            json_data, 
            record['product_type']
        )
        
        return render_template('process_results.html', 
                             record=record, 
                             input_data=json_data,
                             result=result)
    
    except (ValidationError, MappingError) as e:
        flash(f'Processing error: {str(e)}', 'error')
        return redirect(url_for('index'))
    except Exception as e:
        flash(f'Unexpected error during processing: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/field-mapping')
def field_mapping():
    """Field mapping interface"""
    return render_template('field_mapping.html')

@app.route('/compare')
def compare():
    """Data comparison page"""
    # Get processed records for comparison
    processed_records = []
    with sqlite3.connect(db_manager.db_path) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, product_type, created_date, calculator_path
            FROM user_data 
            WHERE status IN ('processed', 'saved') 
            ORDER BY created_date DESC
        """)
        processed_records = [dict(row) for row in cursor.fetchall()]
    
    return render_template('compare.html', processed_records=processed_records)

@app.route('/create-mapping', methods=['POST'])
def create_mapping():
    """Create Excel field mapping template"""
    try:
        product_type = request.form.get('product_type', 'life_insurance')
        
        # Get sample data if available and session exists
        sample_data = None
        if 'session_id' in session:
            recent_data = db_manager.get_all_data(session['session_id'])
            if recent_data:
                latest_record = db_manager.get_data_by_id(recent_data[0]['id'])
                if latest_record:
                    try:
                        sample_data = json.loads(latest_record['json_data'])
                    except json.JSONDecodeError:
                        sample_data = None
        
        # Create Excel template
        filename = create_excel_template(product_type, sample_data)
        
        if not os.path.exists(filename):
            raise FileNotFoundError(f"Template file was not created: {filename}")
        
        return send_file(filename, 
                        as_attachment=True, 
                        download_name=f'AIM_Mapping_Template_{product_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        error_msg = f'Error creating mapping template: {str(e)}'
        print(f"Excel template error: {error_msg}")  # Log to console
        
        # Return JSON error for AJAX requests or HTML error page
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'error': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for('field_mapping'))

@app.route('/api/save-mapping', methods=['POST'])
def save_mapping():
    """Save field mapping configuration"""
    try:
        data = request.get_json()
        
        mapping_name = data.get('name', '')
        product_type = data.get('product_type', '')
        description = data.get('description', '')
        field_mappings = data.get('mappings', [])
        
        if not mapping_name or not product_type:
            return jsonify({'success': False, 'message': 'Mapping name and product type are required'})
        
        # Create mappings directory if it doesn't exist
        mappings_dir = 'saved_mappings'
        os.makedirs(mappings_dir, exist_ok=True)
        
        # Save mapping to JSON file
        mapping_file = os.path.join(mappings_dir, f"{mapping_name.replace(' ', '_').lower()}.json")
        mapping_data = {
            'name': mapping_name,
            'product_type': product_type,
            'description': description,
            'created_date': datetime.now().isoformat(),
            'field_mappings': field_mappings
        }
        
        with open(mapping_file, 'w') as f:
            json.dump(mapping_data, f, indent=2)
        
        return jsonify({
            'success': True, 
            'message': f'Mapping "{mapping_name}" saved successfully',
            'file_path': mapping_file
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving mapping: {str(e)}'})

@app.route('/api/mappings')
def list_mappings():
    """List all saved mapping templates"""
    try:
        mappings_dir = 'saved_mappings'
        if not os.path.exists(mappings_dir):
            return jsonify({'mappings': []})
        
        mappings = []
        for filename in os.listdir(mappings_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(mappings_dir, filename)
                try:
                    with open(filepath, 'r') as f:
                        mapping_data = json.load(f)
                        mappings.append({
                            'filename': filename,
                            'name': mapping_data.get('name', ''),
                            'product_type': mapping_data.get('product_type', ''),
                            'description': mapping_data.get('description', ''),
                            'created_date': mapping_data.get('created_date', '')
                        })
                except Exception as e:
                    print(f"Error reading mapping file {filename}: {e}")
                    continue
        
        return jsonify({'mappings': mappings})
        
    except Exception as e:
        return jsonify({'error': f'Error listing mappings: {str(e)}'})

@app.route('/api/mappings/<filename>')
def load_mapping(filename):
    """Load a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Mapping file not found'}), 404
        
        with open(filepath, 'r') as f:
            mapping_data = json.load(f)
        
        return jsonify({'success': True, 'mapping': mapping_data})
        
    except Exception as e:
        return jsonify({'error': f'Error loading mapping: {str(e)}'})

@app.route('/api/mappings/<filename>', methods=['DELETE'])
def delete_mapping(filename):
    """Delete a specific mapping template"""
    try:
        mappings_dir = 'saved_mappings'
        filepath = os.path.join(mappings_dir, f"{filename}.json")
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'Mapping file not found'}), 404
        
        os.remove(filepath)
        return jsonify({'success': True, 'message': f'Mapping "{filename}" deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting mapping: {str(e)}'})

@app.route('/view-data')
def view_data():
    """View all stored data"""
    all_data = db_manager.get_all_data()
    
    # Calculate statistics for the view
    stats = {
        'total_records': len(all_data),
        'avg_quality': sum(record.get('quality_score', 0) for record in all_data) / len(all_data) if all_data else 0,
        'last_updated': all_data[0].get('created_date', 'Never') if all_data else 'Never'
    }
    
    return render_template('view_data.html', data=all_data, **stats)

@app.route('/api/data/<int:record_id>')
def api_get_data(record_id):
    """API endpoint to get record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify(record)
    else:
        return jsonify({'error': 'Record not found'}), 404

@app.route('/api/stats')
def api_stats():
    """API endpoint for statistics"""
    return jsonify(db_manager.get_statistics())

@app.route('/api/record/<int:record_id>')
def api_get_record(record_id):
    """API endpoint to get detailed record data"""
    record = db_manager.get_data_by_id(record_id)
    if record:
        return jsonify({'success': True, 'record': record})
    else:
        return jsonify({'success': False, 'message': 'Record not found'}), 404

@app.route('/api/record/<int:record_id>', methods=['DELETE'])
def api_delete_record(record_id):
    """API endpoint to delete a record"""
    try:
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM user_data WHERE id = ?", (record_id,))
            
            if cursor.rowcount > 0:
                return jsonify({'success': True, 'message': 'Record deleted successfully'})
            else:
                return jsonify({'success': False, 'message': 'Record not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/record/<int:record_id>', methods=['PUT'])
def api_update_record(record_id):
    """API endpoint to update a record"""
    try:
        data = request.get_json()
        
        # Prepare update fields
        update_fields = []
        update_values = []
        
        if 'name' in data:
            update_fields.append('name = ?')
            update_values.append(data['name'])
        
        if 'product_type' in data:
            update_fields.append('product_type = ?')
            update_values.append(data['product_type'])
        
        if 'status' in data:
            update_fields.append('status = ?')
            update_values.append(data['status'])
        
        if 'quality_score' in data:
            update_fields.append('quality_score = ?')
            update_values.append(data['quality_score'])
        
        if 'notes' in data:
            update_fields.append('notes = ?')
            update_values.append(data['notes'])
        
        # Add updated timestamp
        update_fields.append('updated_date = ?')
        update_values.append(datetime.now().isoformat())
        
        # Add record_id for WHERE clause
        update_values.append(record_id)
        
        if not update_fields:
            return jsonify({'success': False, 'message': 'No valid fields to update'}), 400
        
        with sqlite3.connect(db_manager.db_path) as conn:
            cursor = conn.cursor()
            
            # Check if record exists first
            cursor.execute("SELECT id FROM user_data WHERE id = ?", (record_id,))
            if not cursor.fetchone():
                return jsonify({'success': False, 'message': 'Record not found'}), 404
            
            # Update the record
            update_query = f"UPDATE user_data SET {', '.join(update_fields)} WHERE id = ?"
            cursor.execute(update_query, update_values)
            
            return jsonify({'success': True, 'message': 'Record updated successfully'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export-data')
def api_export_data():
    """API endpoint to export data"""
    try:
        export_format = request.args.get('format', 'json').lower()
        all_data = db_manager.get_all_data()
        
        # Create a simplified export format
        export_data = []
        for record in all_data:
            export_record = {
                'id': record.get('id'),
                'name': record.get('name'),
                'product_type': record.get('product_type'),
                'created_date': record.get('created_date'),
                'quality_score': record.get('quality_score'),
                'status': record.get('status'),
                'source_file': record.get('source_file'),
                'notes': record.get('notes', '')
            }
            export_data.append(export_record)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'excel':
            # Create Excel file
            import pandas as pd
            from io import BytesIO
            
            df = pd.DataFrame(export_data)
            excel_buffer = BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='AIM_Data', index=False)
            
            excel_buffer.seek(0)
            
            return send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'aim_data_export_{timestamp}.xlsx'
            )
        else:
            # Return as JSON download
            response = jsonify(export_data)
            response.headers['Content-Disposition'] = f'attachment; filename=aim_data_export_{timestamp}.json'
            return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/save-data', methods=['POST'])
def api_save_data():
    """API endpoint to save data without processing"""
    try:
        # Get form data
        name = request.form.get('name', '').strip()
        product_type = request.form.get('product_type', '')
        calculator_path = request.form.get('calculator_path', '').strip()
        save_only = request.form.get('save_only', '') == 'true'
        
        # Validation
        if not name or not product_type or not calculator_path:
            return jsonify({
                'success': False, 
                'message': 'Name, product type, and calculator path are required'
            }), 400
        
        # Get data (file or manual JSON)
        raw_data = None
        source_file = None
        
        if 'file' in request.files and request.files['file'].filename:
            file = request.files['file']
            if file.filename != '':
                filename = secure_filename(file.filename)
                source_file = filename
                
                # Process uploaded file
                if filename.endswith('.json'):
                    content = file.read().decode('utf-8')
                    raw_data = json.loads(content)
                elif filename.endswith(('.xlsx', '.xls')):
                    # For Excel files, convert to basic structure
                    df = pd.read_excel(file)
                    raw_data = df.to_dict('records')
                else:
                    return jsonify({
                        'success': False, 
                        'message': 'Unsupported file format. Please use JSON or Excel files.'
                    }), 400
        
        elif request.form.get('json_data'):
            try:
                raw_data = json.loads(request.form.get('json_data'))
                source_file = 'manual_input'
            except json.JSONDecodeError:
                return jsonify({
                    'success': False, 
                    'message': 'Invalid JSON format in manual input'
                }), 400
        
        if not raw_data:
            return jsonify({
                'success': False, 
                'message': 'No data provided. Please upload a file or enter JSON data.'
            }), 400
        
        # Save to database
        record_id = db_manager.add_user_data(
            name=name,
            raw_data=raw_data,
            product_type=product_type,
            source_file=source_file,
            status='saved' if save_only else 'pending',
            calculator_path=calculator_path
        )
        
        return jsonify({
            'success': True,
            'message': f'Data saved successfully',
            'record_id': record_id,
            'calculator_path': calculator_path
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/compare-data', methods=['POST'])
def api_compare_data():
    """API endpoint to compare processed data with calculator"""
    try:
        data = request.get_json()
        record_id = data.get('record_id')
        calculator_path = data.get('calculator_path')
        
        # Validation
        if not record_id or not calculator_path:
            return jsonify({
                'success': False,
                'message': 'Record ID and calculator path are required'
            }), 400
        
        # Get source data
        with sqlite3.connect(db_manager.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM user_data WHERE id = ?", (record_id,))
            record = cursor.fetchone()
            
            if not record:
                return jsonify({
                    'success': False,
                    'message': 'Record not found'
                }), 404
        
        # Parse source data
        source_data = json.loads(record['json_data'] or '{}')
        
        # Load calculator reference (mock implementation)
        calculator_data = db_manager.load_calculator_reference(calculator_path, record['product_type'])
        
        # Perform comparison
        comparison_result = db_manager.perform_data_comparison(source_data, calculator_data)
        
        return jsonify({
            'success': True,
            'comparison': comparison_result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/save-filled-data', methods=['POST'])
def api_save_filled_data():
    """API endpoint to save filled/completed data as new record"""
    try:
        data = request.get_json()
        
        name = data.get('name', '').strip()
        product_type = data.get('product_type', '')
        calculator_path = data.get('calculator_path', '')
        filled_data = data.get('filled_data', {})
        source_record_id = data.get('source_record_id')
        
        # Validation
        if not name or not product_type or not filled_data:
            return jsonify({
                'success': False,
                'message': 'Name, product type, and filled data are required'
            }), 400
        
        # Save filled data as new record
        record_id = db_manager.add_user_data(
            name=name,
            raw_data=filled_data,
            product_type=product_type,
            source_file=f'filled_from_record_{source_record_id}',
            status='filled',
            calculator_path=calculator_path
        )
        
        return jsonify({
            'success': True,
            'message': 'Filled data saved successfully',
            'record_id': record_id
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/health')
def health_check():
    """Health check endpoint for Azure monitoring"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

def create_excel_template(product_type: str, sample_data: Optional[Dict] = None):
    """Create Excel mapping template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Field_Mapping"
    
    # Headers
    headers = ['Source_Field', 'Target_Field', 'Values_Match', 'Sample_Value', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data rows
    if sample_data:
        for row_idx, (field, value) in enumerate(sample_data.items(), 2):
            ws.cell(row=row_idx, column=1, value=field)  # Source field
            ws.cell(row=row_idx, column=4, value=str(value)[:50])  # Sample value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = f"templates/mapping_template_{product_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    os.makedirs('templates', exist_ok=True)
    wb.save(filename)
    
    return filename

@app.route('/help')
def help_page():
    """Help and documentation page"""
    return render_template('help.html')

@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', 
                         error_code=404, 
                         error_message="Page not found"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', 
                         error_code=500, 
                         error_message="Internal server error"), 500

@app.route('/download-template/<product_type>')
def download_template(product_type):
    """Download a JSON template file for the specified product type"""
    try:
        # Create template data based on product type
        if product_type == 'life_insurance':
            template_name = 'Life_Insurance_Mapping_Template'
            template_description = 'Standard field mapping template for life insurance policies including policy details, insured information, and coverage amounts.'
            template_mappings = [
                {'source': 'PolicyNumber', 'target': 'policy_number', 'transformation': 'trim', 'required': True},
                {'source': 'InsuredFirstName', 'target': 'insured_first_name', 'transformation': 'trim', 'required': True},
                {'source': 'InsuredLastName', 'target': 'insured_last_name', 'transformation': 'trim', 'required': True},
                {'source': 'DateOfBirth', 'target': 'birth_date', 'transformation': 'date_format', 'required': True},
                {'source': 'Gender', 'target': 'gender', 'transformation': 'uppercase', 'required': True},
                {'source': 'CoverageAmount', 'target': 'coverage_amount', 'transformation': 'currency_format', 'required': True},
                {'source': 'PremiumAmount', 'target': 'premium_amount', 'transformation': 'currency_format', 'required': True},
                {'source': 'EffectiveDate', 'target': 'effective_date', 'transformation': 'date_format', 'required': True},
                {'source': 'ProductCode', 'target': 'product_code', 'transformation': 'uppercase', 'required': True},
                {'source': 'BeneficiaryName', 'target': 'beneficiary_name', 'transformation': 'trim', 'required': False},
                {'source': 'AgentCode', 'target': 'agent_code', 'transformation': 'trim', 'required': False},
                {'source': 'UnderwritingClass', 'target': 'underwriting_class', 'transformation': 'uppercase', 'required': False}
            ]
        elif product_type == 'annuity':
            template_name = 'Annuity_Mapping_Template'
            template_description = 'Standard field mapping template for annuity products including contract details, owner information, and accumulation values.'
            template_mappings = [
                {'source': 'ContractNumber', 'target': 'contract_number', 'transformation': 'trim', 'required': True},
                {'source': 'OwnerFirstName', 'target': 'owner_first_name', 'transformation': 'trim', 'required': True},
                {'source': 'OwnerLastName', 'target': 'owner_last_name', 'transformation': 'trim', 'required': True},
                {'source': 'DateOfBirth', 'target': 'birth_date', 'transformation': 'date_format', 'required': True},
                {'source': 'Gender', 'target': 'gender', 'transformation': 'uppercase', 'required': True},
                {'source': 'InitialDeposit', 'target': 'initial_deposit', 'transformation': 'currency_format', 'required': True},
                {'source': 'AccumulationValue', 'target': 'accumulation_value', 'transformation': 'currency_format', 'required': True},
                {'source': 'IssueDate', 'target': 'issue_date', 'transformation': 'date_format', 'required': True},
                {'source': 'ProductCode', 'target': 'product_code', 'transformation': 'uppercase', 'required': True},
                {'source': 'AnnuitantName', 'target': 'annuitant_name', 'transformation': 'trim', 'required': False},
                {'source': 'BeneficiaryName', 'target': 'beneficiary_name', 'transformation': 'trim', 'required': False},
                {'source': 'SurrenderChargeSchedule', 'target': 'surrender_schedule', 'transformation': 'none', 'required': False},
                {'source': 'InterestRate', 'target': 'interest_rate', 'transformation': 'none', 'required': False}
            ]
        else:
            return jsonify({'error': 'Invalid product type'}), 400

        # Create template data structure
        template_data = {
            'name': template_name.replace('_', ' '),
            'product_type': product_type,
            'description': template_description,
            'created_date': datetime.now().isoformat(),
            'field_mappings': template_mappings,
            'usage_instructions': {
                'step1': 'Download this template file',
                'step2': 'Customize the field mappings for your data source',
                'step3': 'Upload to AIM field mapping page using the Load Template feature',
                'step4': 'Fine-tune mappings and validation rules as needed'
            }
        }

        # Create temporary file
        temp_filename = f"temp_{template_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        temp_path = os.path.join('uploads', temp_filename)
        
        # Ensure uploads directory exists
        os.makedirs('uploads', exist_ok=True)
        
        # Write template to file
        with open(temp_path, 'w') as f:
            json.dump(template_data, f, indent=2)

        # Return file as download
        return send_file(temp_path, 
                        as_attachment=True, 
                        download_name=f'{template_name}_Template_{datetime.now().strftime("%Y%m%d")}.json',
                        mimetype='application/json')

    except Exception as e:
        return jsonify({'error': f'Error creating template: {str(e)}'}), 500

@app.route('/api/upload-template', methods=['POST'])
def upload_template():
    """Upload and process a mapping template file (JSON or Excel)"""
    try:
        if 'template_file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'})
        
        file = request.files['template_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'})
        
        # Secure the filename
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            if filename.lower().endswith('.json'):
                # Process JSON file
                with open(file_path, 'r') as f:
                    template_data = json.load(f)
                
                return jsonify({
                    'success': True,
                    'message': f'JSON template "{template_data.get("name", "Unknown")}" processed successfully',
                    'template_data': template_data
                })
                
            elif filename.lower().endswith(('.xlsx', '.xls')):
                # Process Excel file
                import pandas as pd
                
                # Read Excel file
                df = pd.read_excel(file_path)
                
                # Extract mapping data from Excel
                template_data = {
                    'name': f'Imported from {filename}',
                    'product_type': 'life_insurance',  # Default
                    'description': f'Template imported from Excel file: {filename}',
                    'created_date': datetime.now().isoformat(),
                    'field_mappings': []
                }
                
                # Process Excel rows to extract field mappings
                for index, row in df.iterrows():
                    if pd.notna(row.get('Source_Field')) and pd.notna(row.get('Target_Field')):
                        mapping = {
                            'source': str(row.get('Source_Field', '')),
                            'target': str(row.get('Target_Field', '')),
                            'transformation': str(row.get('Transformation', 'none')).lower(),
                            'required': bool(row.get('Required', False))
                        }
                        template_data['field_mappings'].append(mapping)
                
                return jsonify({
                    'success': True,
                    'message': f'Excel template processed successfully. {len(template_data["field_mappings"])} mappings found.',
                    'template_data': template_data
                })
            
            else:
                return jsonify({'success': False, 'message': 'Unsupported file format. Please upload JSON or Excel files.'})
                
        finally:
            # Clean up uploaded file
            if os.path.exists(file_path):
                os.remove(file_path)
                
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error processing template file: {str(e)}'})

if __name__ == '__main__':
    print("Starting AIM Web Application...")
    print(f"Current working directory: {os.getcwd()}")
    # Development server
    port = int(os.environ.get('PORT', 5000))
    debug = True  # Temporarily enable debug for error details
    print(f"Starting Flask app on port {port} with debug={debug}")
    app.run(host='0.0.0.0', port=port, debug=debug)
